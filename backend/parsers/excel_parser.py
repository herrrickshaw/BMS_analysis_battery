"""
Extract stock tickers from uploaded Excel files for any supported market.
"""
from __future__ import annotations

import io
import re
from typing import Optional

import openpyxl

from parsers.market_db import _YF_SUFFIX, extract_from_text, lookup, lookup_by_name

_TICKER_HEADERS = {
    'symbol', 'ticker', 'nse code', 'nse symbol', 'bse code', 'scrip',
    'scrip code', 'trading symbol', 'stock symbol', 'nse ticker',
    'instrument', 'stock', 'script', 'equity', 'code', 'yf_ticker',
    'yf ticker', 'stock code', 'security code',
}
_NAME_HEADERS = {
    'name', 'company', 'company name', 'stock name', 'scrip name',
    'instrument name', 'security name', 'issuer', 'issuer name',
    'security', 'description',
}
_ISIN_HEADERS = {'isin', 'isin number', 'isin code', 'isin no'}
_DATE_HEADERS = {
    'date', 'purchase date', 'buy date', 'transaction date', 'trade date',
    'acquisition date', 'date of purchase', 'date of acquisition',
    'invested on', 'purchase on',
}
_QTY_HEADERS = {
    'quantity', 'qty', 'shares', 'units', 'no of shares', 'number of shares',
    'no. of shares', 'holdings', 'nos', 'no', 'shares held',
}
_PRICE_HEADERS = {
    'purchase price', 'buy price', 'cost price', 'avg price', 'average price',
    'avg cost', 'average cost', 'buy rate', 'cost', 'price', 'invested price',
    'acquisition price', 'avg buy price',
}


def _col_role(header: str) -> Optional[str]:
    h = header.lower().strip().rstrip('.')
    if h in _TICKER_HEADERS:
        return 'ticker'
    if h in _NAME_HEADERS:
        return 'name'
    if h in _ISIN_HEADERS:
        return 'isin'
    if h in _DATE_HEADERS:
        return 'date'
    if h in _QTY_HEADERS:
        return 'qty'
    if h in _PRICE_HEADERS:
        return 'price'
    return None


def _find_header_row(ws, max_scan: int = 20):
    for r in range(1, min(max_scan + 1, ws.max_row + 1)):
        role_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            role = _col_role(str(val))
            if role and role not in role_map:
                role_map[role] = c
        if role_map:
            return r, role_map
    return None


def _parse_float(raw: str) -> Optional[float]:
    """Convert a cell value to float, stripping currency symbols and commas."""
    try:
        return float(re.sub(r'[₹$£€,\s]', '', str(raw)))
    except (ValueError, TypeError):
        return None


def _parse_date_cell(raw) -> Optional[str]:
    """Convert a cell value (datetime / date / string) to ISO date string."""
    import datetime as _dt
    if isinstance(raw, (_dt.datetime, _dt.date)):
        d = raw.date() if isinstance(raw, _dt.datetime) else raw
        return d.isoformat()
    if raw is None or str(raw).strip() in ('', 'None', 'nan'):
        return None
    from fetchers.history import parse_date
    d = parse_date(str(raw).strip())
    return d.isoformat() if d else None


def _resolve(raw_ticker: str, raw_name: str, raw_isin: str,
             market: str, sheet: str,
             raw_date: str = '', raw_qty: str = '', raw_price: str = '') -> Optional[dict]:
    """Resolve a row to a yfinance-ready symbol, including optional holding data."""
    info = None
    matched_via = ''

    if raw_ticker:
        # Strip exchange suffixes like -EQ, /NSE, spaces
        sym = re.sub(r'[ \t/-].*$', '', raw_ticker.upper().strip())
        info = lookup(sym, market)
        if info:
            matched_via = 'symbol'

    if not info and raw_isin and len(raw_isin) >= 12:
        if market == 'india' and raw_isin.startswith('IN'):
            from parsers.symbol_db import lookup_by_isin
            nse_info = lookup_by_isin(raw_isin)
            if nse_info:
                info = {'yf_ticker': nse_info['symbol'] + '.NS',
                        'name': nse_info['name'], 'isin': raw_isin}
                matched_via = 'ISIN'

    if not info and raw_name:
        info = lookup_by_name(raw_name, market)
        if info:
            matched_via = 'name (fuzzy)'

    if not info:
        return None

    rec: dict = {
        'yf_ticker':   info['yf_ticker'],
        'symbol':      info['yf_ticker'],
        'name':        info.get('name') or raw_name,
        'isin':        info.get('isin') or raw_isin,
        'sheet':       sheet,
        'matched_via': matched_via,
    }

    # Optional holding fields — included only when present
    if raw_date:
        parsed_date = _parse_date_cell(raw_date)
        if parsed_date:
            rec['purchase_date'] = parsed_date
    if raw_qty:
        qty = _parse_float(raw_qty)
        if qty is not None:
            rec['quantity'] = qty
    if raw_price:
        price = _parse_float(raw_price)
        if price is not None:
            rec['purchase_price'] = price

    return rec


def parse_excel(file_bytes: bytes, market: str = 'india') -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    stocks: dict[str, dict] = {}
    warnings: list[str] = []
    sheets_scanned: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result = _find_header_row(ws)

        if result is None:
            # No recognisable header — scan full sheet as raw text
            text_parts = []
            for row in ws.iter_rows(max_row=500, values_only=True):
                for cell in row:
                    if cell:
                        text_parts.append(str(cell))
            found = extract_from_text(' '.join(text_parts), market)
            if found:
                sheets_scanned.append(sheet_name)
                for s in found:
                    yf = s['yf_ticker']
                    if yf not in stocks:
                        stocks[yf] = {**s, 'symbol': yf, 'sheet': sheet_name}
            continue

        header_row, role_map = result
        sheets_scanned.append(sheet_name)
        ticker_col = role_map.get('ticker')
        name_col   = role_map.get('name')
        isin_col   = role_map.get('isin')
        date_col   = role_map.get('date')
        qty_col    = role_map.get('qty')
        price_col  = role_map.get('price')

        unresolved: list[str] = []

        for row in ws.iter_rows(min_row=header_row + 1, max_row=2000, values_only=True):
            def cell(col):
                return row[col - 1] if col and row[col - 1] is not None else ''

            def cell_str(col):
                v = cell(col)
                return str(v).strip() if v != '' else ''

            raw_ticker = cell_str(ticker_col)
            raw_name   = cell_str(name_col)
            raw_isin   = cell_str(isin_col)
            raw_date   = cell(date_col)       # keep raw for date parsing (may be datetime obj)
            raw_qty    = cell_str(qty_col)
            raw_price  = cell_str(price_col)

            if not any([raw_ticker, raw_name, raw_isin]):
                continue
            if raw_ticker in ('None', 'nan', '-'):
                raw_ticker = ''

            rec = _resolve(raw_ticker, raw_name, raw_isin, market, sheet_name,
                           raw_date=raw_date, raw_qty=raw_qty, raw_price=raw_price)
            if rec:
                yf = rec['yf_ticker']
                if yf not in stocks:
                    stocks[yf] = rec
            elif raw_name and raw_name not in ('None', 'nan'):
                unresolved.append(raw_name)

        if unresolved:
            sample = unresolved[:5]
            warnings.append(
                f"Sheet '{sheet_name}': {len(unresolved)} rows could not be matched "
                f"(e.g. {', '.join(sample)})"
            )

    return {
        'stocks':        list(stocks.values()),
        'warnings':      warnings,
        'sheets_scanned': sheets_scanned,
    }
