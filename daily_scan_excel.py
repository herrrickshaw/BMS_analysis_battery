#!/usr/bin/env python3
"""
Daily NSE/BSE Pegu + Sarvas Scan → Excel Export
=================================================
One-shot script:
  1. Re-runs nse_bse_extractor (fetches live data if network allows,
     falls back to realistic NIFTY500 representative data otherwise)
  2. Applies Pegu scoring + Sarvas scan in pure Python (mirrors
     pegu_sarvas_analysis.R logic exactly)
  3. Exports a formatted multi-sheet Excel workbook

Output: reports/NSE_BSE_Daily_Scan_YYYYMMDD.xlsx

Usage:
  python daily_scan_excel.py
  python daily_scan_excel.py --index NIFTY50 --top-n 30
  python daily_scan_excel.py --mock          # force sample data (no network)
"""

import argparse
import math
import os
import subprocess
import sys
import warnings
from datetime import datetime
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

warnings.filterwarnings("ignore")

TODAY = datetime.now().strftime("%Y-%m-%d")
TODAY_FMT = datetime.now().strftime("%d %b %Y")
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

# ─────────────────────────────────────────────────────────────
# REPRESENTATIVE NIFTY500 SAMPLE DATA
# Used when live fetch is unavailable (network-blocked sandbox)
# Values approximate real fundamentals as of mid-2025
# ─────────────────────────────────────────────────────────────
_SAMPLE_STOCKS = [
    # sym, exch, name, sector, industry, price, pe, pb, peg, roe, op_margin, d_e, curr_ratio, eps_trail, eps_fwd, rev_growth, earn_growth, div_yield, beta, market_cap, w52h, w52l, ma50, ma200, rsi, macd, vol_ratio, ret_1m, ret_3m, ret_6m, target
    ("RELIANCE","NSE","Reliance Industries","Energy","Integrated Oil & Gas",2950,21.3,2.14,0.91,18.3,7.2,44.1,1.18,138.4,164.7,7.4,12.7,3.35,0.93,19834.5e8,3217,2220,2871,2732,58.4,12.3,1.2,3.1,8.4,14.2,3380),
    ("TCS","NSE","Tata Consultancy Services","Technology","IT Services",4120,28.4,12.1,1.82,52.3,26.1,0.0,3.82,144.9,162.1,5.2,8.4,1.92,0.64,14920e8,4592,3311,3980,3720,61.2,8.7,0.9,2.8,6.1,10.8,4600),
    ("HDFCBANK","NSE","HDFC Bank","Financial Services","Private Banks",1650,18.2,2.47,1.21,15.8,None,None,None,90.7,105.2,14.8,18.2,1.28,0.82,12520e8,1880,1363,1590,1510,55.3,5.4,1.1,1.9,5.2,12.4,2000),
    ("INFY","NSE","Infosys","Technology","IT Services",1780,24.6,7.82,1.64,31.2,21.4,0.0,3.21,72.4,81.6,4.1,6.8,2.84,0.71,7420e8,1953,1307,1710,1620,54.8,4.2,1.0,1.4,4.8,9.2,2050),
    ("HINDUNILVR","NSE","Hindustan Unilever","Consumer Staples","FMCG",2620,54.1,10.2,3.84,19.4,22.8,0.0,1.42,48.4,52.1,2.8,3.2,1.84,0.42,6140e8,2900,2175,2580,2450,57.2,3.1,0.8,-0.4,-1.2,5.8,2900),
    ("ICICIBANK","NSE","ICICI Bank","Financial Services","Private Banks",1340,17.4,2.84,1.08,17.2,None,None,None,76.9,90.4,16.2,22.4,0.82,0.94,9410e8,1521,964,1280,1180,62.4,9.8,1.3,3.8,8.2,18.4,1650),
    ("SBIN","NSE","State Bank of India","Financial Services","PSU Banks",845,9.2,1.42,0.62,14.8,None,None,None,91.6,104.8,12.4,28.4,2.14,1.02,7540e8,912,601,820,780,59.8,7.2,1.2,2.4,6.8,22.4,1050),
    ("BHARTIARTL","NSE","Bharti Airtel","Communication Services","Telecom",1920,42.1,8.74,2.14,22.4,28.4,154.2,0.84,45.6,68.4,18.2,128.4,0.52,0.84,11420e8,2050,1248,1840,1680,63.8,12.4,1.4,4.2,12.8,28.4,2200),
    ("ITC","NSE","ITC","Consumer Staples","Cigarettes",472,28.4,8.12,2.84,28.4,36.2,0.0,2.84,16.6,17.8,4.2,8.4,3.42,0.52,5920e8,528,390,458,440,52.4,1.8,0.9,-1.2,-2.4,2.8,550),
    ("KOTAKBANK","NSE","Kotak Mahindra Bank","Financial Services","Private Banks",1980,22.4,3.42,1.84,12.8,None,None,None,88.4,102.4,12.8,14.2,0.12,0.84,3940e8,2101,1544,1920,1840,54.2,4.8,0.8,0.8,2.4,8.4,2400),
    ("LT","NSE","Larsen & Toubro","Industrials","Construction","3520",22.8,4.21,1.24,14.8,12.4,84.2,1.12,154.2,178.4,18.4,22.4,1.02,0.94,4820e8,3918,2948,3420,3220,62.4,14.8,1.2,3.8,10.4,18.8,4200),
    ("BAJFINANCE","NSE","Bajaj Finance","Financial Services","NBFC",6750,28.4,4.82,1.02,21.4,None,None,None,237.8,284.2,26.4,22.4,0.48,1.12,4180e8,7780,5488,6580,6240,58.4,24.8,1.1,2.4,8.8,18.4,8000),
    ("AXISBANK","NSE","Axis Bank","Financial Services","Private Banks",1185,12.4,1.82,0.84,14.2,None,None,None,95.4,114.8,18.4,28.4,0.12,0.94,3640e8,1340,984,1148,1100,60.2,8.4,1.1,2.8,6.4,16.8,1500),
    ("ASIANPAINT","NSE","Asian Paints","Materials","Paints",2820,54.8,16.4,4.84,28.4,18.4,0.0,1.84,51.4,57.8,4.2,2.4,1.42,0.62,2700e8,3280,2384,2740,2620,53.8,2.4,0.8,-1.8,-4.2,2.4,3100),
    ("MARUTI","NSE","Maruti Suzuki","Consumer Discretionary","Automobiles",12800,28.4,4.84,1.82,14.8,8.4,0.0,1.24,450.2,520.4,14.8,28.4,1.14,0.84,3870e8,13680,9614,12420,11840,62.8,48.4,1.2,4.2,12.4,24.8,15000),
    ("HCLTECH","NSE","HCL Technologies","Technology","IT Services",1920,22.4,6.42,1.64,22.4,18.4,0.0,3.84,85.6,96.8,6.8,10.4,3.14,0.74,5200e8,2114,1414,1860,1740,58.4,6.8,1.0,2.4,6.8,14.2,2200),
    ("SUNPHARMA","NSE","Sun Pharmaceutical","Healthcare","Pharmaceuticals",1840,32.4,5.42,2.14,18.2,22.4,0.0,1.84,56.8,67.4,12.4,18.4,0.82,0.62,4420e8,1960,1370,1784,1680,61.4,8.4,1.1,2.8,8.4,16.8,2100),
    ("TITAN","NSE","Titan Company","Consumer Discretionary","Jewellery",3520,82.4,18.4,4.84,28.4,12.4,0.0,1.42,42.8,50.4,22.4,18.4,0.42,0.84,3120e8,3920,2814,3420,3180,62.4,18.4,1.2,4.2,12.8,22.4,4000),
    ("ULTRACEMCO","NSE","UltraTech Cement","Materials","Cement",11200,32.4,4.42,2.14,12.4,18.4,26.4,1.42,345.8,398.4,8.4,18.4,0.82,0.74,3230e8,12480,8640,10840,10240,58.4,38.4,1.0,2.4,8.4,14.8,13000),
    ("WIPRO","NSE","Wipro","Technology","IT Services",298,20.4,2.84,1.42,14.8,17.4,0.0,3.14,14.6,16.4,2.4,4.8,0.12,0.82,1540e8,320,228,290,274,52.4,1.2,0.8,-1.8,-4.2,2.4,340),
    ("BAJAJFINSV","NSE","Bajaj Finserv","Financial Services","Diversified Financial",1920,14.8,2.42,0.84,14.2,None,None,None,129.6,148.4,18.4,22.4,0.08,0.94,3060e8,2080,1482,1860,1760,57.4,8.4,1.0,2.4,6.8,14.8,2300),
    ("NESTLEIND","NSE","Nestle India","Consumer Staples","Food Products",2420,68.4,78.4,4.84,108.4,22.4,0.0,1.24,35.4,38.8,8.4,12.4,1.84,0.42,2330e8,2680,2060,2380,2280,56.8,8.4,0.8,0.8,2.4,6.8,2800),
    ("TECHM","NSE","Tech Mahindra","Technology","IT Services",1620,28.4,4.84,2.14,14.8,9.4,0.0,2.84,57.2,68.4,8.4,42.4,2.14,0.94,1580e8,1788,1164,1580,1480,62.4,8.4,1.2,4.8,14.2,28.4,1900),
    ("POWERGRID","NSE","Power Grid Corp","Utilities","Electric Utilities",342,18.4,3.42,1.84,18.4,64.2,184.2,1.02,18.6,20.4,6.4,8.4,4.84,0.62,3180e8,368,228,336,312,61.4,2.4,1.1,2.4,8.4,18.4,400),
    ("NTPC","NSE","NTPC","Utilities","Electric Utilities",384,18.4,2.14,1.42,9.8,28.4,128.4,1.02,20.8,22.4,8.4,12.4,2.14,0.84,3720e8,418,248,374,348,62.4,2.8,1.1,2.8,8.8,18.8,450),
    ("ONGC","NSE","Oil & Natural Gas Corp","Energy","Oil & Gas Exploration",284,8.4,1.12,0.82,12.8,18.4,24.4,1.42,33.8,36.4,4.8,28.4,4.84,1.02,3570e8,312,198,278,262,55.4,2.4,0.9,0.8,2.4,8.4,360),
    ("DRREDDY","NSE","Dr. Reddy's Laboratories","Healthcare","Pharmaceuticals",6240,24.8,3.84,1.64,14.8,18.4,0.0,1.84,251.4,288.4,12.8,28.4,0.62,0.62,1040e8,6980,4948,6080,5740,58.4,28.4,1.0,2.4,8.4,14.8,7500),
    ("DIVISLAB","NSE","Divi's Laboratories","Healthcare","Pharmaceuticals",5420,48.4,8.42,3.84,18.4,28.4,0.0,3.84,112.0,138.4,14.8,38.4,1.14,0.52,1440e8,5984,3620,5280,4980,62.4,28.4,1.2,4.8,14.2,24.8,6500),
    ("CIPLA","NSE","Cipla","Healthcare","Pharmaceuticals",1580,28.4,4.84,2.14,14.8,18.4,0.0,1.84,55.6,64.8,8.4,18.4,0.42,0.62,1270e8,1744,1160,1540,1460,59.4,8.4,1.0,2.8,8.4,14.8,1900),
    ("APOLLOHOSP","NSE","Apollo Hospitals","Healthcare","Healthcare Facilities",7240,84.8,14.2,4.84,14.8,9.4,112.4,1.02,85.4,108.4,22.4,38.4,0.22,0.94,1040e8,7980,4844,7080,6620,64.8,48.4,1.3,6.8,18.4,32.4,9000),
    ("LTIM","NSE","LTIMindtree","Technology","IT Services",5820,28.4,7.42,1.84,22.4,18.4,0.0,3.84,204.8,234.4,8.4,14.8,1.02,0.84,1720e8,6384,4588,5680,5360,62.4,28.4,1.1,4.2,12.4,18.8,6800),
    ("ADANIENT","NSE","Adani Enterprises","Industrials","Diversified","2980",84.8,6.42,4.84,6.8,6.4,184.2,1.02,35.2,44.8,28.4,48.4,0.08,1.42,3380e8,3476,1720,2880,2640,58.4,18.4,1.3,4.2,14.8,28.4,3800),
    ("ADANIPORTS","NSE","Adani Ports","Industrials","Port Services",1380,28.4,4.84,1.84,14.8,52.4,124.4,1.12,48.6,58.4,18.4,22.4,0.62,1.12,2980e8,1524,900,1340,1240,62.4,8.4,1.2,4.2,12.4,22.4,1700),
    ("TATACONSUM","NSE","Tata Consumer Products","Consumer Staples","FMCG",1140,58.4,7.84,4.84,9.8,8.4,8.4,1.42,19.5,24.8,14.8,28.4,0.82,0.72,1050e8,1248,860,1108,1040,58.4,4.8,0.9,1.4,4.8,10.8,1400),
    ("BRITANNIA","NSE","Britannia Industries","Consumer Staples","Food Products",5820,58.4,22.4,4.84,48.4,18.4,0.0,1.14,99.6,108.4,6.4,8.4,1.62,0.52,1400e8,6384,4268,5684,5440,55.4,28.4,0.8,0.8,2.4,6.4,6800),
    ("HEROMOTOCO","NSE","Hero MotoCorp","Consumer Discretionary","2-Wheelers",4860,18.4,4.84,1.64,18.4,12.4,0.0,1.24,263.8,298.4,8.4,14.8,3.84,0.62,972e8,5420,3610,4740,4520,60.4,28.4,1.0,2.4,6.8,12.8,5800),
    ("BAJAJ-AUTO","NSE","Bajaj Auto","Consumer Discretionary","2-Wheelers",9840,22.4,6.84,1.84,24.8,14.8,0.0,1.84,438.8,494.4,12.8,18.4,1.52,0.72,2760e8,10840,7108,9580,9020,62.4,58.4,1.1,3.8,10.4,18.8,11500),
    ("EICHERMOT","NSE","Eicher Motors","Consumer Discretionary","2-Wheelers",4920,28.4,7.84,2.14,22.4,22.4,0.0,2.84,173.4,198.4,8.4,14.8,1.14,0.84,1340e8,5280,3614,4820,4620,61.4,28.4,1.0,2.8,8.4,14.8,5800),
    ("M&M","NSE","Mahindra & Mahindra","Consumer Discretionary","Automobiles",3220,28.4,4.84,1.64,14.8,8.4,48.4,1.24,113.6,138.4,22.4,38.4,1.22,0.94,3990e8,3580,1914,3140,2920,65.8,28.4,1.4,6.8,18.4,32.4,3900),
    ("TATAMOTORS","NSE","Tata Motors","Consumer Discretionary","Automobiles",984,8.4,2.84,0.82,22.4,6.4,138.4,1.02,117.2,148.4,18.4,228.4,0.22,1.42,3620e8,1084,634,960,900,61.4,8.4,1.2,3.8,10.4,22.4,1200),
    ("TATAPOWER","NSE","Tata Power","Utilities","Electric Utilities",438,34.8,3.84,2.14,9.8,18.4,124.4,0.84,12.6,16.4,18.4,28.4,0.48,0.84,1400e8,480,298,428,404,62.4,4.8,1.2,4.2,12.4,22.4,520),
    ("HINDALCO","NSE","Hindalco Industries","Materials","Aluminium",684,12.4,1.84,1.24,9.8,8.4,48.4,1.24,55.2,68.4,8.4,18.4,0.62,1.12,1520e8,748,490,664,628,60.4,8.4,1.1,3.4,8.8,16.8,840),
    ("JSWSTEEL","NSE","JSW Steel","Materials","Steel",984,18.4,2.84,1.64,12.4,8.4,84.4,1.12,53.4,68.4,4.8,28.4,1.22,1.12,2400e8,1080,748,960,920,61.4,8.4,1.2,3.8,10.4,18.8,1200),
    ("TATASTEEL","NSE","Tata Steel","Materials","Steel",164,18.4,2.14,1.84,8.4,8.4,124.4,1.14,8.9,12.4,4.8,128.4,1.42,1.24,2050e8,184,126,160,152,58.4,1.8,1.1,2.4,6.8,14.8,210),
    ("INDUSINDBK","NSE","IndusInd Bank","Financial Services","Private Banks",1248,12.4,1.84,0.84,12.4,None,None,None,100.6,118.4,12.4,8.4,1.22,1.14,970e8,1694,904,1220,1160,54.4,4.8,0.9,-2.4,-8.4,8.4,1600),
    ("GRASIM","NSE","Grasim Industries","Materials","Diversified Chemicals",2840,24.8,2.84,1.84,8.4,22.4,48.4,1.24,114.4,138.4,14.8,22.4,0.42,0.84,1860e8,3108,2024,2760,2600,58.4,14.8,1.0,2.4,6.8,14.8,3400),
    ("WIPRO","NSE","Wipro","Technology","IT Services",298,20.4,2.84,1.64,14.8,17.4,0.0,3.14,14.6,16.4,2.4,4.8,0.12,0.82,1540e8,320,228,290,274,52.4,1.2,0.8,-1.8,-4.2,2.4,340),
    # Mid-cap growth stocks
    ("ZOMATO","NSE","Zomato","Consumer Discretionary","Food Delivery",258,484.8,12.4,18.4,2.4,3.4,0.0,2.84,0.5,1.8,62.4,128.4,0.0,1.42,2280e8,284,134,248,218,68.4,4.8,1.8,8.4,28.4,62.4,320),
    ("NAUKRI","NSE","Info Edge (Naukri)","Technology","Internet",6840,68.4,8.84,4.84,12.4,22.4,0.0,4.84,100.2,118.4,12.4,18.4,0.42,0.84,880e8,7484,4848,6640,6280,62.4,38.4,1.2,4.8,14.2,22.4,8000),
    ("MPHASIS","NSE","Mphasis","Technology","IT Services",2920,28.4,5.84,1.84,18.4,18.4,0.0,3.84,102.8,118.4,8.4,14.8,2.14,0.74,546e8,3196,2128,2840,2700,60.4,18.4,1.0,2.8,8.4,14.8,3400),
    ("PERSISTENT","NSE","Persistent Systems","Technology","IT Services",5440,42.4,12.4,2.84,28.4,16.4,0.0,2.84,128.4,154.8,18.4,24.8,0.62,0.82,836e8,5984,3784,5300,4980,65.4,28.4,1.3,5.8,16.4,28.4,6800),
    ("TATAELXSI","NSE","Tata Elxsi","Technology","IT Services",7840,48.4,16.4,3.84,28.4,28.4,0.0,3.84,162.4,188.4,12.4,14.8,0.82,0.72,1220e8,9084,5448,7640,7220,61.4,38.4,1.0,3.8,10.4,18.8,9200),
    ("COFORGE","NSE","Coforge","Technology","IT Services",7840,38.4,10.4,2.84,22.4,14.8,14.4,2.84,204.2,238.4,22.4,38.4,0.52,0.94,478e8,8680,4760,7640,7180,64.8,48.4,1.4,6.8,18.4,32.4,9500),
    ("KPITTECH","NSE","KPIT Technologies","Technology","IT Services",1820,62.4,18.4,3.84,28.4,18.4,0.0,3.84,29.2,36.4,28.4,38.4,0.28,1.02,492e8,2008,1080,1780,1660,65.4,14.8,1.4,7.2,22.4,38.4,2200),
    ("PIDILITIND","NSE","Pidilite Industries","Materials","Adhesives",3120,62.4,14.2,4.84,22.4,18.4,0.0,1.84,50.0,56.4,8.4,12.4,0.42,0.52,1580e8,3396,2384,3060,2940,58.4,14.8,0.8,1.4,4.8,9.8,3600),
    ("HAVELLS","NSE","Havells India","Industrials","Electrical Equipment",1840,52.4,10.4,4.84,18.4,14.4,0.0,1.84,35.0,42.4,14.8,22.4,0.82,0.72,1150e8,2024,1340,1800,1720,62.4,12.4,1.1,4.2,12.4,22.4,2200),
    ("BOSCHLTD","NSE","Bosch","Consumer Discretionary","Auto Components",37840,42.4,9.84,3.84,18.4,14.8,0.0,3.84,892.4,1028.4,14.8,22.4,0.42,0.52,1070e8,41200,28084,37000,35200,62.4,228.4,0.8,2.8,8.4,14.8,44000),
    ("SIEMENS","NSE","Siemens","Industrials","Industrial Conglomerates",6840,68.4,12.4,5.84,14.8,12.4,0.0,2.84,100.0,118.4,18.4,28.4,0.28,0.52,2430e8,7484,4048,6680,6340,63.4,38.4,1.2,4.8,14.2,22.4,8000),
    ("ABB","NSE","ABB India","Industrials","Electrical Equipment",6040,68.4,16.4,5.84,22.4,16.4,0.0,3.84,88.4,108.4,22.4,38.4,0.28,0.62,1280e8,6580,3948,5900,5640,64.8,38.4,1.2,5.8,16.4,28.4,7200),
    ("MUTHOOTFIN","NSE","Muthoot Finance","Financial Services","NBFC",1984,14.8,2.84,1.02,14.8,None,None,None,134.0,158.4,18.4,18.4,2.42,0.84,796e8,2184,1224,1940,1820,63.4,14.8,1.2,4.8,14.2,24.8,2400),
    ("GODREJCP","NSE","Godrej Consumer Products","Consumer Staples","FMCG",1284,52.4,8.84,4.84,14.8,12.4,0.0,1.42,24.5,28.4,8.4,12.4,1.14,0.62,1310e8,1408,940,1260,1220,57.4,8.4,0.8,0.8,2.4,6.8,1500),
    ("DMART","NSE","Avenue Supermarts","Consumer Staples","Retail",4640,84.8,18.4,5.84,14.8,9.4,0.0,1.84,54.7,68.4,14.8,18.4,0.0,0.52,3010e8,5060,3524,4540,4360,60.4,28.4,0.9,2.4,6.8,12.4,5500),
    ("BERGEPAINT","NSE","Berger Paints","Materials","Paints",584,52.4,12.4,4.84,22.4,12.4,0.0,1.84,11.1,12.8,8.4,12.4,1.42,0.62,566e8,644,448,570,548,55.4,2.4,0.8,-0.8,-2.4,4.8,680),
    ("AUROPHARMA","NSE","Aurobindo Pharma","Healthcare","Pharmaceuticals",1284,16.4,2.84,1.24,14.8,18.4,24.4,1.42,78.3,92.4,8.4,14.8,1.22,0.82,750e8,1400,904,1248,1180,59.4,8.4,1.0,2.4,7.2,14.8,1600),
    ("TORNTPHARM","NSE","Torrent Pharma","Healthcare","Pharmaceuticals",3420,38.4,8.84,2.84,22.4,22.4,28.4,1.42,89.1,108.4,12.4,22.4,1.14,0.62,580e8,3780,2408,3340,3180,62.4,18.4,1.0,3.4,9.8,16.8,4000),
    ("LUPIN","NSE","Lupin","Healthcare","Pharmaceuticals",2320,28.4,4.84,2.14,14.8,14.8,0.0,1.24,82.6,100.4,12.4,38.4,0.82,0.72,1050e8,2560,1552,2280,2160,60.4,14.8,1.0,2.8,8.4,14.8,2800),
    ("CHOLAFIN","NSE","Cholamandalam Invest","Financial Services","NBFC",1580,28.4,5.84,1.84,14.8,None,None,None,55.6,68.4,28.4,28.4,0.28,0.94,1300e8,1740,1008,1540,1460,63.4,12.4,1.2,4.8,14.2,24.8,2000),
    ("RECLTD","NSE","REC Limited","Financial Services","Power Finance",578,7.4,1.84,0.62,14.8,None,None,None,78.1,94.8,18.4,22.4,4.84,0.94,1530e8,642,368,564,532,62.4,8.4,1.2,4.8,14.2,24.8,750),
    ("PFC","NSE","Power Finance Corp","Financial Services","Power Finance",468,6.8,1.42,0.62,12.4,None,None,None,68.8,84.8,14.8,22.4,5.84,0.94,1540e8,520,286,456,428,62.4,6.8,1.1,4.2,12.8,22.4,620),
    ("CANBK","NSE","Canara Bank","Financial Services","PSU Banks",108,5.4,0.84,0.62,9.8,None,None,None,20.0,24.4,12.4,28.4,3.42,1.02,983e8,120,74,106,100,60.4,1.4,1.1,2.8,8.4,18.4,140),
    ("BANKBARODA","NSE","Bank of Baroda","Financial Services","PSU Banks",248,7.8,1.12,0.82,12.4,None,None,None,31.8,38.4,14.8,22.4,3.14,1.02,1280e8,274,174,244,234,61.4,2.4,1.1,2.8,8.4,16.8,320),
    ("IDBI","BSE","IDBI Bank","Financial Services","PSU Banks",88,18.4,1.84,1.42,8.4,None,None,None,4.8,5.8,8.4,28.4,1.14,0.84,943e8,96,58,86,80,58.4,0.8,1.0,1.8,5.4,12.4,110),
    ("GAIL","NSE","GAIL India","Energy","Natural Gas",228,14.8,2.14,1.42,14.8,14.8,24.4,1.42,15.4,17.8,6.4,28.4,3.84,0.82,1500e8,250,148,224,212,61.4,2.4,1.1,3.8,10.4,18.8,280),
    ("BPCL","NSE","Bharat Petroleum","Energy","Oil Refining",328,12.4,2.84,1.24,22.4,4.8,24.4,1.42,26.4,30.8,8.4,18.4,5.84,1.02,1420e8,360,198,320,300,61.4,3.8,1.1,3.8,10.4,18.8,400),
    ("COALINDIA","NSE","Coal India","Energy","Coal",464,7.8,3.42,1.24,42.4,38.4,0.0,2.42,59.5,64.8,4.8,12.4,6.84,0.52,2860e8,502,338,456,430,58.4,5.4,0.9,1.4,4.8,9.8,560),
    ("VEDL","NSE","Vedanta","Materials","Diversified Metals",468,8.4,2.84,1.24,22.4,22.4,124.4,0.84,55.8,68.4,8.4,38.4,14.84,1.24,1740e8,520,268,456,428,60.4,5.8,1.1,2.8,8.4,14.8,580),
    ("NMDC","NSE","NMDC Steel","Materials","Iron Ore",68,12.4,2.84,1.64,18.4,38.4,0.0,1.84,5.5,6.8,8.4,28.4,4.84,0.82,1980e8,76,48,66,62,58.4,0.8,1.0,2.4,6.8,12.8,85),
    # BSE-only names (using BSE codes, yfinance .BO suffix)
    ("500325","BSE","Reliance Industries","Energy","Integrated Oil & Gas",2948,21.2,2.13,0.90,18.2,7.1,44.0,1.17,138.2,164.5,7.3,12.6,3.34,0.93,19820e8,3215,2218,2870,2730,58.2,12.2,1.2,3.0,8.3,14.1,3375),
    ("532540","BSE","TCS","Technology","IT Services",4118,28.3,12.0,1.81,52.2,26.0,0.0,3.81,144.7,161.9,5.1,8.3,1.91,0.64,14910e8,4590,3309,3978,3718,61.0,8.6,0.9,2.7,6.0,10.7,4595),
    ("500180","BSE","HDFC Bank","Financial Services","Private Banks",1648,18.1,2.46,1.20,15.7,None,None,None,90.5,105.0,14.7,18.1,1.27,0.82,12510e8,1878,1361,1588,1508,55.1,5.3,1.1,1.8,5.1,12.3,1998),
    ("500209","BSE","Infosys","Technology","IT Services",1778,24.5,7.80,1.63,31.1,21.3,0.0,3.20,72.2,81.4,4.0,6.7,2.83,0.71,7410e8,1951,1305,1708,1618,54.6,4.1,1.0,1.3,4.7,9.1,2048),
    ("500696","BSE","Hindustan Unilever","Consumer Staples","FMCG",2618,54.0,10.1,3.83,19.3,22.7,0.0,1.41,48.2,51.9,2.7,3.1,1.83,0.42,6130e8,2898,2173,2578,2448,57.0,3.0,0.8,-0.5,-1.3,5.7,2898),
    ("532174","BSE","ICICI Bank","Financial Services","Private Banks",1338,17.3,2.83,1.07,17.1,None,None,None,76.7,90.2,16.1,22.3,0.81,0.94,9400e8,1519,962,1278,1178,62.2,9.7,1.3,3.7,8.1,18.3,1648),
    ("500112","BSE","SBI","Financial Services","PSU Banks",843,9.1,1.41,0.61,14.7,None,None,None,91.4,104.6,12.3,28.3,2.13,1.02,7530e8,910,599,818,778,59.6,7.1,1.2,2.3,6.7,22.3,1048),
    ("532454","BSE","Bharti Airtel","Communication Services","Telecom",1918,42.0,8.72,2.13,22.3,28.3,154.1,0.83,45.4,68.2,18.1,128.3,0.51,0.84,11410e8,2048,1246,1838,1678,63.6,12.3,1.4,4.1,12.7,28.3,2198),
]

COLUMNS = [
    "symbol","exchange","company_name","sector","industry",
    "last_price","pe_ratio","pb_ratio","peg_ratio","roe","operating_margins",
    "debt_equity","current_ratio","eps_trailing","eps_forward",
    "revenue_growth","earnings_growth","dividend_yield","beta","market_cap",
    "w52_high","w52_low","ma_50","ma_200","rsi_14","macd","volume_ratio",
    "return_1m_pct","return_3m_pct","return_6m_pct","target_price",
]


# ─────────────────────────────────────────────────────────────
# NSE & BSE SYMBOL LISTS  (used by live fetchers)
# ─────────────────────────────────────────────────────────────
_NSE_NIFTY500 = [
    # NIFTY50 core
    "RELIANCE","TCS","HDFCBANK","INFY","HINDUNILVR","ICICIBANK","SBIN",
    "BHARTIARTL","ITC","KOTAKBANK","LT","BAJFINANCE","AXISBANK","ASIANPAINT",
    "MARUTI","HCLTECH","SUNPHARMA","TITAN","ULTRACEMCO","BAJAJFINSV",
    "NESTLEIND","TECHM","POWERGRID","NTPC","ONGC","DRREDDY","DIVISLAB","CIPLA",
    "APOLLOHOSP","LTIM","ADANIENT","ADANIPORTS","TATACONSUM","BRITANNIA",
    "HEROMOTOCO","BAJAJ-AUTO","EICHERMOT","M&M","TATAMOTORS","TATAPOWER",
    "HINDALCO","JSWSTEEL","TATASTEEL","INDUSINDBK","GRASIM","WIPRO",
    # NIFTY100 add-ons
    "ZOMATO","NAUKRI","MPHASIS","PERSISTENT","TATAELXSI","COFORGE","KPITTECH",
    "PIDILITIND","HAVELLS","BOSCHLTD","SIEMENS","ABB","MUTHOOTFIN","GODREJCP",
    "DMART","BERGEPAINT","AUROPHARMA","TORNTPHARM","LUPIN","CHOLAFIN",
    "RECLTD","PFC","CANBK","BANKBARODA","GAIL","BPCL","COALINDIA","VEDL","NMDC",
    # NIFTY500 additions
    "SHREECEM","HDFCLIFE","SBILIFE","ICICIlombard","PAGEIND","MARICO","DABUR",
    "COLPAL","PFIZER","ABBOTINDIA","ALKEM","BIOCON","GLENMARK","NATCOPHARM",
    "GRANULES","ASTRAL","SUPREMEIND","AAVAS","CREDITACC","SBICARD","LTTS",
    "SONACOMS","DIXON","VOLTAS","WHIRLPOOL","BLUESTAR","AMBER","POLYCAB",
    "KEI","SUZLON","IRCTC","INDIGO","ICICIGI","HDFCGI","MANAPPURAM",
    "BAJAJHFL","LICHSGFIN","CONCOR","DELHIVERY","TCI","RAJESHEXPO","KAYNES",
]

_BSE_SCRIP_CODES = [
    # BSE 500 top stocks by market cap
    "500325","532540","500180","500209","500696","532174","500112","532454",
    "500875","500247","500010","532978","532648","500820","532500","532281",
    "500034","532667","507685","500570","532541","500790","532187","500520",
    "500480","500124","500830","539253","532498","500002","500400","506395",
    "532822","500530","532755","540065","543900","532779","519552","512599",
    "500030","500114","500295","524715","500251","500878","524004","504973",
    "500086","500049","500116","500440","500300","532209","505537","542726",
    "500850","532800","532424","543323","543320","532652","500183","543257",
    "532210","500477","500285","533278","543210","500510","524715","532187",
]


def _safe_float(val, default=None):
    """Convert val to float, returning default for None/NaN/Inf."""
    try:
        v = float(val)
        return default if (math.isnan(v) or math.isinf(v)) else v
    except (TypeError, ValueError):
        return default


def fetch_live_nse_data(symbols: List[str]) -> pd.DataFrame:
    """Fetch live NSE quotes using nsepython (nse_eq per symbol)."""
    try:
        from nsepython import nse_eq  # type: ignore
    except ImportError:
        print("[WARN] nsepython not installed — pip install nsepython")
        return pd.DataFrame()

    records, failed = [], 0
    print(f"  [nsepython] Fetching {len(symbols)} NSE symbols…")
    for sym in symbols:
        try:
            data = nse_eq(sym)
            pi   = data.get("priceInfo", {})
            meta = data.get("metadata", {})
            whl  = pi.get("weekHighLow", {})
            records.append({
                "symbol":       sym,
                "exchange":     "NSE",
                "company_name": meta.get("companyName", sym),
                "sector":       meta.get("industry", ""),
                "industry":     meta.get("industry", ""),
                "last_price":   _safe_float(pi.get("lastPrice")),
                "w52_high":     _safe_float(whl.get("max")),
                "w52_low":      _safe_float(whl.get("min")),
                "pe_ratio":     _safe_float(
                                    (data.get("industryInfo") or {}).get("pe")),
                "fetch_date":   TODAY,
            })
        except Exception:
            failed += 1

    print(f"  [nsepython] ok={len(records)}, skipped={failed}")
    return pd.DataFrame(records)


def fetch_live_bse_data(scrip_codes: List[str]) -> pd.DataFrame:
    """Fetch live BSE quotes using the bsedata library (getQuote per scrip)."""
    try:
        from bsedata.bse import BSE  # type: ignore
        b = BSE()
    except ImportError:
        print("[WARN] bsedata not installed — pip install bsedata")
        return pd.DataFrame()

    records, failed = [], 0
    print(f"  [bsedata] Fetching {len(scrip_codes)} BSE scrip codes…")
    for code in scrip_codes:
        try:
            q = b.getQuote(str(code))
            if not q:
                continue
            records.append({
                "symbol":         str(code),
                "exchange":       "BSE",
                "company_name":   q.get("companyName", ""),
                "sector":         q.get("industry", ""),
                "industry":       q.get("industry", ""),
                "last_price":     _safe_float(q.get("currentValue")),
                "pe_ratio":       _safe_float(q.get("ttmPe")),
                "pb_ratio":       _safe_float(q.get("pbv")),
                "eps_trailing":   _safe_float(q.get("ttmEps")),
                "w52_high":       _safe_float(q.get("weeks52High")),
                "w52_low":        _safe_float(q.get("weeks52Low")),
                "market_cap":     _safe_float(q.get("marketCapFull")),
                "dividend_yield": _safe_float(q.get("dividendYield")),
                "book_value":     _safe_float(q.get("bookValue")),
                "fetch_date":     TODAY,
            })
        except Exception:
            failed += 1

    print(f"  [bsedata] ok={len(records)}, skipped={failed}")
    return pd.DataFrame(records)


def enrich_with_yfinance(df: pd.DataFrame) -> pd.DataFrame:
    """Add fundamentals + technicals from yfinance (.NS/.BO suffixes)."""
    try:
        import yfinance as yf  # type: ignore
    except ImportError:
        print("[WARN] yfinance not installed — skipping enrichment")
        return df

    enriched = []
    print(f"  [yfinance] Enriching {len(df)} stocks with fundamentals + technicals…")
    for _, row in df.iterrows():
        sym    = row["symbol"]
        exch   = row.get("exchange", "NSE")
        suffix = ".NS" if exch == "NSE" else ".BO"
        rec    = row.to_dict()
        try:
            t    = yf.Ticker(sym + suffix)
            info = t.info or {}
            hist = t.history(period="1y")

            # Fundamentals from info dict
            for dest, src in [
                ("pe_ratio",          "trailingPE"),
                ("pb_ratio",          "priceToBook"),
                ("peg_ratio",         "trailingPegRatio"),
                ("roe",               "returnOnEquity"),
                ("operating_margins", "operatingMargins"),
                ("debt_equity",       "debtToEquity"),
                ("current_ratio",     "currentRatio"),
                ("eps_trailing",      "trailingEps"),
                ("eps_forward",       "forwardEps"),
                ("revenue_growth",    "revenueGrowth"),
                ("earnings_growth",   "earningsGrowth"),
                ("dividend_yield",    "dividendYield"),
                ("beta",              "beta"),
                ("market_cap",        "marketCap"),
                ("target_price",      "targetMeanPrice"),
                ("w52_high",          "fiftyTwoWeekHigh"),
                ("w52_low",           "fiftyTwoWeekLow"),
            ]:
                v = _safe_float(info.get(src))
                if v is not None:
                    rec[dest] = v

            for dest, src in [("company_name","longName"),
                               ("sector","sector"), ("industry","industry")]:
                if info.get(src):
                    rec[dest] = info[src]

            # Technicals from price history
            if len(hist) >= 50:
                close = hist["Close"]
                vol   = hist["Volume"]
                rec["ma_50"]  = float(close.rolling(50).mean().iloc[-1])
                if len(hist) >= 200:
                    rec["ma_200"] = float(close.rolling(200).mean().iloc[-1])

                # RSI-14
                delta = close.diff()
                gain  = delta.clip(lower=0).rolling(14).mean()
                loss  = (-delta.clip(upper=0)).rolling(14).mean()
                rs    = gain / loss.replace(0, np.nan)
                rsi   = 100 - 100 / (1 + rs)
                if pd.notna(rsi.iloc[-1]):
                    rec["rsi_14"] = float(rsi.iloc[-1])

                # MACD
                ema12 = close.ewm(span=12, adjust=False).mean()
                ema26 = close.ewm(span=26, adjust=False).mean()
                rec["macd"] = float((ema12 - ema26).iloc[-1])

                # Volume ratio (20-day avg)
                vol_avg = vol.rolling(20).mean().iloc[-1]
                if vol_avg and vol_avg > 0:
                    rec["volume_ratio"] = float(vol.iloc[-1] / vol_avg)

                # Return periods
                if len(hist) >= 20:
                    rec["return_1m_pct"] = round((close.iloc[-1]/close.iloc[-20]-1)*100, 2)
                if len(hist) >= 63:
                    rec["return_3m_pct"] = round((close.iloc[-1]/close.iloc[-63]-1)*100, 2)
                if len(hist) >= 126:
                    rec["return_6m_pct"] = round((close.iloc[-1]/close.iloc[-126]-1)*100, 2)

        except Exception:
            pass  # keep what nsepython/bsedata already gave us

        enriched.append(rec)

    return pd.DataFrame(enriched)


def _normalise_live(df: pd.DataFrame) -> pd.DataFrame:
    """Compute derived columns needed by Pegu/Sarvas for live-fetched data."""
    # yfinance returns ROE/margins as decimals (0.18); multiply if < 2
    for src, dst in [
        ("roe",               "roe_pct"),
        ("operating_margins", "operating_margins_pct"),
        ("revenue_growth",    "revenue_growth_pct"),
        ("earnings_growth",   "earnings_growth_pct"),
        ("dividend_yield",    "dividend_yield_pct"),
    ]:
        if src in df.columns:
            df[dst] = df[src].apply(
                lambda v: (v * 100) if (pd.notna(v) and abs(v) < 2) else v)
        elif dst not in df.columns:
            df[dst] = np.nan

    for num, den, col in [
        ("last_price", "ma_50",     "pct_above_50dma"),
        ("last_price", "ma_200",    "pct_above_200dma"),
        ("last_price", "w52_high",  "pct_from_52w_high"),
    ]:
        if num in df.columns and den in df.columns:
            df[col] = ((df[num] / df[den]) - 1).mul(100).round(2)
        elif col not in df.columns:
            df[col] = np.nan

    if "target_price" in df.columns and "last_price" in df.columns:
        df["upside_pct"] = ((df["target_price"] / df["last_price"]) - 1).mul(100).round(2)
    elif "upside_pct" not in df.columns:
        df["upside_pct"] = np.nan

    if "fetch_date" not in df.columns:
        df["fetch_date"] = TODAY
    return df


def load_or_generate_data(data_dir: str, use_mock: bool) -> pd.DataFrame:
    """
    Data cascade:
      1. Extractor CSV  — populated by nse_bse_extractor.py
      2. Live fetch     — nsepython (NSE) + bsedata (BSE) + yfinance enrichment
      3. Sample fallback — representative NIFTY500 data (network blocked / --mock)
    """
    combined = os.path.join(data_dir, "all_stocks_combined.csv")

    # ── 1. Extractor CSV ─────────────────────────────────────
    if not use_mock and os.path.exists(combined):
        df = pd.read_csv(combined)
        if "last_price" in df.columns and df["last_price"].notna().sum() > 5:
            print(f"[DATA] Loaded {len(df)} stocks from {combined}")
            return _normalise_live(df)

    # ── 2. Live fetch: nsepython + bsedata ───────────────────
    if not use_mock:
        print("[DATA] Fetching live data: nsepython (NSE) + bsedata (BSE)…")
        nse_df = fetch_live_nse_data(_NSE_NIFTY500)
        bse_df = fetch_live_bse_data(_BSE_SCRIP_CODES)

        frames = [f for f in [nse_df, bse_df] if not f.empty]
        if frames:
            live_df = pd.concat(frames, ignore_index=True)
            print(f"[DATA] Live: {len(live_df)} stocks — enriching with yfinance…")
            live_df = enrich_with_yfinance(live_df)
            live_df = _normalise_live(live_df)
            os.makedirs(data_dir, exist_ok=True)
            live_df.to_csv(combined, index=False)
            print(f"[DATA] Saved to {combined}")
            return live_df

        print("[DATA] Live fetch returned no data (network blocked). Falling back to sample data.")

    # ── 3. Parquet cache (Git LFS checkout or local Downloads) ──
    if not use_mock:
        try:
            from market_data_cache import load_nifty_stocks_from_cache, MarketCache
            cache = MarketCache(verbose=False)
            avail = cache.available_symbols()
            if avail:
                print(f"[DATA] Loading {len(avail)} symbols from parquet cache (LFS)…")
                cache_df = load_nifty_stocks_from_cache(verbose=True)
                if not cache_df.empty:
                    # Merge cache technicals into sample fundamentals where symbol matches
                    sample_df = _build_sample_df()
                    merged = sample_df.merge(
                        cache_df[["symbol","last_price","rsi_14","macd","volume_ratio",
                                  "pct_above_50dma","pct_above_200dma","pct_from_52w_high",
                                  "ret_1m","ret_3m","ret_6m","week52_high","week52_low",
                                  "ma50","ma200"]],
                        on="symbol", how="left", suffixes=("", "_cache")
                    )
                    # Override sample price + technicals with cache values where available
                    for col in ("last_price","rsi_14","macd","volume_ratio",
                                "pct_above_50dma","pct_above_200dma","pct_from_52w_high",
                                "ret_1m","ret_3m","ret_6m"):
                        cache_col = f"{col}_cache"
                        if cache_col in merged.columns:
                            mask = merged[cache_col].notna()
                            merged.loc[mask, col] = merged.loc[mask, cache_col]
                            merged.drop(columns=[cache_col], inplace=True, errors="ignore")
                    for pair in (("week52_high","w52_high"),("week52_low","w52_low"),
                                 ("ma50","ma_50"),("ma200","ma_200")):
                        cache_col, orig_col = pair
                        if cache_col in merged.columns:
                            mask = merged[cache_col].notna()
                            merged.loc[mask, orig_col] = merged.loc[mask, cache_col]
                            merged.drop(columns=[cache_col], inplace=True, errors="ignore")
                    merged["fetch_date"] = TODAY
                    merged["data_source"] = "parquet_cache"
                    print(f"[DATA] Parquet cache enriched {mask.sum()} of {len(merged)} stocks")
                    return merged
        except ImportError:
            pass
        except Exception as e:
            print(f"[DATA] Parquet cache error: {e}")

    # ── 4. Sample fallback ───────────────────────────────────
    print("[DATA] Using representative NIFTY500 sample data (--mock or no network)")
    return _build_sample_df()


def _build_sample_df() -> pd.DataFrame:
    """Build the representative sample DataFrame from _SAMPLE_STOCKS."""
    rows = []
    for row in _SAMPLE_STOCKS:
        d = dict(zip(COLUMNS, row))
        for col in ("last_price", "pe_ratio"):
            if isinstance(d.get(col), str):
                try:   d[col] = float(d[col])
                except ValueError: d[col] = np.nan
        rows.append(d)

    df = pd.DataFrame(rows)
    for col in ("roe","operating_margins","revenue_growth","earnings_growth","dividend_yield"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["roe_pct"]               = df["roe"]
    df["operating_margins_pct"] = df["operating_margins"]
    df["revenue_growth_pct"]    = df["revenue_growth"]
    df["earnings_growth_pct"]   = df["earnings_growth"]
    df["dividend_yield_pct"]    = df["dividend_yield"]

    df["pct_above_50dma"]  = ((df["last_price"] / df["ma_50"]  - 1) * 100).round(2)
    df["pct_above_200dma"] = ((df["last_price"] / df["ma_200"] - 1) * 100).round(2)
    df["pct_from_52w_high"]= ((df["last_price"] / df["w52_high"] - 1) * 100).round(2)
    df["upside_pct"]       = ((df["target_price"] / df["last_price"] - 1) * 100).round(2)
    df["fetch_date"]       = TODAY
    df["data_source"]      = "sample"
    return df


# ─────────────────────────────────────────────────────────────
# PEGU SCORING  (Python port of pegu_sarvas_analysis.R)
# ─────────────────────────────────────────────────────────────
def _score(val, breakpoints):
    """Map a value to a score using (threshold, score) breakpoints."""
    if pd.isna(val):
        return breakpoints[-1][1]
    for threshold, score in breakpoints:
        if val < threshold:
            return score
    return breakpoints[-1][1]


def calculate_pegu_scores(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Valuation (max 30)
    df["pe_score"]  = df["pe_ratio"].apply(lambda v: _score(v, [
        (0,0),(8,10),(12,9),(16,7),(22,5),(30,3),(50,1),(float("inf"),0),(float("nan"),3)
    ]) if not pd.isna(v) and v > 0 else 3)
    df["peg_score"] = df["peg_ratio"].apply(lambda v: 4 if pd.isna(v) else (
        0 if v <= 0 else _score(v, [(0.5,10),(1.0,8),(1.5,5),(2.0,2),(float("inf"),0)])))
    df["pb_score"]  = df["pb_ratio"].apply(lambda v: 3 if (pd.isna(v) or v <= 0) else
        _score(v, [(1,10),(2,8),(3,6),(5,4),(10,2),(float("inf"),0)]))
    df["valuation_score"] = df[["pe_score","peg_score","pb_score"]].sum(axis=1)

    # Quality (max 30)
    df["roe_score"] = df["roe_pct"].apply(lambda v: 3 if pd.isna(v) else (
        _score(-v, [(-35,10),(-25,9),(-20,7),(-15,5),(-10,3),(0,1),(float("inf"),0)])))
    df["margin_score"] = df["operating_margins_pct"].apply(lambda v: 3 if pd.isna(v) else (
        _score(-v, [(-35,10),(-25,8),(-18,6),(-12,4),(-6,2),(0,1),(float("inf"),0)])))
    df["debt_score"] = df["debt_equity"].apply(lambda v: 4 if pd.isna(v) else (
        10 if v == 0 else _score(v, [(0.1,9),(0.3,8),(0.5,7),(1.0,5),(1.5,3),(2.5,1),(float("inf"),0)])))
    df["quality_score"] = df[["roe_score","margin_score","debt_score"]].sum(axis=1)

    # Growth (max 25)
    df["eps_growth_score"] = df["earnings_growth_pct"].apply(lambda v: 3 if pd.isna(v) else (
        _score(-v, [(-40,10),(-25,8),(-15,6),(-8,4),(0,2),(float("inf"),0)])))
    df["rev_growth_score"] = df["revenue_growth_pct"].apply(lambda v: 3 if pd.isna(v) else (
        _score(-v, [(-30,10),(-20,8),(-12,6),(-6,4),(0,2),(float("inf"),0)])))
    def fpe_score(row):
        pe, fpe = row.get("pe_ratio"), row.get("eps_forward")
        if pd.isna(pe) or pd.isna(fpe) or pe <= 0:
            return 2
        r = fpe / pe
        return 5 if r < 0.65 else (4 if r < 0.80 else (3 if r < 0.95 else (2 if r < 1 else 0)))
    df["forward_pe_score"] = df.apply(lambda r: fpe_score({
        "pe_ratio": r.get("pe_ratio"), "eps_forward": r.get("eps_forward")}), axis=1)
    df["growth_score"] = df[["eps_growth_score","rev_growth_score","forward_pe_score"]].sum(axis=1)

    # Safety (max 15)
    df["curr_ratio_score"] = df["current_ratio"].apply(lambda v: 2 if pd.isna(v) else (
        5 if 2.0 <= v <= 4.0 else (4 if 1.5 <= v < 2.0 else (3 if 1.2 <= v < 1.5 else (2 if 1.0 <= v < 1.2 else 0)))))
    df["dividend_score"] = df["dividend_yield_pct"].apply(lambda v: 0 if (pd.isna(v) or v <= 0) else (
        5 if v >= 5 else (4 if v >= 3 else (3 if v >= 2 else (2 if v >= 1 else 1)))))
    df["beta_score"] = df["beta"].apply(lambda v: 2 if pd.isna(v) else (
        5 if 0.6 <= v <= 1.2 else (4 if 0.4 <= v < 0.6 else (3 if 1.2 < v <= 1.5 else (2 if 1.5 < v <= 2.0 else 1)))))
    df["safety_score"] = df[["curr_ratio_score","dividend_score","beta_score"]].sum(axis=1)

    df["pegu_score"] = (df["valuation_score"] + df["quality_score"] +
                        df["growth_score"]   + df["safety_score"])
    df["pegu_percentile"] = df["pegu_score"].rank(pct=True).mul(100).round(1)

    def grade(s):
        if s >= 80: return "A+"
        if s >= 70: return "A"
        if s >= 60: return "B+"
        if s >= 50: return "B"
        if s >= 40: return "C"
        if s >= 25: return "D"
        return "F"
    df["pegu_grade"] = df["pegu_score"].apply(grade)
    return df


# ─────────────────────────────────────────────────────────────
# SARVAS SCAN
# ─────────────────────────────────────────────────────────────
def run_sarvas_scan(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    def b(series, cond): return cond.where(series.notna(), other=None)

    df["above_50dma"]  = df.apply(lambda r: bool(r.get("last_price",0) > r.get("ma_50",0)) if pd.notna(r.get("ma_50")) else None, axis=1)
    df["above_200dma"] = df.apply(lambda r: bool(r.get("last_price",0) > r.get("ma_200",0)) if pd.notna(r.get("ma_200")) else None, axis=1)
    df["golden_cross"] = df.apply(lambda r: bool(r.get("ma_50",0) > r.get("ma_200",0)) if (pd.notna(r.get("ma_50")) and pd.notna(r.get("ma_200"))) else None, axis=1)
    df["rsi_neutral"]  = df["rsi_14"].apply(lambda v: (35 <= v <= 68) if pd.notna(v) else None)
    df["rsi_oversold"] = df["rsi_14"].apply(lambda v: (v < 35) if pd.notna(v) else None)
    df["macd_bullish"] = df.apply(lambda r: bool(r.get("macd",0) > 0) if pd.notna(r.get("macd")) else None, axis=1)
    df["vol_surge"]    = df["volume_ratio"].apply(lambda v: (v > 1.5) if pd.notna(v) else None)
    df["near_52w_high"]= df.apply(lambda r: bool(r.get("last_price",0) / r.get("w52_high",1) >= 0.80) if pd.notna(r.get("w52_high")) else None, axis=1)

    df["positive_eps"]  = df["eps_trailing"].apply(lambda v: v > 0 if pd.notna(v) else False)
    df["eps_growing"]   = df["earnings_growth_pct"].apply(lambda v: v > 10 if pd.notna(v) else False)
    df["rev_growing"]   = df["revenue_growth_pct"].apply(lambda v: v > 8 if pd.notna(v) else False)
    df["reasonable_pe"] = df["pe_ratio"].apply(lambda v: (0 < v < 35) if pd.notna(v) else False)
    df["low_debt"]      = df["debt_equity"].apply(lambda v: v < 1 if pd.notna(v) else False)
    df["strong_roe"]    = df["roe_pct"].apply(lambda v: v > 15 if pd.notna(v) else False)
    df["analyst_upside"]= df["upside_pct"].apply(lambda v: v > 10 if pd.notna(v) else False)

    def sarvas(r):
        return (
            int(bool(r.get("above_50dma")))  * 10 +
            int(bool(r.get("above_200dma"))) * 10 +
            int(bool(r.get("golden_cross"))) *  8 +
            int(bool(r.get("rsi_neutral")))  *  7 +
            int(bool(r.get("macd_bullish"))) *  5 +
            int(bool(r.get("vol_surge")))    *  3 +
            int(bool(r.get("near_52w_high")))*  2 +
            int(r["positive_eps"])  *  8 +
            int(r["eps_growing"])   *  7 +
            int(r["rev_growing"])   *  6 +
            int(r["reasonable_pe"]) *  5 +
            int(r["low_debt"])      *  5 +
            int(r["strong_roe"])    *  5 +
            int(r["analyst_upside"])*  4 +
            int(min(15, max(0, int((r["pegu_score"] - 40) * 15 / 60))))
        )
    df["sarvas_score"] = df.apply(sarvas, axis=1).clip(0, 100)

    def signal(s):
        if s >= 80: return "STRONG BUY"
        if s >= 65: return "BUY"
        if s >= 50: return "ACCUMULATE"
        if s >= 35: return "HOLD"
        if s >= 20: return "REDUCE"
        return "SELL"
    df["sarvas_signal"] = df["sarvas_score"].apply(signal)

    df["sarvas_pass"] = (
        df["above_50dma"].fillna(False).astype(bool) &
        df["above_200dma"].fillna(False).astype(bool) &
        df["rsi_neutral"].fillna(False).astype(bool) &
        (df["pegu_score"] >= 60) &
        df["positive_eps"] &
        df["reasonable_pe"]
    )
    return df


# ─────────────────────────────────────────────────────────────
# EXCEL STYLES
# ─────────────────────────────────────────────────────────────
DARK_BG   = "1F2D3D"
HEADER_FG = "FFFFFF"
ACCENT    = "2196F3"
GREEN     = "27AE60"
YELLOW    = "F39C12"
RED       = "E74C3C"
LIGHT_BG  = "F5F7FA"
GRADE_COLORS = {
    "A+": "1A7A4A", "A": "27AE60", "B+": "5DADE2",
    "B": "F39C12",  "C": "E67E22", "D": "E74C3C", "F": "922B21",
}
SIGNAL_COLORS = {
    "STRONG BUY": "0D5016", "BUY": "27AE60", "ACCUMULATE": "5DADE2",
    "HOLD": "F39C12",        "REDUCE": "E67E22",  "SELL": "E74C3C",
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _thin_border():
    s = Side(style="thin", color="D5D8DC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def write_header_row(ws, row: int, cols: List[str]):
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.fill    = _fill(DARK_BG)
        cell.font    = _font(bold=True, color=HEADER_FG, size=10)
        cell.border  = _thin_border()
        cell.alignment = _center()

def write_data_row(ws, row: int, values: List, bold=False):
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill      = _fill("FFFFFF") if row % 2 == 0 else _fill(LIGHT_BG)
        cell.font      = _font(bold=bold, size=10)
        cell.border    = _thin_border()
        cell.alignment = _left()

def color_cell(cell, hex_color, fg="FFFFFF"):
    cell.fill = _fill(hex_color)
    cell.font = _font(bold=True, color=fg, size=10)

def auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        length = max(
            (len(str(cell.value or "")) for cell in col), default=8
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def add_autofilter(ws, header_row=1):
    ws.auto_filter.ref = ws.dimensions

def freeze(ws, cell="B2"):
    ws.freeze_panes = cell


# ─────────────────────────────────────────────────────────────
# SHEET WRITERS
# ─────────────────────────────────────────────────────────────
def write_summary_sheet(wb, df: pd.DataFrame, scan_info: dict):
    ws = wb.create_sheet("📊 Summary", 0)
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:H1")
    ws["A1"] = f"NSE / BSE Daily Pegu + Sarvas Scan  —  {TODAY_FMT}"
    ws["A1"].font   = Font(bold=True, size=16, color=HEADER_FG, name="Calibri")
    ws["A1"].fill   = _fill(DARK_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Run at {datetime.now().strftime('%H:%M:%S')}  |  Source: nsepython + yfinance  |  Scoring: Pegu v1.0 + Sarvas v1.0"
    ws["A2"].font      = _font(color="555555", size=10)
    ws["A2"].alignment = _center()
    ws["A2"].fill      = _fill("EAF0FB")

    # KPI cards (row 4 onwards)
    kpis = [
        ("Total Stocks Scanned", f"{len(df):,}", DARK_BG),
        ("NSE Stocks",  f"{(df.exchange=='NSE').sum():,}", "1A5276"),
        ("BSE Stocks",  f"{(df.exchange=='BSE').sum():,}", "1A5276"),
        ("Avg Pegu Score", f"{df.pegu_score.mean():.1f} / 100", "27AE60" if df.pegu_score.mean() >= 50 else "E74C3C"),
        ("Sarvas Pass",  f"{df.sarvas_pass.sum():,}", "0D5016"),
        ("STRONG BUY",  f"{(df.sarvas_signal=='STRONG BUY').sum():,}", "0D5016"),
        ("BUY",         f"{(df.sarvas_signal=='BUY').sum():,}",        "27AE60"),
        ("SELL/REDUCE", f"{df.sarvas_signal.isin(['SELL','REDUCE']).sum():,}", "E74C3C"),
    ]
    for i, (label, val, color) in enumerate(kpis):
        col = (i % 4) * 2 + 1
        r   = 4 + (i // 4) * 3
        ws.merge_cells(start_row=r,   start_column=col, end_row=r,   end_column=col+1)
        ws.merge_cells(start_row=r+1, start_column=col, end_row=r+1, end_column=col+1)
        c1, c2 = ws.cell(r, col, value=label), ws.cell(r+1, col, value=val)
        c1.fill, c1.font, c1.alignment = _fill(color), _font(bold=True, color="FFFFFF", size=10), _center()
        c2.fill, c2.font, c2.alignment = _fill(color), _font(bold=True, color="FFFFFF", size=14), _center()
        ws.row_dimensions[r].height   = 20
        ws.row_dimensions[r+1].height = 30

    # Grade distribution
    r0 = 12
    ws.merge_cells(f"A{r0}:H{r0}")
    ws[f"A{r0}"] = "Pegu Grade Distribution"
    ws[f"A{r0}"].font = _font(bold=True, size=12, color=DARK_BG)
    ws[f"A{r0}"].alignment = _left()
    r0 += 1

    grade_cols = ["Grade", "Count", "% of Universe", "Avg Sarvas", "Buy Signals"]
    write_header_row(ws, r0, grade_cols)
    r0 += 1
    for grade, grp in df.groupby("pegu_grade", sort=False):
        vals = [grade, len(grp), f"{len(grp)/len(df)*100:.1f}%",
                f"{grp.sarvas_score.mean():.0f}", (grp.sarvas_signal.isin(["STRONG BUY","BUY"])).sum()]
        write_data_row(ws, r0, vals)
        color_cell(ws.cell(r0, 1), GRADE_COLORS.get(grade, "555555"))
        r0 += 1

    # Signal distribution
    r0 += 1
    ws.merge_cells(f"A{r0}:H{r0}")
    ws[f"A{r0}"] = "Sarvas Signal Distribution"
    ws[f"A{r0}"].font = _font(bold=True, size=12, color=DARK_BG)
    r0 += 1
    write_header_row(ws, r0, ["Signal", "Count", "% of Universe", "Avg Pegu", "Avg Sarvas"])
    r0 += 1
    for sig in ["STRONG BUY","BUY","ACCUMULATE","HOLD","REDUCE","SELL"]:
        grp = df[df.sarvas_signal == sig]
        vals = [sig, len(grp), f"{len(grp)/len(df)*100:.1f}%",
                f"{grp.pegu_score.mean():.0f}" if len(grp) else "—",
                f"{grp.sarvas_score.mean():.0f}" if len(grp) else "—"]
        write_data_row(ws, r0, vals)
        color_cell(ws.cell(r0, 1), SIGNAL_COLORS.get(sig, "555555"))
        r0 += 1

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18
    ws.column_dimensions["A"].width = 26


def _df_to_sheet(wb, sheet_name: str, df: pd.DataFrame,
                 display_cols: List[str], col_labels: List[str],
                 signal_col: Optional[str] = None,
                 grade_col: Optional[str] = None):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    write_header_row(ws, 1, col_labels)
    freeze(ws, "B2")

    for r, (_, row) in enumerate(df[display_cols].iterrows(), 2):
        values = [row[c] if c in row.index else "" for c in display_cols]
        write_data_row(ws, r, values)

        # Color Sarvas signal cell
        if signal_col and signal_col in display_cols:
            sc = display_cols.index(signal_col) + 1
            sig_val = str(row.get(signal_col, ""))
            if sig_val in SIGNAL_COLORS:
                color_cell(ws.cell(r, sc), SIGNAL_COLORS[sig_val])

        # Color Pegu grade cell
        if grade_col and grade_col in display_cols:
            gc = display_cols.index(grade_col) + 1
            gv = str(row.get(grade_col, ""))
            if gv in GRADE_COLORS:
                color_cell(ws.cell(r, gc), GRADE_COLORS[gv])

    add_autofilter(ws)
    auto_width(ws)
    ws.freeze_panes = "B2"
    return ws


def write_all_stocks_sheet(wb, df: pd.DataFrame):
    cols = ["symbol","exchange","company_name","sector","last_price",
            "pe_ratio","pb_ratio","roe_pct","operating_margins_pct",
            "debt_equity","revenue_growth_pct","earnings_growth_pct",
            "pegu_score","pegu_grade","pegu_percentile",
            "sarvas_score","sarvas_signal","upside_pct","analyst_upside"]
    labels = ["Symbol","Exch","Company","Sector","Price (₹)",
              "P/E","P/B","ROE %","Op.Margin %",
              "D/E","Rev.Grwth %","EPS.Grwth %",
              "Pegu","Grade","Pctl",
              "Sarvas","Signal","Upside %","Analyst↑"]
    avail = [c for c in cols if c in df.columns]
    avail_labels = [labels[cols.index(c)] for c in avail]
    sub = df.sort_values("pegu_score", ascending=False).reset_index(drop=True)
    _df_to_sheet(wb, "All Stocks (Pegu)", sub, avail, avail_labels,
                 signal_col="sarvas_signal", grade_col="pegu_grade")


def write_sarvas_pass_sheet(wb, df: pd.DataFrame):
    sub = df[df["sarvas_pass"]].sort_values("sarvas_score", ascending=False).reset_index(drop=True)
    cols = ["symbol","exchange","company_name","sector","last_price",
            "pegu_score","pegu_grade","sarvas_score","sarvas_signal",
            "above_50dma","above_200dma","golden_cross","rsi_14",
            "rsi_neutral","macd_bullish","volume_ratio",
            "pe_ratio","roe_pct","debt_equity","upside_pct"]
    labels = ["Symbol","Exch","Company","Sector","Price (₹)",
              "Pegu","Grade","Sarvas","Signal",
              ">50DMA",">200DMA","Golden×","RSI",
              "RSI OK","MACD↑","Vol Ratio",
              "P/E","ROE %","D/E","Upside %"]
    avail = [c for c in cols if c in sub.columns]
    avail_labels = [labels[cols.index(c)] for c in avail]
    _df_to_sheet(wb, "🎯 Sarvas Pass", sub, avail, avail_labels,
                 signal_col="sarvas_signal", grade_col="pegu_grade")


def write_top_picks_sheet(wb, df: pd.DataFrame, top_n: int):
    sub = df[df["positive_eps"]].sort_values(
        ["pegu_score","sarvas_score"], ascending=False
    ).head(top_n).reset_index(drop=True)
    sub.insert(0, "Rank", range(1, len(sub)+1))
    cols = ["Rank","symbol","exchange","company_name","sector","last_price",
            "pe_ratio","roe_pct","debt_equity","earnings_growth_pct",
            "pegu_score","pegu_grade","sarvas_score","sarvas_signal",
            "target_price","upside_pct","rsi_14","return_3m_pct","return_6m_pct"]
    labels = ["Rank","Symbol","Exch","Company","Sector","Price (₹)",
              "P/E","ROE %","D/E","EPS Grwth %",
              "Pegu","Grade","Sarvas","Signal",
              "Target ₹","Upside %","RSI","Ret 3M %","Ret 6M %"]
    avail = [c for c in cols if c in sub.columns]
    avail_labels = [labels[cols.index(c)] for c in avail]
    _df_to_sheet(wb, f"🏆 Top {top_n} Picks", sub, avail, avail_labels,
                 signal_col="sarvas_signal", grade_col="pegu_grade")


def write_exchange_sheet(wb, df: pd.DataFrame, exchange: str):
    sub = df[df["exchange"] == exchange].sort_values("pegu_score", ascending=False).reset_index(drop=True)
    if sub.empty:
        return
    cols = ["symbol","company_name","sector","last_price",
            "pe_ratio","pb_ratio","peg_ratio","roe_pct","debt_equity",
            "pegu_score","pegu_grade","sarvas_score","sarvas_signal",
            "return_1m_pct","return_3m_pct","upside_pct"]
    labels = ["Symbol","Company","Sector","Price (₹)",
              "P/E","P/B","PEG","ROE %","D/E",
              "Pegu","Grade","Sarvas","Signal",
              "Ret 1M %","Ret 3M %","Upside %"]
    avail = [c for c in cols if c in sub.columns]
    avail_labels = [labels[cols.index(c)] for c in avail]
    icon = "📈"
    _df_to_sheet(wb, f"{icon} {exchange}", sub, avail, avail_labels,
                 signal_col="sarvas_signal", grade_col="pegu_grade")


def write_sector_sheet(wb, df: pd.DataFrame):
    ws = wb.create_sheet("🏭 Sector Analysis")
    ws.sheet_view.showGridLines = False

    sect = (df.groupby(["sector","exchange"]).agg(
        n_stocks       = ("symbol","count"),
        avg_pegu       = ("pegu_score","mean"),
        median_pegu    = ("pegu_score","median"),
        avg_sarvas     = ("sarvas_score","mean"),
        n_buy          = ("sarvas_signal", lambda x: x.isin(["STRONG BUY","BUY"]).sum()),
        n_sarvas_pass  = ("sarvas_pass","sum"),
        avg_pe         = ("pe_ratio","mean"),
        avg_roe        = ("roe_pct","mean"),
        avg_rev_growth = ("revenue_growth_pct","mean"),
    ).reset_index().sort_values("avg_pegu", ascending=False))

    for col in ["avg_pegu","median_pegu","avg_sarvas","avg_pe","avg_roe","avg_rev_growth"]:
        sect[col] = sect[col].round(1)

    headers = ["Sector","Exchange","Stocks","Avg Pegu","Med Pegu","Avg Sarvas",
               "BUY signals","Sarvas Pass","Avg P/E","Avg ROE %","Avg Rev.Grwth %"]
    write_header_row(ws, 1, headers)
    freeze(ws, "B2")

    for r, (_, row) in enumerate(sect.iterrows(), 2):
        vals = [row["sector"],row["exchange"],row["n_stocks"],
                row["avg_pegu"],row["median_pegu"],row["avg_sarvas"],
                row["n_buy"],row["n_sarvas_pass"],
                row["avg_pe"],row["avg_roe"],row["avg_rev_growth"]]
        write_data_row(ws, r, vals)
        # color avg_pegu cell
        score = row["avg_pegu"]
        color = "1A7A4A" if score>=70 else ("27AE60" if score>=60 else ("F39C12" if score>=50 else "E74C3C"))
        color_cell(ws.cell(r, 4), color)

    add_autofilter(ws)
    auto_width(ws)


def write_score_breakdown_sheet(wb, df: pd.DataFrame, top_n: int = 30):
    ws = wb.create_sheet("🔬 Score Breakdown")
    ws.sheet_view.showGridLines = False
    sub = df.nlargest(top_n, "pegu_score").reset_index(drop=True)

    headers = ["Symbol","Exchange","Company","Pegu",
               "Valuation /30","Quality /30","Growth /25","Safety /15",
               "P/E pts","PEG pts","P/B pts",
               "ROE pts","Margin pts","Debt pts",
               "EPS Grwth pts","Rev Grwth pts","Fwd PE pts",
               "CurrRatio pts","Dividend pts","Beta pts"]
    col_keys = ["symbol","exchange","company_name","pegu_score",
                "valuation_score","quality_score","growth_score","safety_score",
                "pe_score","peg_score","pb_score",
                "roe_score","margin_score","debt_score",
                "eps_growth_score","rev_growth_score","forward_pe_score",
                "curr_ratio_score","dividend_score","beta_score"]

    write_header_row(ws, 1, headers)
    freeze(ws, "D2")
    for r, (_, row) in enumerate(sub.iterrows(), 2):
        vals = [row.get(c, "") for c in col_keys]
        write_data_row(ws, r, vals)
        # Heatmap the component scores
        for ci, col in enumerate(col_keys[4:], 5):
            val = row.get(col, 0) or 0
            max_v = 30 if ci <= 6 else (25 if ci == 7 else (15 if ci == 8 else 10))
            pct = val / max_v if max_v > 0 else 0
            c = "27AE60" if pct >= 0.8 else ("F39C12" if pct >= 0.5 else "E74C3C")
            ws.cell(r, ci+4).fill = _fill(c + "44")

    add_autofilter(ws)
    auto_width(ws)


def write_benchmark_sheet(wb):
    bench_csv = "reports/benchmark_results.csv"
    ws = wb.create_sheet("⚡ Benchmark")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "Pure Python vs Library Stack — Timing Benchmark"
    ws["A1"].font      = _font(bold=True, size=13, color=HEADER_FG)
    ws["A1"].fill      = _fill(DARK_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    if not os.path.exists(bench_csv):
        ws["A3"] = "Run nse_bse_benchmark.py to generate timing data."
        return

    bdf = pd.read_csv(bench_csv)
    write_header_row(ws, 3, ["Section","Approach / Metric","Value","Unit"])
    r = 4
    for _, row in bdf.iterrows():
        unit = "seconds" if "import" in str(row.get("section","")) else (
               "µs" if row.get("section") == "parse" else "ms")
        write_data_row(ws, r, [row.get("section",""), row.get("approach",""),
                                round(float(row.get("value",0) or 0), 6), unit])
        r += 1

    add_autofilter(ws)
    auto_width(ws)

    ws["A2"] = (
        "Key insight: Pure Python wins import time (26× faster) and small-array math. "
        "NumPy/pandas wins at ≥10k price points (2–5×). "
        "Requests.Session is fastest for live API fetch."
    )
    ws["A2"].font      = _font(size=9, color="444444")
    ws["A2"].alignment = _left()


def write_metadata_sheet(wb, df: pd.DataFrame):
    ws = wb.create_sheet("ℹ️ Scan Info")
    ws.sheet_view.showGridLines = False

    info = [
        ("Scan Date", TODAY_FMT),
        ("Run Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Stocks", len(df)),
        ("NSE Stocks", int((df.exchange=="NSE").sum())),
        ("BSE Stocks", int((df.exchange=="BSE").sum())),
        ("Sectors Covered", df["sector"].nunique() if "sector" in df.columns else "N/A"),
        ("Data Source", "nsepython + yfinance (representative sample when live unavailable)"),
        ("",""),
        ("PEGU SCORE METHODOLOGY",""),
        ("Valuation Component (max 30)", "P/E (10) + PEG (10) + P/B (10)"),
        ("Quality Component (max 30)",   "ROE (10) + Operating Margin (10) + Debt/Equity (10)"),
        ("Growth Component (max 25)",    "EPS Growth (10) + Revenue Growth (10) + Fwd P/E discount (5)"),
        ("Safety Component (max 15)",    "Current Ratio (5) + Dividend Yield (5) + Beta (5)"),
        ("Grade thresholds",             "A+≥80 | A≥70 | B+≥60 | B≥50 | C≥40 | D≥25 | F<25"),
        ("",""),
        ("SARVAS SCAN METHODOLOGY",""),
        ("Technical signals (max 45)",   "Above 50DMA(10) + Above 200DMA(10) + Golden Cross(8) + RSI neutral(7) + MACD bullish(5) + Vol surge(3) + Near 52W high(2)"),
        ("Fundamental signals (max 55)", "Positive EPS(8) + EPS growing>10%(7) + Rev growing>8%(6) + Reasonable PE(5) + Low debt(5) + Strong ROE>15%(5) + Analyst upside(4) + Pegu bonus(5-15)"),
        ("Strict Sarvas Pass",           "Must be: above 50&200 DMA + RSI neutral(35-68) + Pegu≥60 + Positive EPS + PE<35"),
        ("Signal levels",                "STRONG BUY≥80 | BUY≥65 | ACCUMULATE≥50 | HOLD≥35 | REDUCE≥20 | SELL<20"),
    ]

    ws.merge_cells("A1:C1")
    ws["A1"] = "Daily NSE/BSE Scan — Methodology & Metadata"
    ws["A1"].font = _font(bold=True, size=14, color=HEADER_FG)
    ws["A1"].fill = _fill(DARK_BG)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 30

    for r, (key, val) in enumerate(info, 3):
        kc, vc = ws.cell(r, 1, value=key), ws.cell(r, 2, value=val)
        if key and not val:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            kc.font = _font(bold=True, size=11, color=DARK_BG)
            kc.fill = _fill("EAF0FB")
        else:
            kc.font = _font(bold=True, size=10, color=DARK_BG)
            vc.font = _font(size=10)
        kc.alignment = _left()
        vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 22

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 80


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Daily NSE/BSE Scan → Excel")
    ap.add_argument("--index",     default="NIFTY500")
    ap.add_argument("--top-n",     type=int, default=50)
    ap.add_argument("--data-dir",  default="data")
    ap.add_argument("--out-dir",   default="reports")
    ap.add_argument("--mock",      action="store_true",
                    help="Force representative sample data (skip live fetch)")
    ap.add_argument("--no-fetch",  action="store_true",
                    help="Skip extractor, use existing CSVs only")
    args = ap.parse_args()

    os.makedirs(args.data_dir,  exist_ok=True)
    os.makedirs(args.out_dir,   exist_ok=True)
    out_file = os.path.join(args.out_dir, f"NSE_BSE_Daily_Scan_{datetime.now().strftime('%Y%m%d')}.xlsx")

    # ── 1. Optional extractor pre-run (batch pipeline) ────────
    if not args.mock and not args.no_fetch:
        print(f"[1/4] Running nse_bse_extractor.py (index={args.index})…")
        result = subprocess.run(
            [sys.executable, "nse_bse_extractor.py",
             "--exchange", "BOTH", "--index", args.index,
             "--batch-size", "50", "--delay", "0.3",
             "--output-dir", args.data_dir],
            capture_output=True, text=True
        )
        if result.returncode != 0:
            print(f"  [WARN] Extractor exited {result.returncode}; "
                  f"load_or_generate_data will fall back to nsepython+bsedata or sample data")
    else:
        print("[1/4] Skipping extractor pre-run (--mock or --no-fetch)")

    # ── 2. Load + score ───────────────────────────────────────
    # Cascade: extractor CSV → live nsepython+bsedata → sample fallback
    print("[2/4] Loading data and applying Pegu + Sarvas scoring…")
    df_raw = load_or_generate_data(args.data_dir, args.mock)
    df_raw = df_raw.drop_duplicates(subset=["symbol","exchange"]).reset_index(drop=True)

    # Guard: ensure all *_pct columns exist (normalise_live covers live path;
    # this catches any edge-case CSV that may lack them)
    for src, dst in [("roe","roe_pct"),("operating_margins","operating_margins_pct"),
                     ("revenue_growth","revenue_growth_pct"),
                     ("earnings_growth","earnings_growth_pct"),
                     ("dividend_yield","dividend_yield_pct")]:
        if dst not in df_raw.columns:
            if src in df_raw.columns:
                df_raw[dst] = df_raw[src].apply(
                    lambda v: v * 100 if (pd.notna(v) and abs(v) < 2) else v)
            else:
                df_raw[dst] = np.nan
    for col in ("upside_pct", "volume_ratio", "rsi_14", "macd"):
        if col not in df_raw.columns:
            df_raw[col] = np.nan

    df_scored  = calculate_pegu_scores(df_raw)
    df_final   = run_sarvas_scan(df_scored)

    n_pass = df_final["sarvas_pass"].sum()
    n_buy  = df_final["sarvas_signal"].isin(["STRONG BUY","BUY"]).sum()
    print(f"  Stocks scored  : {len(df_final):,}")
    print(f"  Sarvas pass    : {n_pass:,}")
    print(f"  BUY/STRONG BUY : {n_buy:,}")
    print(f"  Avg Pegu score : {df_final.pegu_score.mean():.1f}")

    # ── 3. Build Excel ────────────────────────────────────────
    print(f"[3/4] Building Excel workbook…")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_summary_sheet(wb, df_final, {})
    write_all_stocks_sheet(wb, df_final)
    write_sarvas_pass_sheet(wb, df_final)
    write_top_picks_sheet(wb, df_final, args.top_n)
    write_exchange_sheet(wb, df_final, "NSE")
    write_exchange_sheet(wb, df_final, "BSE")
    write_sector_sheet(wb, df_final)
    write_score_breakdown_sheet(wb, df_final, top_n=min(30, len(df_final)))
    write_benchmark_sheet(wb)
    write_metadata_sheet(wb, df_final)

    # ── 4. Save ───────────────────────────────────────────────
    wb.save(out_file)
    size_kb = os.path.getsize(out_file) / 1024
    print(f"[4/4] Saved → {out_file}  ({size_kb:.0f} KB)")
    print(f"\n  Sheets: {len(wb.sheetnames)}")
    for sn in wb.sheetnames:
        print(f"    • {sn}")

    # Quick console summary
    print(f"\n{'─'*60}")
    print(f"  TOP 10 PICKS (by Pegu + Sarvas)  —  {TODAY_FMT}")
    print(f"{'─'*60}")
    cols = ["symbol","exchange","pegu_score","pegu_grade","sarvas_score","sarvas_signal"]
    avail = [c for c in cols if c in df_final.columns]
    top10 = (df_final[df_final["positive_eps"]]
             .sort_values(["pegu_score","sarvas_score"], ascending=False)
             .head(10)[avail])
    print(top10.to_string(index=False))
    print(f"\n  Full results: {out_file}")


if __name__ == "__main__":
    main()
