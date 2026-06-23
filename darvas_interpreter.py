#!/usr/bin/env python3
"""
darvas_interpreter.py
=====================
Darvas Box interpreter — exact entry price, stop loss, and exit targets.

The Nicolas Darvas method (1960):
  · Box TOP    = a high that holds for CONFIRM_DAYS consecutive bars
  · Box BOTTOM = a low that holds for CONFIRM_DAYS consecutive bars within the box
  · BREAKOUT   = close above box top  →  BUY signal
  · BREAKDOWN  = close below box bottom  →  EXIT / AVOID
  · STOP LOSS  = box bottom (or a small buffer below it)
  · ENTRY      = box top + entry buffer (confirmed breakout price)
  · TARGET 1   = entry + 1 × risk  (1:1 R/R)
  · TARGET 2   = entry + 2 × risk  (2:1 R/R)
  · TARGET 3   = entry + 3 × risk  (3:1 R/R)

Box states:
  BREAKOUT_FRESH   — price broke above box top within last 5 bars  → BUY NOW
  BREAKOUT_OLD     — breakout happened >5 bars ago; await pullback  → BUY ON PULLBACK
  NEAR_TOP         — price within 2% of confirmed box top           → PLACE BUY STOP
  IN_BOX           — price inside a confirmed box                   → WAIT
  BREAKDOWN        — price closed below box bottom                  → AVOID / EXIT
  FORMING          — top confirmed, bottom still forming            → MONITOR
  NO_BOX           — no valid box in lookback window               → NO ACTION

Usage (terminal):
    python3 darvas_interpreter.py RELIANCE
    python3 darvas_interpreter.py RELIANCE TCS INFY HDFCBANK
    python3 darvas_interpreter.py RELIANCE --rr 2.5 --confirm 3
    python3 darvas_interpreter.py --file symbols.txt
    python3 darvas_interpreter.py RELIANCE --excel reports/darvas_signals.xlsx

Integration (import):
    from darvas_interpreter import interpret_symbol, interpret_dataframe
    sig = interpret_symbol("RELIANCE")
    df_with_signals = interpret_dataframe(df)   # df has columns from nifty500_scan

Dependencies:
    pip3 install yfinance pandas numpy openpyxl
"""
from __future__ import annotations

import argparse
import math
import sys
import warnings
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

warnings.filterwarnings("ignore")

try:
    import numpy as np
    import pandas as pd
except ImportError:
    sys.exit("pip3 install pandas numpy yfinance openpyxl")

try:
    import yfinance as yf
except ImportError:
    sys.exit("pip3 install yfinance")


# ── configuration defaults ────────────────────────────────────────────────────

CONFIRM_DAYS     = 3      # bars a high/low must hold to confirm box boundary
ENTRY_BUFFER_PCT = 0.004  # 0.4% above box top for entry price
STOP_BUFFER_PCT  = 0.002  # 0.2% below box bottom for hard stop
NEAR_TOP_PCT     = 0.020  # "near top" if within 2% of box top
BREAKOUT_FRESH   = 5      # breakout within this many bars = "fresh"
LOOKBACK_BARS    = 252    # max bars of history to scan for boxes (~1 year)
RR_TARGETS       = [1.0, 2.0, 3.0]  # risk:reward multiples for targets


# ── data structures ───────────────────────────────────────────────────────────

@dataclass
class DarvasBox:
    top:           float
    bottom:        float
    top_bar:       int         # index in the price series
    bottom_bar:    int
    confirmed:     bool        # both top AND bottom confirmed
    breakout_bar:  Optional[int] = None
    breakdown_bar: Optional[int] = None
    width_pct:     float = 0.0

    def __post_init__(self):
        if self.top > 0:
            self.width_pct = round((self.top - self.bottom) / self.top * 100, 2)


@dataclass
class DarvasSignal:
    ticker:        str
    name:          str = ""
    as_of:         str = ""

    # Box geometry
    box_top:       Optional[float] = None
    box_bottom:    Optional[float] = None
    box_width_pct: Optional[float] = None
    box_age_days:  Optional[int]   = None
    box_confirmed: bool = False

    # Prices
    current_price: Optional[float] = None
    entry_price:   Optional[float] = None   # where to BUY
    stop_loss:     Optional[float] = None   # where to EXIT if wrong
    target_1:      Optional[float] = None   # 1:1 R/R
    target_2:      Optional[float] = None   # 2:1 R/R
    target_3:      Optional[float] = None   # 3:1 R/R

    # Risk metrics
    risk_pct:      Optional[float] = None   # (entry - stop) / entry × 100
    reward_1_pct:  Optional[float] = None
    reward_2_pct:  Optional[float] = None
    reward_3_pct:  Optional[float] = None
    rr_ratio:      float = 2.0

    # State
    state:         str = "NO_BOX"   # see module docstring
    action:        str = "NO ACTION"
    notes:         str = ""

    # Supporting indicators
    rsi:           Optional[float] = None
    volume_ratio:  Optional[float] = None   # today vs 20-day avg
    above_ema50:   Optional[bool]  = None
    above_ema200:  Optional[bool]  = None


# ── core box-detection algorithm ──────────────────────────────────────────────

def _find_boxes(
    highs:  list[float],
    lows:   list[float],
    closes: list[float],
    confirm_days: int = CONFIRM_DAYS,
) -> list[DarvasBox]:
    """
    Scan a price series and return all identified Darvas boxes in chronological order.

    A box top is confirmed when a high H is NOT exceeded for confirm_days bars.
    A box bottom is confirmed when a low L is NOT breached for confirm_days bars
    (measured from within the box formation period).
    Scanning moves forward past each breakout; overlapping boxes are not produced.
    """
    n = len(highs)
    boxes: list[DarvasBox] = []
    i = 0

    while i < n - confirm_days:
        H = highs[i]

        # ── Check if highs[i] qualifies as a box TOP ──────────────────────────
        top_holds = all(highs[i + k] < H for k in range(1, confirm_days + 1)
                        if i + k < n)
        if not top_holds:
            i += 1
            continue

        # Top confirmed at bar i
        box_top     = H
        top_bar     = i
        search_from = i + confirm_days  # first bar AFTER confirmation

        # ── Scan forward to find box BOTTOM and eventual outcome ──────────────
        running_low  = min(lows[i: search_from + 1])
        bottom_bar   = i + lows[i: search_from + 1].index(running_low)
        bottom_confirmed = False
        breakout_bar     = None
        breakdown_bar    = None

        j = search_from
        while j < n:
            # Breakout: close clears the box top
            if closes[j] > box_top:
                breakout_bar = j
                break

            # Update running low
            if lows[j] < running_low:
                running_low = lows[j]
                bottom_bar  = j
                bottom_confirmed = False   # reset — new low, re-confirm

            # Check if current running low is confirmed (holds for confirm_days)
            if not bottom_confirmed:
                lo_holds = all(lows[j + k] > running_low for k in range(1, confirm_days + 1)
                               if j + k < n)
                if lo_holds:
                    bottom_confirmed = True

            # Breakdown: close falls below the running low
            if closes[j] < running_low:
                breakdown_bar = j
                break

            j += 1

        box = DarvasBox(
            top=box_top,
            bottom=running_low,
            top_bar=top_bar,
            bottom_bar=bottom_bar,
            confirmed=bottom_confirmed,
            breakout_bar=breakout_bar,
            breakdown_bar=breakdown_bar,
        )
        boxes.append(box)

        # Skip past the breakout/breakdown so we don't double-count
        if breakout_bar is not None:
            i = breakout_bar + 1
        elif breakdown_bar is not None:
            i = breakdown_bar + 1
        else:
            # Box still active (no conclusion yet) — it's the latest box
            break

    return boxes


def _rsi14(closes: list[float], period: int = 14) -> Optional[float]:
    if len(closes) < period + 1:
        return None
    s = pd.Series(closes, dtype=float)
    delta = s.diff().dropna()
    gain = delta.clip(lower=0).ewm(com=period - 1, min_periods=period).mean()
    loss = (-delta).clip(lower=0).ewm(com=period - 1, min_periods=period).mean()
    rs = gain / loss.replace(0, float("nan"))
    return round(float((100 - 100 / (1 + rs)).iloc[-1]), 2)


def _ema(closes: list[float], period: int) -> Optional[float]:
    if len(closes) < period:
        return None
    s = pd.Series(closes, dtype=float)
    return round(float(s.ewm(span=period, adjust=False).mean().iloc[-1]), 2)


# ── signal builder ────────────────────────────────────────────────────────────

def _build_signal(
    ticker:       str,
    hist:         pd.DataFrame,
    info:         dict,
    rr_ratio:     float = 2.0,
    confirm_days: int   = CONFIRM_DAYS,
    entry_buffer: float = ENTRY_BUFFER_PCT,
    stop_buffer:  float = STOP_BUFFER_PCT,
    near_top_pct: float = NEAR_TOP_PCT,
) -> DarvasSignal:

    sig = DarvasSignal(
        ticker    = ticker,
        name      = info.get("longName") or info.get("shortName") or ticker,
        as_of     = datetime.now().strftime("%Y-%m-%d"),
        rr_ratio  = rr_ratio,
    )

    if hist.empty or len(hist) < confirm_days * 3 + 5:
        sig.state  = "NO_BOX"
        sig.action = "INSUFFICIENT DATA"
        return sig

    # Trim to lookback window
    hist = hist.tail(LOOKBACK_BARS).reset_index(drop=True)

    highs  = hist["High"].tolist()
    lows   = hist["Low"].tolist()
    closes = hist["Close"].tolist()
    vols   = hist["Volume"].tolist()

    cmp = closes[-1]
    sig.current_price = round(cmp, 2)

    # Supporting indicators
    sig.rsi         = _rsi14(closes)
    ema50           = _ema(closes, 50)
    ema200          = _ema(closes, 200)
    sig.above_ema50  = (cmp > ema50)  if ema50  is not None else None
    sig.above_ema200 = (cmp > ema200) if ema200 is not None else None

    avg_vol_20 = float(pd.Series(vols).rolling(20).mean().iloc[-1]) if len(vols) >= 20 else None
    sig.volume_ratio = round(vols[-1] / avg_vol_20, 2) if avg_vol_20 else None

    # ── Detect boxes ──────────────────────────────────────────────────────────
    boxes = _find_boxes(highs, lows, closes, confirm_days=confirm_days)

    if not boxes:
        sig.state  = "NO_BOX"
        sig.action = "NO VALID BOX FOUND IN HISTORY"
        return sig

    latest = boxes[-1]   # most recent box
    n      = len(closes)
    last_i = n - 1       # index of today

    sig.box_top       = round(latest.top,    2)
    sig.box_bottom    = round(latest.bottom, 2)
    sig.box_width_pct = latest.width_pct
    sig.box_confirmed = latest.confirmed
    sig.box_age_days  = last_i - latest.top_bar

    risk          = latest.top * entry_buffer           # entry buffer amount
    entry         = round(latest.top * (1 + entry_buffer), 2)
    stop          = round(latest.bottom * (1 - stop_buffer), 2)
    risk_per_unit = entry - stop

    sig.entry_price   = entry
    sig.stop_loss     = stop
    sig.risk_pct      = round((entry - stop) / entry * 100, 2) if entry > 0 else None

    if risk_per_unit > 0:
        sig.target_1     = round(entry + 1.0 * risk_per_unit, 2)
        sig.target_2     = round(entry + 2.0 * risk_per_unit, 2)
        sig.target_3     = round(entry + 3.0 * risk_per_unit, 2)
        sig.reward_1_pct = round((sig.target_1 - entry) / entry * 100, 2)
        sig.reward_2_pct = round((sig.target_2 - entry) / entry * 100, 2)
        sig.reward_3_pct = round((sig.target_3 - entry) / entry * 100, 2)

    # ── Classify state ────────────────────────────────────────────────────────

    # Case 1: Box ended with a confirmed breakout previously → look for next box
    if latest.breakout_bar is not None:
        bars_since_breakout = last_i - latest.breakout_bar
        if bars_since_breakout <= BREAKOUT_FRESH:
            sig.state  = "BREAKOUT_FRESH"
            sig.action = "BUY NOW — Momentum breakout confirmed"
            sig.notes  = (f"Broke out {bars_since_breakout} bar(s) ago  |  "
                          f"Volume ratio: {sig.volume_ratio}")
        else:
            sig.state  = "BREAKOUT_OLD"
            sig.action = "BUY ON PULLBACK — Wait for retest of box top"
            sig.notes  = (f"Breakout {bars_since_breakout} bars ago  |  "
                          f"Pullback entry near ₹{sig.box_top:,.2f}")
        return sig

    # Case 2: Box ended with a breakdown
    if latest.breakdown_bar is not None:
        sig.state  = "BREAKDOWN"
        sig.action = "AVOID / EXIT — Price broke below box floor"
        sig.entry_price = None
        sig.stop_loss   = None
        sig.target_1 = sig.target_2 = sig.target_3 = None
        sig.notes  = f"Broke down below ₹{sig.box_bottom:,.2f}"
        return sig

    # Case 3: Box still active — price is inside or near the box
    if not latest.confirmed:
        sig.state  = "FORMING"
        sig.action = "MONITOR — Box top confirmed, bottom still forming"
        sig.notes  = (f"Box top: ₹{sig.box_top:,.2f}  |  "
                      f"Tentative floor: ₹{sig.box_bottom:,.2f}")
        return sig

    # Box is fully confirmed and price is still within it
    gap_to_top = (latest.top - cmp) / latest.top if latest.top > 0 else 1.0

    if cmp > latest.top:
        # Intraday breakout (close hasn't confirmed yet)
        sig.state  = "BREAKOUT_FRESH"
        sig.action = "BUY NOW — Intraday breakout above box top"
        sig.notes  = (f"Price {cmp:,.2f} > box top {latest.top:,.2f}  |  "
                      f"Volume ratio: {sig.volume_ratio}")
    elif gap_to_top <= near_top_pct:
        sig.state  = "NEAR_TOP"
        sig.action = (f"PLACE BUY STOP at ₹{entry:,.2f} — "
                      f"Price is {gap_to_top*100:.1f}% from breakout")
        sig.notes  = (f"Box width: {sig.box_width_pct:.1f}%  |  "
                      f"Box age: {sig.box_age_days} days")
    else:
        pct_from_bottom = (cmp - latest.bottom) / (latest.top - latest.bottom) * 100 \
                          if (latest.top - latest.bottom) > 0 else 0
        sig.state  = "IN_BOX"
        sig.action = (f"WAIT FOR BREAKOUT — Price is "
                      f"{pct_from_bottom:.0f}% up within the box")
        sig.notes  = (f"Box: ₹{sig.box_bottom:,.2f} – ₹{sig.box_top:,.2f}  |  "
                      f"Age: {sig.box_age_days} days")

    return sig


# ── public API ────────────────────────────────────────────────────────────────

def interpret_symbol(
    ticker:       str,
    confirm_days: int   = CONFIRM_DAYS,
    rr_ratio:     float = 2.0,
    period:       str   = "1y",
) -> DarvasSignal:
    """Fetch live data and return a DarvasSignal for one NSE symbol."""
    yf_ticker = f"{ticker}.NS"
    try:
        t    = yf.Ticker(yf_ticker)
        hist = t.history(period=period, auto_adjust=True)
        info = {}
        try:
            info = t.info
        except Exception:
            pass
    except Exception as exc:
        sig = DarvasSignal(ticker=ticker)
        sig.state  = "NO_BOX"
        sig.action = f"DATA ERROR: {exc}"
        return sig

    return _build_signal(
        ticker, hist, info,
        rr_ratio=rr_ratio,
        confirm_days=confirm_days,
    )


def interpret_dataframe(
    df: pd.DataFrame,
    hist_map: Optional[dict] = None,
    confirm_days: int   = CONFIRM_DAYS,
    rr_ratio:     float = 2.0,
) -> pd.DataFrame:
    """
    Add Darvas interpretation columns to an existing scan DataFrame.

    Parameters
    ----------
    df        : DataFrame with at minimum a 'ticker' column (from nifty500_scan).
    hist_map  : Optional dict of {bare_symbol: pd.DataFrame} from batch_download.
                If provided, no network calls are made.
                If None, data is fetched live per symbol (slow for large lists).
    """
    records = []
    total = len(df)

    for i, (_, row) in enumerate(df.iterrows(), 1):
        sym = str(row.get("ticker", ""))
        print(f"  Darvas interpret {i}/{total}  {sym:<14}", end="\r", flush=True)

        if hist_map and sym in hist_map:
            sig = _build_signal(sym, hist_map[sym], {}, rr_ratio=rr_ratio,
                                confirm_days=confirm_days)
        else:
            sig = interpret_symbol(sym, confirm_days=confirm_days, rr_ratio=rr_ratio)

        records.append({
            "ticker":        sym,
            "dv_state":      sig.state,
            "dv_action":     sig.action,
            "dv_entry":      sig.entry_price,
            "dv_stop":       sig.stop_loss,
            "dv_target1":    sig.target_1,
            "dv_target2":    sig.target_2,
            "dv_target3":    sig.target_3,
            "dv_risk_pct":   sig.risk_pct,
            "dv_rwd1_pct":   sig.reward_1_pct,
            "dv_rwd2_pct":   sig.reward_2_pct,
            "dv_rwd3_pct":   sig.reward_3_pct,
            "dv_box_top":    sig.box_top,
            "dv_box_bottom": sig.box_bottom,
            "dv_box_w_pct":  sig.box_width_pct,
            "dv_box_age":    sig.box_age_days,
            "dv_notes":      sig.notes,
        })

    print()
    interp_df = pd.DataFrame(records)
    return df.merge(interp_df, on="ticker", how="left")


# ── terminal print ────────────────────────────────────────────────────────────

STATE_EMOJI = {
    "BREAKOUT_FRESH": "🚀",
    "BREAKOUT_OLD":   "📈",
    "NEAR_TOP":       "👀",
    "IN_BOX":         "📦",
    "FORMING":        "🔍",
    "BREAKDOWN":      "⛔",
    "NO_BOX":         "—",
}

STATE_COLOUR = {
    "BREAKOUT_FRESH": "\033[92m",  # bright green
    "BREAKOUT_OLD":   "\033[32m",  # green
    "NEAR_TOP":       "\033[33m",  # yellow
    "IN_BOX":         "\033[34m",  # blue
    "FORMING":        "\033[35m",  # purple
    "BREAKDOWN":      "\033[91m",  # bright red
    "NO_BOX":         "\033[90m",  # grey
}
RESET = "\033[0m"


def _p(label: str, value, fmt: str = "", colour: str = "") -> None:
    label_s  = f"  {label:<28}"
    if value is None:
        print(f"{label_s}  —")
        return
    if fmt:
        value_s = format(value, fmt)
    else:
        value_s = str(value)
    print(f"{label_s}  {colour}{value_s}{RESET if colour else ''}")


def print_signal(sig: DarvasSignal) -> None:
    W = 64
    colour = STATE_COLOUR.get(sig.state, "")
    emoji  = STATE_EMOJI.get(sig.state, "")

    print()
    print("━" * W)
    print(f"  DARVAS INTERPRETER  —  {sig.ticker}  ({sig.name})")
    print(f"  As of: {sig.as_of}")
    print("━" * W)

    _p("Current Price (₹)",  sig.current_price, ",.2f")
    _p("RSI (14)",           sig.rsi,           ".1f")
    _p("Volume Ratio",       sig.volume_ratio,  ".2f")
    _p("Above EMA-50",       sig.above_ema50)
    _p("Above EMA-200",      sig.above_ema200)

    print()
    _p("Status",  f"{emoji}  {sig.state}", colour=colour)
    _p("Action",  sig.action,             colour=colour)
    if sig.notes:
        _p("Notes",  sig.notes)

    print()
    print(f"  {'── BOX GEOMETRY ':─<{W-2}}")
    _p("Box Top (₹)",       sig.box_top,       ",.2f")
    _p("Box Bottom (₹)",    sig.box_bottom,     ",.2f")
    _p("Box Width",         f"{sig.box_width_pct:.1f}%" if sig.box_width_pct else None)
    _p("Box Age (days)",    sig.box_age_days)
    _p("Box Confirmed",     sig.box_confirmed)

    print()
    print(f"  {'── TRADE LEVELS ':─<{W-2}}")
    entry_colour = "\033[92m" if sig.entry_price else ""
    stop_colour  = "\033[91m"
    tgt_colour   = "\033[96m"

    _p("Entry Price (₹)",   sig.entry_price,   ",.2f", colour=entry_colour)
    _p("Stop Loss (₹)",     sig.stop_loss,      ",.2f", colour=stop_colour)

    if sig.risk_pct is not None:
        _p("Risk from Entry",  f"-{sig.risk_pct:.2f}%",  colour=stop_colour)

    print()
    _p("Target 1 — 1:1 (₹)",  sig.target_1,  ",.2f", colour=tgt_colour)
    if sig.reward_1_pct is not None:
        _p("  Reward",           f"+{sig.reward_1_pct:.2f}%")

    _p("Target 2 — 2:1 (₹)",  sig.target_2,  ",.2f", colour=tgt_colour)
    if sig.reward_2_pct is not None:
        _p("  Reward",           f"+{sig.reward_2_pct:.2f}%")

    _p("Target 3 — 3:1 (₹)",  sig.target_3,  ",.2f", colour=tgt_colour)
    if sig.reward_3_pct is not None:
        _p("  Reward",           f"+{sig.reward_3_pct:.2f}%")

    print("━" * W)


# ── Excel export ──────────────────────────────────────────────────────────────

def write_signals_excel(signals: list[DarvasSignal], xl_path: Path) -> None:
    """Write all signals to a date-wise sheet in an Excel workbook."""
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("pip3 install openpyxl")

    as_of      = datetime.now()
    sheet_name = as_of.strftime("%d-%b-%y")

    if xl_path.exists():
        wb = load_workbook(xl_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, 0)
    else:
        wb = Workbook()
        ws  = wb.active
        ws.title = sheet_name

    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color="999999")
    bdr   = Border(left=thin, right=thin, top=thin,  bottom=thin)
    bdr_h = Border(left=thin, right=thin, top=thick, bottom=thick)

    C = {
        "dark":     "0D2137",
        "navy":     "1E3A5F",
        "mid":      "2B4C7E",
        "fresh":    "C6EFCE", "fresh_fg":  "276221",
        "old":      "DDFFDD", "old_fg":    "276221",
        "near":     "FFEB9C", "near_fg":   "9C6500",
        "inbox":    "DDEEFF", "inbox_fg":  "003366",
        "forming":  "EEE0FF", "forming_fg":"4B0082",
        "breakdown":"FFC7CE", "brkdn_fg":  "9C0006",
        "nobox":    "F0F0F0", "nobox_fg":  "888888",
        "entry":    "276221",
        "stop":     "9C0006",
        "target":   "003366",
        "alt":      "F5F8FF",
        "white":    "FFFFFF",
        "hdr_fg":   "FFFFFF",
    }

    STATE_FILL = {
        "BREAKOUT_FRESH": (C["fresh"],    C["fresh_fg"]),
        "BREAKOUT_OLD":   (C["old"],      C["old_fg"]),
        "NEAR_TOP":       (C["near"],     C["near_fg"]),
        "IN_BOX":         (C["inbox"],    C["inbox_fg"]),
        "FORMING":        (C["forming"],  C["forming_fg"]),
        "BREAKDOWN":      (C["breakdown"],C["brkdn_fg"]),
        "NO_BOX":         (C["nobox"],    C["nobox_fg"]),
    }

    def fill(h): return PatternFill("solid", fgColor=h)
    def fnt(bold=False, color="000000", size=10, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    def mono(color="000000", size=10):
        return Font(name="Courier New", color=color, size=size)
    def aln(h="left", v="center"):
        return Alignment(horizontal=h, vertical=v, wrap_text=False)

    # ── Title ─────────────────────────────────────────────────────────────────
    N_COLS = 19
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    c = ws.cell(1, 1, f"Darvas Box Interpreter  —  {as_of.strftime('%A, %d %B %Y  %H:%M')}")
    c.font = fnt(bold=True, size=13, color=C["hdr_fg"])
    c.fill = fill(C["dark"])
    c.alignment = aln("center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{get_column_letter(N_COLS)}2")
    c = ws.cell(2, 1, "Entry = box top + 0.4%  |  Stop = box bottom − 0.2%  |  "
                       "T1 = 1:1 R/R  |  T2 = 2:1  |  T3 = 3:1")
    c.font = fnt(italic=True, size=9, color="AAAAAA")
    c.fill = fill(C["dark"])
    c.alignment = aln("center")
    ws.row_dimensions[2].height = 14

    # ── Column headers ────────────────────────────────────────────────────────
    HDRS = [
        ("Ticker",      10), ("Name",         26), ("Status",      16),
        ("Action",      36), ("CMP (₹)",       10), ("Entry (₹)",   10),
        ("Stop (₹)",    10), ("Risk %",         8),
        ("Target 1 (₹)",11), ("Rwd 1 %",        8),
        ("Target 2 (₹)",11), ("Rwd 2 %",        8),
        ("Target 3 (₹)",11), ("Rwd 3 %",        8),
        ("Box Top (₹)", 10), ("Box Bot (₹)",   10),
        ("Box W %",      8), ("Box Age",         8),
        ("Notes",       40),
    ]
    for col_i, (h, w) in enumerate(HDRS, 1):
        c = ws.cell(3, col_i, h)
        c.font = fnt(bold=True, color=C["hdr_fg"], size=9)
        c.fill = fill(C["navy"])
        c.alignment = aln("center")
        c.border = bdr_h
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[3].height = 18

    # Sort: BREAKOUT_FRESH → BREAKOUT_OLD → NEAR_TOP → IN_BOX → FORMING → rest
    ORDER = {"BREAKOUT_FRESH": 0, "BREAKOUT_OLD": 1, "NEAR_TOP": 2,
             "IN_BOX": 3, "FORMING": 4, "BREAKDOWN": 5, "NO_BOX": 6}
    signals_sorted = sorted(signals, key=lambda s: ORDER.get(s.state, 9))

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_i, sig in enumerate(signals_sorted, 4):
        bg_state, fg_state = STATE_FILL.get(sig.state, (C["white"], "000000"))
        alt_bg = C["alt"] if row_i % 2 == 0 else C["white"]

        def cell(col, val, fmt=None, colour="000000", bold=False, bg=alt_bg, mno=False):
            c = ws.cell(row_i, col, val)
            c.font = mono(color=colour) if mno else fnt(bold=bold, color=colour)
            c.fill = fill(bg)
            c.alignment = aln("right" if col > 4 and col != 4 else "left")
            c.border = bdr
            if fmt: c.number_format = fmt
            return c

        # Ticker
        cell(1,  sig.ticker,        bold=True)
        # Name
        cell(2,  sig.name or "")
        # Status (coloured by state)
        c = ws.cell(row_i, 3, sig.state)
        c.font = fnt(bold=True, color=fg_state)
        c.fill = fill(bg_state)
        c.alignment = aln("center")
        c.border = bdr
        # Action
        cell(4,  sig.action or "")
        # CMP
        cell(5,  sig.current_price, "#,##0.00", mno=True)
        # Entry
        cell(6,  sig.entry_price,   "#,##0.00", colour=C["entry"], bold=True, mno=True)
        # Stop
        cell(7,  sig.stop_loss,     "#,##0.00", colour=C["stop"],  bold=True, mno=True)
        # Risk %
        cell(8,  f"-{sig.risk_pct:.2f}%" if sig.risk_pct else None,
             colour=C["stop"], mno=True)
        # T1 / Rwd1
        cell(9,  sig.target_1,  "#,##0.00", colour=C["target"], mno=True)
        cell(10, f"+{sig.reward_1_pct:.2f}%" if sig.reward_1_pct else None,
             colour=C["target"], mno=True)
        # T2 / Rwd2
        cell(11, sig.target_2,  "#,##0.00", colour=C["target"], mno=True)
        cell(12, f"+{sig.reward_2_pct:.2f}%" if sig.reward_2_pct else None,
             colour=C["target"], mno=True)
        # T3 / Rwd3
        cell(13, sig.target_3,  "#,##0.00", colour=C["target"], mno=True)
        cell(14, f"+{sig.reward_3_pct:.2f}%" if sig.reward_3_pct else None,
             colour=C["target"], mno=True)
        # Box geometry
        cell(15, sig.box_top,       "#,##0.00", mno=True)
        cell(16, sig.box_bottom,    "#,##0.00", mno=True)
        cell(17, f"{sig.box_width_pct:.1f}%" if sig.box_width_pct else None, mno=True)
        cell(18, sig.box_age_days)
        # Notes
        c = ws.cell(row_i, 19, sig.notes or "")
        c.font = fnt(italic=True, size=9, color="555555")
        c.fill = fill(alt_bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        c.border = bdr

        ws.row_dimensions[row_i].height = 15

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(N_COLS)}{3 + len(signals_sorted)}"

    xl_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xl_path)
    print(f"\n  ✓  Saved → {xl_path}  (sheet: {sheet_name})")

    # Count by state
    from collections import Counter
    counts = Counter(s.state for s in signals)
    for state, n in sorted(counts.items(), key=lambda x: ORDER.get(x[0], 9)):
        print(f"     {state:<18}  {n}")


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Darvas Box interpreter — entry, stop loss, exit targets",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("symbols", nargs="*", metavar="SYMBOL",
                        help="NSE symbols (e.g. RELIANCE TCS INFY)")
    parser.add_argument("--file", metavar="PATH",
                        help="Text file with one symbol per line")
    parser.add_argument("--confirm", type=int, default=CONFIRM_DAYS, metavar="N",
                        help=f"Days a high/low must hold to confirm box (default {CONFIRM_DAYS})")
    parser.add_argument("--rr", type=float, default=2.0, metavar="RATIO",
                        help="Risk:Reward ratio for targets (default 2.0)")
    parser.add_argument("--excel", metavar="PATH",
                        help="Also write results to Excel (date-wise sheets)")
    parser.add_argument("--buy-only", action="store_true",
                        help="Print / export only actionable signals (BREAKOUT / NEAR_TOP)")
    args = parser.parse_args()

    # Build symbol list
    symbols: list[str] = [s.upper() for s in args.symbols]
    if args.file:
        try:
            extra = [ln.strip().upper() for ln in
                     Path(args.file).read_text().splitlines()
                     if ln.strip() and not ln.startswith("#")]
            symbols += extra
        except Exception as exc:
            print(f"  [WARN] Could not read {args.file}: {exc}")

    if not symbols:
        parser.print_help()
        print("\n  Example:  python3 darvas_interpreter.py RELIANCE TCS INFY")
        sys.exit(0)

    # Deduplicate
    seen: set[str] = set()
    symbols = [s for s in symbols if not (s in seen or seen.add(s))]  # type: ignore

    print(f"\n  Darvas Interpreter  —  {len(symbols)} symbol(s)  |  "
          f"Confirm: {args.confirm} days  |  R/R: {args.rr}")
    print()

    signals: list[DarvasSignal] = []
    for i, sym in enumerate(symbols, 1):
        print(f"  [{i}/{len(symbols)}] {sym} …", end=" ", flush=True)
        sig = interpret_symbol(sym, confirm_days=args.confirm, rr_ratio=args.rr)
        signals.append(sig)
        emoji = STATE_EMOJI.get(sig.state, "")
        print(f"{emoji} {sig.state}")

    # Filter if --buy-only
    to_print = [s for s in signals
                if s.state in ("BREAKOUT_FRESH","BREAKOUT_OLD","NEAR_TOP")] \
               if args.buy_only else signals

    for sig in to_print:
        print_signal(sig)

    if args.buy_only and not to_print:
        print("\n  No actionable signals (BREAKOUT / NEAR_TOP) found today.\n")

    if args.excel:
        xl_path = Path(args.excel)
        print(f"\n  Writing Excel → {xl_path} …")
        write_signals_excel(
            [s for s in signals if not args.buy_only or
             s.state in ("BREAKOUT_FRESH","BREAKOUT_OLD","NEAR_TOP")],
            xl_path,
        )
    elif len(symbols) > 3:
        # Auto-save when scanning many symbols
        default_xl = Path(__file__).parent / "reports" / "darvas_signals_history.xlsx"
        print(f"\n  Auto-saving Excel → {default_xl} …")
        write_signals_excel(signals, default_xl)


if __name__ == "__main__":
    main()
