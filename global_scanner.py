#!/usr/bin/env python3
"""
global_scanner.py
=================
Batch Pegu + Sarvas scanner for 38,000+ global tickers.

Processes tickers in parallel chunks across 20 markets, writing results
incrementally to CSV + a final consolidated Excel workbook.

Usage:
    python global_scanner.py                        # all markets
    python global_scanner.py --markets US IN JP     # specific markets
    python global_scanner.py --batch-size 50        # tune API concurrency
    python global_scanner.py --resume               # resume from checkpoint
    python global_scanner.py --mock                 # use cached/sample data only

Output:
    reports/global_scan_YYYYMMDD/
        <MARKET>_scan.csv        per-market incremental results
        checkpoint.json          progress tracking for resume
    reports/Global_Research_YYYYMMDD.xlsx   final consolidated workbook
"""

from __future__ import annotations

import argparse
import json
import math
import os
import sys
import time
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

TODAY      = datetime.now().strftime("%Y-%m-%d")
TIMESTAMP  = datetime.now().strftime("%Y%m%d_%H%M%S")
DATE_LABEL = datetime.now().strftime("%Y%m%d")
REPO_ROOT  = Path(__file__).parent
REPORTS    = REPO_ROOT / "reports"
DATA_DIR   = REPO_ROOT / "data"

# ── Optional imports ──────────────────────────────────────────────────────────

try:
    import yfinance as yf
    _YF = True
except ImportError:
    _YF = False

try:
    from market_data_cache import MarketCache
    _CACHE = MarketCache(verbose=False)
    _CACHE_SYMS = set(_CACHE.available_symbols())
except Exception:
    _CACHE = None
    _CACHE_SYMS = set()

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _XL = True
except ImportError:
    _XL = False


# ── Ticker universe ───────────────────────────────────────────────────────────

def load_universe() -> pd.DataFrame:
    """
    Assemble the full global ticker universe.

    Sources (in priority order):
      1. data/india_tickers_full.csv  — 4,805 NSE+BSE stocks with Darvas data
      2. global_tickers.py             — 550 curated international tickers
      3. data/global_universe.json     — extended lists (written by build_universe.py)
    """
    frames: List[pd.DataFrame] = []

    # ── India: 4,805 tickers with LTP + Darvas signal ───────────────────────
    india_csv = DATA_DIR / "india_tickers_full.csv"
    if india_csv.exists():
        df = pd.read_csv(india_csv)
        df["market"] = "IN"
        df["yf_symbol"] = df["Symbol"] + df["Suffix"].fillna(".NS")
        df.rename(columns={
            "Symbol": "symbol", "LTP": "last_price", "Suffix": "suffix",
            "Change%": "change_pct", "Darvas_Signal": "darvas_signal",
        }, inplace=True)
        frames.append(df)
        print(f"[UNIVERSE] India: {len(df):,} tickers from data/india_tickers_full.csv")

    # ── International: global_tickers.py ────────────────────────────────────
    try:
        sys.path.insert(0, str(REPO_ROOT))
        from global_tickers import MARKETS, tickers_for
        intl_rows = []
        for code, mkt in MARKETS.items():
            if code == "IN":
                continue
            sfx = mkt["suffix"]
            for sym in mkt["tickers"]:
                intl_rows.append({
                    "symbol": sym, "suffix": sfx, "market": code,
                    "yf_symbol": f"{sym}{sfx}",
                    "exchange": mkt["exchange"], "currency": mkt["currency"],
                })
            for sym in mkt.get("extra", {}).get("tickers", []):
                xsfx = mkt["extra"]["suffix"]
                intl_rows.append({
                    "symbol": sym, "suffix": xsfx, "market": code,
                    "yf_symbol": f"{sym}{xsfx}",
                    "exchange": mkt["exchange"], "currency": mkt["currency"],
                })
        intl_df = pd.DataFrame(intl_rows)
        frames.append(intl_df)
        print(f"[UNIVERSE] International: {len(intl_df):,} tickers from global_tickers.py")
    except ImportError:
        pass

    # ── Extended universe: data/global_universe.json ─────────────────────────
    ext_path = DATA_DIR / "global_universe.json"
    if ext_path.exists():
        with open(ext_path) as f:
            ext = json.load(f)
        ext_rows = []
        for market_key, syms in ext.items():
            parts = market_key.split("_", 1)
            mkt_code = parts[0]
            suffix_map = {
                "US":"","JP":".T","KR":".KS","CN_SSE":".SS","CN_SZSE":".SZ",
                "HK":".HK","UK":".L","DE":".DE","FR":".PA","AU":".AX",
                "CA":".TO","BR":".SA","CN":".SS",
            }
            sfx = suffix_map.get(mkt_code, "")
            for sym in syms:
                ext_rows.append({
                    "symbol": sym, "suffix": sfx, "market": mkt_code,
                    "yf_symbol": f"{sym}{sfx}",
                })
        ext_df = pd.DataFrame(ext_rows)
        frames.append(ext_df)
        print(f"[UNIVERSE] Extended: {len(ext_df):,} tickers from global_universe.json")

    if not frames:
        raise RuntimeError("No ticker source found. Run build_universe.py first.")

    universe = pd.concat(frames, ignore_index=True)
    universe.drop_duplicates(subset=["yf_symbol"], inplace=True)
    universe.reset_index(drop=True, inplace=True)
    print(f"[UNIVERSE] Total: {len(universe):,} unique tickers across {universe['market'].nunique()} markets")
    return universe


# ── Data fetching ─────────────────────────────────────────────────────────────

def fetch_batch(symbols: List[str], period: str = "1y") -> Dict[str, dict]:
    """
    Fetch fundamental + technical snapshot for a batch of yfinance symbols.
    Returns dict: yf_symbol → data_dict.
    """
    result: Dict[str, dict] = {}

    if not _YF:
        return result

    try:
        raw = yf.download(
            symbols, period=period, auto_adjust=True,
            threads=True, progress=False, timeout=20,
        )
        if raw.empty:
            return result

        closes = raw["Close"] if isinstance(raw.columns, pd.MultiIndex) else raw[["Close"]]

        for sym in symbols:
            try:
                # OHLC for technicals
                if isinstance(raw.columns, pd.MultiIndex):
                    df = raw.xs(sym, axis=1, level=1).dropna(how="all")
                else:
                    df = raw.copy()

                if df.empty or len(df) < 20:
                    continue

                close = df["Close"]
                last  = float(close.iloc[-1])
                dma50  = float(close.rolling(50).mean().iloc[-1])
                dma200 = float(close.rolling(200).mean().iloc[-1])
                high52 = float(close.tail(252).max())
                low52  = float(close.tail(252).min())

                delta = close.diff()
                gain  = delta.clip(lower=0).rolling(14).mean()
                loss  = (-delta.clip(upper=0)).rolling(14).mean()
                rsi   = float((100 - 100 / (1 + gain / loss.replace(0, np.nan))).iloc[-1])

                ema12 = close.ewm(span=12, adjust=False).mean()
                ema26 = close.ewm(span=26, adjust=False).mean()
                macd  = float((ema12 - ema26).iloc[-1])

                vol_ratio = None
                if "Volume" in df.columns:
                    avg = df["Volume"].tail(20).mean()
                    if avg > 0:
                        vol_ratio = float(df["Volume"].iloc[-1] / avg)

                result[sym] = {
                    "last_price": round(last, 4),
                    "ma50": round(dma50, 4),
                    "ma200": round(dma200, 4),
                    "week52_high": round(high52, 4),
                    "week52_low": round(low52, 4),
                    "rsi_14": round(rsi, 2),
                    "macd": round(macd, 6),
                    "volume_ratio": round(vol_ratio, 3) if vol_ratio else None,
                    "pct_above_50dma":  round((last/dma50-1)*100, 2),
                    "pct_above_200dma": round((last/dma200-1)*100, 2),
                    "pct_from_52w_high":round((last/high52-1)*100, 2),
                }

            except Exception:
                continue

    except Exception:
        pass

    # Fetch fundamentals per ticker (slow but thorough)
    for sym in symbols:
        if sym not in result:
            continue
        try:
            t = yf.Ticker(sym)
            info = t.info or {}
            result[sym].update({
                "pe_ratio":         info.get("trailingPE"),
                "pb_ratio":         info.get("priceToBook"),
                "peg_ratio":        info.get("pegRatio"),
                "roe":              _pct(info.get("returnOnEquity")),
                "operating_margins":_pct(info.get("operatingMargins")),
                "debt_to_equity":   info.get("debtToEquity"),
                "current_ratio":    info.get("currentRatio"),
                "eps_trailing":     info.get("trailingEps"),
                "eps_forward":      info.get("forwardEps"),
                "revenue_growth":   _pct(info.get("revenueGrowth")),
                "earnings_growth":  _pct(info.get("earningsGrowth")),
                "dividend_yield":   _pct(info.get("dividendYield")),
                "beta":             info.get("beta"),
                "market_cap":       info.get("marketCap"),
                "target_price":     info.get("targetMeanPrice"),
                "sector":           info.get("sector"),
                "industry":         info.get("industry"),
                "name":             info.get("longName") or info.get("shortName"),
            })
        except Exception:
            pass

    return result


def _pct(v):
    if v is None:
        return None
    return round(v * 100, 2) if abs(v) < 2 else round(v, 2)


# ── Pegu scoring ──────────────────────────────────────────────────────────────

def _score(val, breakpoints):
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return breakpoints[-1][1]
    for threshold, score in breakpoints:
        if val < threshold:
            return score
    return breakpoints[-1][1]


def pegu_score(row: dict) -> Tuple[float, str]:
    """Compute Pegu score (0-100) and grade for a stock data dict."""

    # Valuation (30 pts)
    pe  = _score(row.get("pe_ratio"),  [(0,0),(10,4),(15,8),(20,12),(25,10),(35,6),(50,2),(1e9,0)])
    peg = _score(row.get("peg_ratio"), [(0,0),(0.5,6),(1.0,10),(1.5,8),(2.0,4),(3.0,2),(1e9,0)])
    pb  = _score(row.get("pb_ratio"),  [(0,0),(1,6),(2,10),(3,8),(5,4),(10,2),(1e9,0)])
    val = min(pe + peg + pb, 30)

    # Quality (30 pts)
    roe = _score(row.get("roe"),              [(0,0),(5,3),(10,7),(15,12),(20,18),(25,22),(30,25),(1e9,30)])
    opm = _score(row.get("operating_margins"),[(0,0),(5,2),(10,5),(15,8),(20,10),(25,12),(30,14),(1e9,16)])
    de  = row.get("debt_to_equity")
    de_s= 14 if de is None else (14 if de < 0 else _score(de, [(0,14),(20,12),(40,10),(80,7),(120,4),(200,2),(1e9,0)]))
    qual = min(roe + opm + de_s, 30)

    # Growth (25 pts)
    epsg = _score(row.get("earnings_growth"), [(-1e9,0),(0,1),(5,4),(10,8),(15,12),(20,16),(25,20),(30,24),(1e9,25)])
    revg = _score(row.get("revenue_growth"),  [(-1e9,0),(0,1),(5,2),(10,5),(15,7),(20,9),(25,11),(1e9,13)])
    fpe  = row.get("pe_ratio"); efwd = row.get("eps_forward"); etr = row.get("eps_trailing")
    fwd_pe = (row.get("last_price") / efwd) if efwd and efwd > 0 and row.get("last_price") else None
    fpe_s  = _score(fwd_pe, [(0,0),(10,4),(15,6),(20,5),(25,4),(35,2),(1e9,0)])
    grw = min(epsg + revg + fpe_s, 25)

    # Safety (15 pts)
    cr  = _score(row.get("current_ratio"),  [(0,0),(0.5,0),(1.0,4),(1.5,7),(2.0,10),(3.0,12),(1e9,15)])
    div = _score(row.get("dividend_yield"), [(0,0),(1,2),(2,4),(3,5),(4,5),(6,4),(1e9,3)])
    beta= row.get("beta")
    b_s = 5 if beta is None else _score(abs(beta), [(0,0),(0.5,5),(0.8,4),(1.0,3),(1.2,2),(1.5,1),(1e9,0)])
    saf = min(cr + div + b_s, 15)

    total = val + qual + grw + saf
    if   total >= 80: grade = "A+"
    elif total >= 70: grade = "A"
    elif total >= 60: grade = "B+"
    elif total >= 50: grade = "B"
    elif total >= 40: grade = "C"
    elif total >= 25: grade = "D"
    else:             grade = "F"

    return round(total, 1), grade


# ── Sarvas scan ───────────────────────────────────────────────────────────────

def sarvas_signal(row: dict, pegu: float) -> Tuple[int, str]:
    """Return (score, signal) for Sarvas scan."""
    pts = 0

    # Technical (45 pts)
    ltp    = row.get("last_price", 0) or 0
    ma50   = row.get("ma50") or 0
    ma200  = row.get("ma200") or 0
    rsi    = row.get("rsi_14")
    macd   = row.get("macd")
    vr     = row.get("volume_ratio")
    h52    = row.get("week52_high") or 0

    if ma50  and ltp > ma50:             pts += 8
    if ma200 and ltp > ma200:            pts += 10
    if ma50  and ma200 and ma50 > ma200: pts += 10  # golden cross
    if rsi   and 35 <= rsi <= 68:        pts += 7
    if macd  and macd > 0:               pts += 5
    if vr    and vr > 1.5:              pts += 5
    if h52   and ltp >= 0.9 * h52:      pts += 7

    # Fundamental (55 pts)
    eps = row.get("eps_trailing")
    eg  = row.get("earnings_growth")
    rg  = row.get("revenue_growth")
    pe  = row.get("pe_ratio")
    de  = row.get("debt_to_equity")
    roe = row.get("roe")
    tp  = row.get("target_price")

    if eps and eps > 0:                  pts += 10
    if eg  and eg > 10:                  pts += 8
    if rg  and rg > 5:                   pts += 8
    if pe  and 5 < pe < 30:             pts += 7
    if de  is not None and de < 50:      pts += 7
    if roe and roe > 15:                 pts += 7
    if tp  and ltp and tp > ltp * 1.1:  pts += 8
    if pegu >= 60:                       pts += 5  # Pegu bonus

    if   pts >= 80: sig = "STRONG BUY"
    elif pts >= 65: sig = "BUY"
    elif pts >= 50: sig = "ACCUMULATE"
    elif pts >= 35: sig = "HOLD"
    elif pts >= 20: sig = "REDUCE"
    else:           sig = "SELL"

    return pts, sig


# ── Main scan loop ────────────────────────────────────────────────────────────

def scan_universe(
    universe: pd.DataFrame,
    markets: Optional[List[str]] = None,
    batch_size: int = 50,
    workers: int = 4,
    mock: bool = False,
    resume: bool = False,
) -> pd.DataFrame:
    """
    Run Pegu + Sarvas on the full universe, returning a results DataFrame.
    Writes per-market CSVs incrementally for fault tolerance.
    """
    out_dir = REPORTS / f"global_scan_{DATE_LABEL}"
    out_dir.mkdir(parents=True, exist_ok=True)

    checkpoint_file = out_dir / "checkpoint.json"
    checkpoint: dict = {}
    if resume and checkpoint_file.exists():
        with open(checkpoint_file) as f:
            checkpoint = json.load(f)
        print(f"[SCAN] Resuming — {len(checkpoint)} markets already done")

    if markets:
        universe = universe[universe["market"].isin(markets)].copy()

    all_results: List[pd.DataFrame] = []

    for mkt, grp in universe.groupby("market"):
        if mkt in checkpoint:
            csv_path = out_dir / f"{mkt}_scan.csv"
            if csv_path.exists():
                all_results.append(pd.read_csv(csv_path))
                print(f"[SCAN] {mkt}: loaded from checkpoint ({len(grp):,} tickers)")
                continue

        print(f"\n[SCAN] {mkt}: {len(grp):,} tickers", flush=True)
        yf_syms = grp["yf_symbol"].dropna().tolist()

        rows = []

        if mock or not _YF:
            # Use any pre-existing data from the row, score with NaN-safe defaults
            for _, r in grp.iterrows():
                d = r.to_dict()
                pg, grade = pegu_score(d)
                ss, sig   = sarvas_signal(d, pg)
                d.update({
                    "pegu_score": pg, "pegu_grade": grade,
                    "sarvas_score": ss, "sarvas_signal": sig,
                    "data_source": "local",
                })
                rows.append(d)
        else:
            # Batch fetch via yfinance
            batches = [yf_syms[i:i+batch_size] for i in range(0, len(yf_syms), batch_size)]
            total_b = len(batches)

            def process_batch(batch_idx_syms):
                idx, batch = batch_idx_syms
                data = fetch_batch(batch)
                return idx, data

            with ThreadPoolExecutor(max_workers=workers) as ex:
                futures = {ex.submit(process_batch, (i, b)): i for i, b in enumerate(batches)}
                fetched: Dict[str, dict] = {}
                for fut in as_completed(futures):
                    idx, data = fut.result()
                    fetched.update(data)
                    done = len(fetched)
                    print(f"  {mkt}: {done}/{len(yf_syms)} fetched", end="\r", flush=True)
                    time.sleep(0.1)

            print()

            for _, r in grp.iterrows():
                sym = r.get("yf_symbol", "")
                d   = r.to_dict()
                d.update(fetched.get(sym, {}))
                pg, grade = pegu_score(d)
                ss, sig   = sarvas_signal(d, pg)
                d.update({
                    "pegu_score": pg, "pegu_grade": grade,
                    "sarvas_score": ss, "sarvas_signal": sig,
                    "data_source": "live" if sym in fetched else "no_data",
                })
                rows.append(d)

        mkt_df = pd.DataFrame(rows)
        mkt_df["scan_date"] = TODAY
        csv_path = out_dir / f"{mkt}_scan.csv"
        mkt_df.to_csv(csv_path, index=False)
        all_results.append(mkt_df)
        checkpoint[mkt] = {"tickers": len(rows), "ts": TODAY}
        with open(checkpoint_file, "w") as f:
            json.dump(checkpoint, f)
        print(f"  {mkt}: {len(rows):,} stocks scored → {csv_path.name}")

    if not all_results:
        return pd.DataFrame()

    combined = pd.concat(all_results, ignore_index=True)
    combined.sort_values("pegu_score", ascending=False, inplace=True)
    combined.reset_index(drop=True, inplace=True)
    return combined


# ── Excel export ──────────────────────────────────────────────────────────────

_CLR = {
    "header_bg":  "1F4E79", "header_fg":  "FFFFFF",
    "A+": "00B050", "A": "70AD47", "B+": "C6EFCE", "B": "FFEB9C",
    "C":  "FFC7CE", "D": "FF0000", "F":  "990000",
    "BUY_BG": "C6EFCE", "SELL_BG": "FFC7CE", "HOLD_BG": "FFEB9C",
    "alt_row": "F2F7FC",
}

_GRADE_COLOR = {"A+":"00B050","A":"70AD47","B+":"92D050","B":"FFEB9C","C":"FFC7CE","D":"FF0000","F":"990000"}
_SIG_COLOR   = {"STRONG BUY":"00B050","BUY":"70AD47","ACCUMULATE":"C6EFCE","HOLD":"FFEB9C","REDUCE":"FFC7CE","SELL":"FF0000"}


def export_excel(df: pd.DataFrame, out_dir: Path) -> Path:
    if not _XL or df.empty:
        return None

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    xlsx_path = REPORTS / f"Global_Research_{DATE_LABEL}.xlsx"
    wb = Workbook()

    def hfill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def header_row(ws, cols):
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font      = Font(bold=True, color="FFFFFF", size=10)
            c.fill      = hfill("1F4E79")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

    def autofit(ws, df_cols):
        for i, col in enumerate(df_cols, 1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max() if col in df.columns else 10)
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 25)

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    summary_data = {
        "Total Tickers Scanned": len(df),
        "Markets Covered":       df["market"].nunique() if "market" in df.columns else "—",
        "Data With Live Prices": (df["data_source"] == "live").sum() if "data_source" in df.columns else "—",
        "Avg Pegu Score":        round(df["pegu_score"].mean(), 1),
        "Grade A+ (≥80)":        (df["pegu_grade"] == "A+").sum(),
        "Grade A  (≥70)":        (df["pegu_grade"] == "A").sum(),
        "Grade B+ (≥60)":        (df["pegu_grade"] == "B+").sum(),
        "Grade B  (≥50)":        (df["pegu_grade"] == "B").sum(),
        "Grade C  (≥40)":        (df["pegu_grade"] == "C").sum(),
        "Grade D  (≥25)":        (df["pegu_grade"] == "D").sum(),
        "Grade F  (<25)":        (df["pegu_grade"] == "F").sum(),
        "STRONG BUY Signals":    (df["sarvas_signal"] == "STRONG BUY").sum(),
        "BUY Signals":           (df["sarvas_signal"] == "BUY").sum(),
        "ACCUMULATE Signals":    (df["sarvas_signal"] == "ACCUMULATE").sum(),
        "HOLD Signals":          (df["sarvas_signal"] == "HOLD").sum(),
        "REDUCE Signals":        (df["sarvas_signal"] == "REDUCE").sum(),
        "SELL Signals":          (df["sarvas_signal"] == "SELL").sum(),
        "Scan Date":             TODAY,
    }
    ws.cell(1, 1, "Global Research Summary").font = Font(bold=True, size=14, color="1F4E79")
    for ri, (k, v) in enumerate(summary_data.items(), 3):
        ws.cell(ri, 1, k).font   = Font(bold=True)
        ws.cell(ri, 2, v)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    # ── Sheet 2: Top 500 by Pegu ─────────────────────────────────────────────
    top500 = df.head(500)
    cols = ["yf_symbol","market","name","sector","last_price","pegu_score","pegu_grade",
            "sarvas_score","sarvas_signal","pe_ratio","pb_ratio","roe","operating_margins",
            "earnings_growth","revenue_growth","rsi_14","pct_above_200dma","market_cap"]
    cols = [c for c in cols if c in top500.columns]
    ws2 = wb.create_sheet("Top 500 (Pegu)")
    header_row(ws2, cols)
    for ri, (_, row) in enumerate(top500[cols].iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            ws2.cell(ri, ci, row[col] if pd.notna(row[col]) else None)
        # Color grade
        if "pegu_grade" in cols:
            gi = cols.index("pegu_grade") + 1
            grade = row.get("pegu_grade", "")
            if grade in _GRADE_COLOR:
                ws2.cell(ri, gi).fill = hfill(_GRADE_COLOR[grade])
    autofit(ws2, cols)

    # ── Sheet 3: Strong Buy signals ──────────────────────────────────────────
    buys = df[df["sarvas_signal"].isin(["STRONG BUY","BUY"])].head(500)
    ws3  = wb.create_sheet("Strong Buy & Buy")
    header_row(ws3, cols)
    for ri, (_, row) in enumerate(buys[cols].iterrows(), 2):
        for ci, col in enumerate(cols, 1):
            ws3.cell(ri, ci, row[col] if pd.notna(row[col]) else None)
        if "sarvas_signal" in cols:
            si = cols.index("sarvas_signal") + 1
            sig = row.get("sarvas_signal","")
            ws3.cell(ri, si).fill = hfill(_SIG_COLOR.get(sig, "FFFFFF"))
    autofit(ws3, cols)

    # ── Sheet 4: Per-market breakdown ────────────────────────────────────────
    ws4 = wb.create_sheet("By Market")
    mkt_cols = ["Market","Tickers","Avg Pegu","A+","A","B+","Buys","Holds","Sells"]
    header_row(ws4, mkt_cols)
    for ri, (mkt, mg) in enumerate(df.groupby("market"), 2):
        ws4.cell(ri, 1, mkt)
        ws4.cell(ri, 2, len(mg))
        ws4.cell(ri, 3, round(mg["pegu_score"].mean(), 1))
        ws4.cell(ri, 4, (mg["pegu_grade"]=="A+").sum())
        ws4.cell(ri, 5, (mg["pegu_grade"]=="A").sum())
        ws4.cell(ri, 6, (mg["pegu_grade"]=="B+").sum())
        ws4.cell(ri, 7, mg["sarvas_signal"].isin(["STRONG BUY","BUY"]).sum())
        ws4.cell(ri, 8, (mg["sarvas_signal"]=="HOLD").sum())
        ws4.cell(ri, 9, mg["sarvas_signal"].isin(["REDUCE","SELL"]).sum())
    for col in ["A","B","C","D","E","F","G","H","I"]:
        ws4.column_dimensions[col].width = 12

    # ── Sheet 5: All tickers (chunked for Excel row limit) ──────────────────
    MAX_ROWS = 1_048_000
    all_cols = list(df.columns)
    chunk = df.head(MAX_ROWS)
    ws5 = wb.create_sheet("All Tickers")
    header_row(ws5, all_cols)
    for ri, (_, row) in enumerate(chunk.iterrows(), 2):
        for ci, col in enumerate(all_cols, 1):
            v = row[col]
            ws5.cell(ri, ci, v if pd.notna(v) else None)
    print(f"  [XL] All Tickers sheet: {len(chunk):,} rows")

    wb.save(xlsx_path)
    print(f"\n[EXCEL] Saved → {xlsx_path}")
    return xlsx_path


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Global 38K-ticker Pegu+Sarvas scanner")
    ap.add_argument("--markets",    nargs="+", help="Market codes to scan (e.g. IN US JP)")
    ap.add_argument("--batch-size", type=int, default=50,  help="yfinance batch size")
    ap.add_argument("--workers",    type=int, default=4,   help="Parallel fetch threads")
    ap.add_argument("--mock",       action="store_true",   help="Skip live fetch, use cached data only")
    ap.add_argument("--resume",     action="store_true",   help="Resume from last checkpoint")
    ap.add_argument("--excel-only", action="store_true",   help="Skip scan, build Excel from existing CSVs")
    args = ap.parse_args()

    print(f"{'='*60}")
    print(f"  Global Stock Research — {TODAY}")
    print(f"{'='*60}\n")

    universe = load_universe()

    if args.excel_only:
        out_dir = REPORTS / f"global_scan_{DATE_LABEL}"
        csvs = list(out_dir.glob("*_scan.csv"))
        frames = [pd.read_csv(c) for c in csvs]
        results = pd.concat(frames, ignore_index=True).sort_values("pegu_score", ascending=False)
    else:
        results = scan_universe(
            universe,
            markets    = args.markets,
            batch_size = args.batch_size,
            workers    = args.workers,
            mock       = args.mock,
            resume     = args.resume,
        )

    if results.empty:
        print("[WARN] No results produced.")
        return

    print(f"\n[RESULTS] {len(results):,} stocks scored")
    top = results[results["pegu_grade"].isin(["A+","A"])].head(20)
    if not top.empty:
        print("\nTop 20 by Pegu Score:")
        disp_cols = ["yf_symbol","market","pegu_score","pegu_grade","sarvas_signal"]
        disp_cols = [c for c in disp_cols if c in top.columns]
        print(top[disp_cols].to_string(index=False))

    out_dir = REPORTS / f"global_scan_{DATE_LABEL}"
    export_excel(results, out_dir)

    # Save master CSV
    master_csv = REPORTS / f"Global_Research_{DATE_LABEL}.csv"
    results.to_csv(master_csv, index=False)
    print(f"[CSV]   Saved → {master_csv}")


if __name__ == "__main__":
    main()
