#!/usr/bin/env python3
"""
morning_briefing.py
===================
Standalone India market morning briefing — runs entirely from yfinance.
No Cassandra, no NSE API cookies, no R pipeline required.

What it does:
  1. Fetches live quotes + 1-year history for a watchlist of NSE stocks
  2. Computes RSI-14, EMA-20/50/200, MACD, Bollinger Bands, Stochastics, ATR
  3. Runs Darvas/Buffett + Piotroski scoring (reuses daily_scanner.py logic)
  4. Fetches NIFTY 50 & SENSEX index levels
  5. Prints a formatted morning briefing to stdout
  6. Optionally saves an HTML report to reports/morning_briefing.html

Usage:
    python3 morning_briefing.py
    python3 morning_briefing.py --symbols RELIANCE TCS INFY HDFCBANK
    python3 morning_briefing.py --nifty50       # scan full NIFTY 50
    python3 morning_briefing.py --no-excel      # terminal only
    python3 morning_briefing.py --no-fundamentals  # faster, no PE/ROE

Dependencies (standard PyPI — no Claude, no Anthropic SDK required):
    pip3 install yfinance pandas numpy openpyxl requests

Standalone execution (no Claude Code, no app server needed):
    python3 morning_briefing.py
"""
from __future__ import annotations

import argparse
import math
import sys
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

warnings.filterwarnings("ignore")

# ── dependency check ──────────────────────────────────────────────────────────
try:
    import numpy as np
    import pandas as pd
except ImportError:
    sys.exit("Install dependencies:  pip install yfinance pandas numpy requests")

try:
    import yfinance as yf
except ImportError:
    sys.exit("Install yfinance:  pip install yfinance")

# ── constants ─────────────────────────────────────────────────────────────────

NIFTY50_SYMBOLS = [
    "ADANIENT", "ADANIPORTS", "APOLLOHOSP", "ASIANPAINT", "AXISBANK",
    "BAJAJ-AUTO", "BAJFINANCE", "BAJAJFINSV", "BPCL", "BHARTIARTL",
    "BRITANNIA", "CIPLA", "COALINDIA", "DIVISLAB", "DRREDDY",
    "EICHERMOT", "GRASIM", "HCLTECH", "HDFCBANK", "HDFCLIFE",
    "HEROMOTOCO", "HINDALCO", "HINDUNILVR", "ICICIBANK", "ITC",
    "INDUSINDBK", "INFY", "JSWSTEEL", "KOTAKBANK", "LT",
    "M&M", "MARUTI", "NTPC", "NESTLEIND", "ONGC",
    "POWERGRID", "RELIANCE", "SBILIFE", "SHRIRAMFIN", "SBIN",
    "SUNPHARMA", "TCS", "TATACONSUM", "TATAMOTORS", "TATASTEEL",
    "TECHM", "TITAN", "ULTRACEMCO", "WIPRO", "ZOMATO",
]

DEFAULT_WATCHLIST = [
    "RELIANCE", "TCS", "INFY", "HDFCBANK", "ICICIBANK",
    "AXISBANK", "KOTAKBANK", "SBIN", "BHARTIARTL", "ITC",
    "HINDUNILVR", "MARUTI", "BAJFINANCE", "LT", "TITAN",
    "SUNPHARMA", "WIPRO", "HCLTECH", "TATAMOTORS", "ADANIPORTS",
]

INDICES = {
    "NIFTY 50":  "^NSEI",
    "SENSEX":    "^BSESN",
    "NIFTY BANK":"^NSEBANK",
    "NIFTY IT":  "^CNXIT",
    "NIFTY MID": "NIFTY_MID_SELECT.NS",
}


# ── technical indicator functions ────────────────────────────────────────────

def _rsi(closes: pd.Series, period: int = 14) -> float | None:
    if len(closes) < period + 1:
        return None
    delta = closes.diff().dropna()
    gain = delta.clip(lower=0).ewm(com=period - 1, min_periods=period).mean()
    loss = (-delta).clip(lower=0).ewm(com=period - 1, min_periods=period).mean()
    rs = gain / loss.replace(0, float("nan"))
    val = 100 - (100 / (1 + rs))
    return round(float(val.iloc[-1]), 2)


def _ema(closes: pd.Series, period: int) -> float | None:
    if len(closes) < period:
        return None
    return round(float(closes.ewm(span=period, adjust=False).mean().iloc[-1]), 2)


def _macd(closes: pd.Series) -> tuple[float | None, float | None]:
    if len(closes) < 35:
        return None, None
    ema12 = closes.ewm(span=12, adjust=False).mean()
    ema26 = closes.ewm(span=26, adjust=False).mean()
    macd_line = ema12 - ema26
    signal = macd_line.ewm(span=9, adjust=False).mean()
    return round(float(macd_line.iloc[-1]), 4), round(float(signal.iloc[-1]), 4)


def _bollinger(closes: pd.Series, period: int = 20) -> tuple[float | None, float | None, float | None]:
    if len(closes) < period:
        return None, None, None
    sma = closes.rolling(period).mean()
    std = closes.rolling(period).std()
    upper = sma + 2 * std
    lower = sma - 2 * std
    u, l, c = float(upper.iloc[-1]), float(lower.iloc[-1]), float(closes.iloc[-1])
    pct_b = (c - l) / (u - l) if (u - l) != 0 else None
    return round(u, 2), round(l, 2), round(pct_b, 4) if pct_b is not None else None


def _stochastic(hist: pd.DataFrame, period: int = 14, smooth: int = 3) -> tuple[float | None, float | None]:
    if len(hist) < period + smooth:
        return None, None
    low_min  = hist["Low"].rolling(period).min()
    high_max = hist["High"].rolling(period).max()
    denom = (high_max - low_min).replace(0, float("nan"))
    k = 100 * (hist["Close"] - low_min) / denom
    k_smooth = k.rolling(smooth).mean()
    d_smooth = k_smooth.rolling(smooth).mean()
    return round(float(k_smooth.iloc[-1]), 2), round(float(d_smooth.iloc[-1]), 2)


def _atr(hist: pd.DataFrame, period: int = 14) -> float | None:
    if len(hist) < period + 1:
        return None
    high, low, prev_close = hist["High"], hist["Low"], hist["Close"].shift(1)
    tr = pd.concat([
        high - low,
        (high - prev_close).abs(),
        (low  - prev_close).abs()
    ], axis=1).max(axis=1)
    return round(float(tr.ewm(com=period - 1, min_periods=period).mean().iloc[-1]), 4)


def _volume_ratio(hist: pd.DataFrame, period: int = 20) -> float | None:
    if len(hist) < period + 1:
        return None
    avg_vol = hist["Volume"].rolling(period).mean().iloc[-1]
    if avg_vol == 0:
        return None
    return round(float(hist["Volume"].iloc[-1] / avg_vol), 2)


def _return_pct(closes: pd.Series, n_days: int) -> float | None:
    if len(closes) < n_days + 1:
        return None
    start = float(closes.iloc[-(n_days + 1)])
    end   = float(closes.iloc[-1])
    if start == 0:
        return None
    return round((end - start) / start * 100, 2)


# ── data fetching ─────────────────────────────────────────────────────────────

def fetch_stock_data(symbol: str, period: str = "1y") -> dict | None:
    """
    Download 1-year daily history for an NSE symbol and compute all indicators.
    Returns a flat dict ready for the scanner, or None on failure.
    """
    ticker_str = f"{symbol}.NS"
    try:
        ticker = yf.Ticker(ticker_str)
        hist = ticker.history(period=period, auto_adjust=True)
        if hist.empty or len(hist) < 30:
            return None

        closes = hist["Close"].dropna()
        info   = {}
        try:
            info = ticker.fast_info  # lightweight; avoids heavy .info call
        except Exception:
            pass

        # Try full info for fundamentals (may be slow but only for watchlist)
        full_info = {}
        try:
            full_info = ticker.info
        except Exception:
            pass

        cmp      = round(float(closes.iloc[-1]), 2)
        high_52w = round(float(hist["High"].tail(252).max()), 2)
        low_52w  = round(float(hist["Low"].tail(252).min()), 2)

        macd_val, macd_sig = _macd(closes)
        bb_u, bb_l, bb_pct = _bollinger(closes)
        stk, std_k          = _stochastic(hist)
        atr                 = _atr(hist)
        vr                  = _volume_ratio(hist)

        return {
            "ticker":         symbol,
            "name":           full_info.get("longName") or full_info.get("shortName") or symbol,
            "sector":         full_info.get("sector", ""),
            "industry":       full_info.get("industry", ""),
            "cmp":            cmp,
            "high_52w":       high_52w,
            "low_52w":        low_52w,
            "ema_20":         _ema(closes, 20),
            "ema_50":         _ema(closes, 50),
            "ema_200":        _ema(closes, 200),
            "rsi":            _rsi(closes),
            "rsi_signal":     _rsi_signal(_rsi(closes)),
            "macd":           macd_val,
            "macd_signal":    macd_sig,
            "bb_upper":       bb_u,
            "bb_lower":       bb_l,
            "bb_pct":         bb_pct,
            "stoch_k":        stk,
            "stoch_d":        std_k,
            "atr_14":         atr,
            "volume_ratio":   vr,
            "volume":         int(hist["Volume"].iloc[-1]),
            # fundamentals
            "pe":             full_info.get("trailingPE"),
            "pb":             full_info.get("priceToBook"),
            "roe":            _pct(full_info.get("returnOnEquity")),
            "opm":            _pct(full_info.get("operatingMargins")),
            "debt_to_equity": full_info.get("debtToEquity"),
            "market_cap":     full_info.get("marketCap"),
            "beta":           full_info.get("beta"),
            "current_ratio":  full_info.get("currentRatio"),
            "revenue_growth": _pct(full_info.get("revenueGrowth")),
            "eps":            full_info.get("trailingEps"),
            "dividend_yield": full_info.get("dividendYield"),
            # returns
            "ret_1d":  _return_pct(closes, 1),
            "ret_1w":  _return_pct(closes, 5),
            "ret_1m":  _return_pct(closes, 21),
            "ret_3m":  _return_pct(closes, 63),
            "ret_6m":  _return_pct(closes, 126),
            "ret_1y":  _return_pct(closes, 252),
            "_exchange": "NSE",
        }
    except Exception as exc:
        print(f"  [WARN] {symbol}: {exc}")
        return None


def _pct(v) -> float | None:
    """Convert yfinance decimal ratio (e.g. 0.15) → percentage (15.0)."""
    try:
        return round(float(v) * 100, 2) if v is not None else None
    except Exception:
        return None


def _rsi_signal(rsi: float | None) -> str:
    if rsi is None:
        return "HOLD"
    if rsi >= 70:
        return "OVERBOUGHT"
    if rsi <= 30:
        return "OVERSOLD"
    if rsi >= 60:
        return "STRONG"
    if rsi <= 40:
        return "WEAK"
    return "HOLD"


def fetch_index_data() -> dict:
    """Fetch current levels + 1-day change for major India indices."""
    results = {}
    tickers_str = " ".join(INDICES.values())
    try:
        raw = yf.download(tickers_str, period="5d", auto_adjust=True, progress=False)
        closes = raw["Close"] if "Close" in raw.columns else raw.xs("Close", axis=1, level=0)

        for name, sym in INDICES.items():
            try:
                col = closes[sym].dropna()
                if len(col) >= 2:
                    level   = round(float(col.iloc[-1]), 2)
                    prev    = round(float(col.iloc[-2]), 2)
                    chg     = round(level - prev, 2)
                    chg_pct = round((chg / prev) * 100, 2) if prev else 0
                    results[name] = {
                        "level":   level,
                        "change":  chg,
                        "chg_pct": chg_pct,
                        "trend":   "▲" if chg >= 0 else "▼",
                    }
            except Exception:
                pass
    except Exception as exc:
        print(f"  [WARN] Index fetch failed: {exc}")
    return results


# ── scanner (inline — mirrors daily_scanner.py logic) ────────────────────────

def _col(df: pd.DataFrame, name: str) -> pd.Series:
    return pd.to_numeric(df[name], errors="coerce") if name in df.columns \
        else pd.Series(np.nan, index=df.index)


def _vcrit(passes: pd.Series, valid: pd.Series) -> pd.Series:
    out = pd.Series(None, index=passes.index, dtype=object)
    out[valid] = passes[valid]
    return out


def _pts(crit: pd.Series, weight: int = 1) -> pd.Series:
    return (crit == True).astype(int) * weight  # noqa: E712


def _r(v) -> float | None:
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
    except (TypeError, ValueError):
        return None


def scan_darvas(df: pd.DataFrame) -> list[dict]:
    if df.empty:
        return []
    d = df.reset_index(drop=True)

    cmp    = _col(d, "cmp");    h52  = _col(d, "high_52w"); l52  = _col(d, "low_52w")
    e20    = _col(d, "ema_20"); e50  = _col(d, "ema_50");   e200 = _col(d, "ema_200")
    rsi    = _col(d, "rsi");    roe  = _col(d, "roe");      pe   = _col(d, "pe")
    de     = _col(d, "debt_to_equity")
    macd   = _col(d, "macd");   msig = _col(d, "macd_signal")
    vr     = _col(d, "volume_ratio"); bb_pct = _col(d, "bb_pct")
    stk    = _col(d, "stoch_k"); std = _col(d, "stoch_d")

    h52s = h52.where(h52 > 0)
    v1  = cmp.notna() & h52.gt(0) & h52.notna()
    v2  = cmp.notna() & e50.notna()
    v3  = cmp.notna() & h52.notna() & l52.notna() & (h52 - l52).gt(0)
    v4  = rsi.notna()
    v5  = roe.notna(); v6 = pe.notna(); v7 = de.notna()
    v8  = cmp.notna() & e200.notna()
    v9  = macd.notna() & msig.notna()
    v10 = vr.notna()
    v11 = e20.notna() & e50.notna()
    v12 = bb_pct.notna()
    v13 = stk.notna() & std.notna()

    c1  = _vcrit((h52 - cmp) / h52s <= 0.15, v1)
    c2  = _vcrit(cmp > e50, v2)
    c3  = _vcrit((cmp - l52) / (h52 - l52) >= 0.60, v3)
    c4  = _vcrit(rsi.between(40, 68), v4)
    c5  = _vcrit(roe > 12, v5)
    c6  = _vcrit(pe.between(5, 30), v6)
    c7  = _vcrit(de < 1.0, v7)
    c8  = _vcrit(cmp > e200, v8)
    c9  = _vcrit(macd > msig, v9)
    c10 = _vcrit(vr > 1.5, v10)
    c11 = _vcrit(e20 > e50, v11)
    c12 = _vcrit(bb_pct > 0.5, v12)
    c13 = _vcrit(stk.between(20, 80) & (stk > std), v13)

    score = (_pts(c1, 2) + _pts(c2, 2) + _pts(c3) + _pts(c4) +
             _pts(c5) + _pts(c6) + _pts(c7) + _pts(c8) +
             _pts(c9) + _pts(c10) + _pts(c11) + _pts(c12) + _pts(c13))

    signal = np.where(score >= 5, "BUY", np.where(score >= 3, "WATCH", "AVOID"))

    names_col = d["name"].fillna("").astype(str).tolist() if "name" in d.columns else [""] * len(d)
    tickers   = d["ticker"].fillna("").astype(str).tolist() if "ticker" in d.columns else [""] * len(d)
    sectors   = d["sector"].fillna("").astype(str).tolist() if "sector" in d.columns else [""] * len(d)

    results = []
    for i in range(len(d)):
        results.append({
            "ticker":       tickers[i],
            "name":         names_col[i],
            "sector":       sectors[i],
            "cmp":          _r(cmp.iat[i]),
            "high_52w":     _r(h52.iat[i]),
            "rsi":          _r(rsi.iat[i]),
            "ema_50":       _r(e50.iat[i]),
            "pe":           _r(pe.iat[i]),
            "roe":          _r(roe.iat[i]),
            "ret_1d":       _r(_col(d, "ret_1d").iat[i]),
            "ret_1w":       _r(_col(d, "ret_1w").iat[i]),
            "ret_1m":       _r(_col(d, "ret_1m").iat[i]),
            "score":        int(score.iat[i]),
            "signal":       str(signal[i]),
            "criteria": {
                "near_52w_high":  bool(c1.iat[i]) if c1.iat[i] is not None else None,
                "above_ema50":    bool(c2.iat[i]) if c2.iat[i] is not None else None,
                "range_strength": bool(c3.iat[i]) if c3.iat[i] is not None else None,
                "rsi_healthy":    bool(c4.iat[i]) if c4.iat[i] is not None else None,
                "buffett_roe":    bool(c5.iat[i]) if c5.iat[i] is not None else None,
                "buffett_pe":     bool(c6.iat[i]) if c6.iat[i] is not None else None,
                "buffett_de":     bool(c7.iat[i]) if c7.iat[i] is not None else None,
                "above_ema200":   bool(c8.iat[i]) if c8.iat[i] is not None else None,
                "macd_bull":      bool(c9.iat[i]) if c9.iat[i] is not None else None,
                "volume_surge":   bool(c10.iat[i]) if c10.iat[i] is not None else None,
                "ema_cross":      bool(c11.iat[i]) if c11.iat[i] is not None else None,
                "bb_upper_half":  bool(c12.iat[i]) if c12.iat[i] is not None else None,
                "stoch_bullish":  bool(c13.iat[i]) if c13.iat[i] is not None else None,
            },
        })
    return results


def scan_piotroski(df: pd.DataFrame) -> list[dict]:
    if df.empty:
        return []
    d = df.reset_index(drop=True)

    cmp  = _col(d, "cmp");   h52 = _col(d, "high_52w")
    e20  = _col(d, "ema_20"); e50 = _col(d, "ema_50")
    rsi  = _col(d, "rsi");   roe = _col(d, "roe")
    opm  = _col(d, "opm");   pb  = _col(d, "pb")
    de   = _col(d, "debt_to_equity")
    cr   = _col(d, "current_ratio"); rg = _col(d, "revenue_growth")
    vr   = _col(d, "volume_ratio")
    atr  = _col(d, "atr_14")
    stk  = _col(d, "stoch_k"); std = _col(d, "stoch_d")

    v1 = roe.notna(); v2 = de.notna(); v3 = opm.notna()
    v4 = rg.notna();  v5 = cmp.notna() & e50.notna()
    v6 = rsi.notna()
    v7 = cmp.notna() & h52.gt(0) & h52.notna()
    v8 = cr.notna();  v9 = pb.notna()
    v10 = e20.notna() & e50.notna()
    v11 = atr.notna() & cmp.gt(0)

    c1  = _vcrit(roe > 0, v1)
    c2  = _vcrit(de < 0.5, v2)
    c3  = _vcrit(opm > 15, v3)
    c4  = _vcrit(rg > 0, v4)
    c5  = _vcrit(cmp > e50, v5)
    c6  = _vcrit(rsi.between(40, 70), v6)
    c7  = _vcrit((h52 - cmp) / h52.where(h52 > 0) <= 0.20, v7)
    c8  = _vcrit(cr > 1.5, v8)
    c9  = _vcrit(pb < 3, v9)
    c10 = _vcrit(e20 > e50, v10)
    c11 = _vcrit((atr / cmp) * 100 < 3, v11)

    score = (_pts(c1) + _pts(c2) + _pts(c3) + _pts(c4) + _pts(c5) +
             _pts(c6) + _pts(c7) + _pts(c8) + _pts(c9) + _pts(c10) + _pts(c11))

    max_s = (v1.astype(int) + v2.astype(int) + v3.astype(int) + v4.astype(int) +
             v5.astype(int) + v6.astype(int) + v7.astype(int) + v8.astype(int) +
             v9.astype(int) + v10.astype(int) + v11.astype(int))

    buy_thr   = np.ceil(max_s * 0.65).astype(int)
    watch_thr = np.ceil(max_s * 0.40).astype(int)
    signal = np.where(score >= buy_thr, "BUY",
             np.where(score >= watch_thr, "WATCH", "AVOID"))

    tickers = d["ticker"].fillna("").astype(str).tolist() if "ticker" in d.columns else [""] * len(d)
    names   = d["name"].fillna("").astype(str).tolist()   if "name" in d.columns   else [""] * len(d)
    sectors = d["sector"].fillna("").astype(str).tolist() if "sector" in d.columns else [""] * len(d)

    results = []
    for i in range(len(d)):
        results.append({
            "ticker":    tickers[i],
            "name":      names[i],
            "sector":    sectors[i],
            "cmp":       _r(cmp.iat[i]),
            "rsi":       _r(rsi.iat[i]),
            "roe":       _r(roe.iat[i]),
            "opm":       _r(opm.iat[i]),
            "pe":        _r(_col(d, "pe").iat[i]),
            "score":     int(score.iat[i]),
            "max_score": int(max_s.iat[i]),
            "signal":    str(signal[i]),
        })
    return results


# ── report generation ─────────────────────────────────────────────────────────

def _tick(v: bool | None) -> str:
    if v is True:  return "✓"
    if v is False: return "✗"
    return "—"


def _arrow(v: float | None) -> str:
    if v is None: return ""
    return f"+{v:.2f}%" if v >= 0 else f"{v:.2f}%"


def print_briefing(
    index_data: dict,
    darvas_results: list[dict],
    piotroski_results: list[dict],
    as_of: datetime,
) -> None:

    W = 72
    print()
    print("═" * W)
    print(f"  🌅  INDIA MARKET MORNING BRIEFING  —  {as_of.strftime('%d %b %Y  %H:%M IST')}")
    print("═" * W)

    # ── 1. Index levels ───────────────────────────────────────────────────────
    print()
    print("  BENCHMARK LEVELS")
    print("  " + "─" * (W - 2))
    for name, data in index_data.items():
        level = f"{data['level']:>12,.2f}"
        chg   = f"  {data['trend']} {data['change']:+,.2f}  ({data['chg_pct']:+.2f}%)"
        print(f"  {name:<16} {level}  {chg}")

    if not index_data:
        print("  [No index data available]")

    # ── 2. Darvas / Buffett BUYs ──────────────────────────────────────────────
    print()
    print("  DARVAS / BUFFETT SCAN")
    print("  " + "─" * (W - 2))

    buys   = [r for r in darvas_results if r["signal"] == "BUY"]
    watchs = [r for r in darvas_results if r["signal"] == "WATCH"]
    avoids = [r for r in darvas_results if r["signal"] == "AVOID"]

    print(f"  Scanned {len(darvas_results)} stocks  │  "
          f"BUY: {len(buys)}  WATCH: {len(watchs)}  AVOID: {len(avoids)}")
    print()

    if buys:
        print(f"  {'TICKER':<14}{'CMP':>8}{'RSI':>6}{'52W-Hi':>9}{'EMA50':>9}  "
              f"{'1D':>6}{'1W':>7}{'1M':>7}  SCORE  CRITERIA")
        print("  " + "─" * (W - 2))
        for r in sorted(buys, key=lambda x: -x["score"]):
            crit = r["criteria"]
            flags = (f"{_tick(crit.get('near_52w_high'))}H "
                     f"{_tick(crit.get('above_ema50'))}E50 "
                     f"{_tick(crit.get('rsi_healthy'))}RSI "
                     f"{_tick(crit.get('macd_bull'))}MAC "
                     f"{_tick(crit.get('above_ema200'))}E200 "
                     f"{_tick(crit.get('buffett_roe'))}ROE")
            print(f"  {r['ticker']:<14}"
                  f"{r['cmp'] or 0:>8.2f}"
                  f"{r['rsi'] or 0:>6.1f}"
                  f"{r['high_52w'] or 0:>9.2f}"
                  f"{r['ema_50'] or 0:>9.2f}"
                  f"  {_arrow(r.get('ret_1d')):>6}"
                  f"{_arrow(r.get('ret_1w')):>7}"
                  f"{_arrow(r.get('ret_1m')):>7}"
                  f"  {r['score']:>3}/13  {flags}")
    else:
        print("  No BUY signals today.")

    if watchs:
        print()
        print(f"  WATCH ({len(watchs)}): " + ", ".join(r["ticker"] for r in watchs[:15]))

    # ── 3. Piotroski BUYs ────────────────────────────────────────────────────
    print()
    print("  PIOTROSKI SCAN")
    print("  " + "─" * (W - 2))

    p_buys   = [r for r in piotroski_results if r["signal"] == "BUY"]
    p_watchs = [r for r in piotroski_results if r["signal"] == "WATCH"]

    print(f"  BUY: {len(p_buys)}  WATCH: {len(p_watchs)}")
    print()

    if p_buys:
        print(f"  {'TICKER':<14}{'CMP':>8}{'RSI':>6}{'ROE%':>7}{'OPM%':>7}{'PE':>6}  SCORE")
        print("  " + "─" * (W - 2))
        for r in sorted(p_buys, key=lambda x: -x["score"]):
            print(f"  {r['ticker']:<14}"
                  f"{r['cmp'] or 0:>8.2f}"
                  f"{r['rsi'] or 0:>6.1f}"
                  f"{r['roe'] or 0:>7.1f}"
                  f"{r['opm'] or 0:>7.1f}"
                  f"{r['pe'] or 0:>6.1f}"
                  f"  {r['score']}/{r['max_score']}")
    else:
        print("  No Piotroski BUY signals today.")

    # ── 4. Notable movers ─────────────────────────────────────────────────────
    print()
    print("  NOTABLE MOVERS (1-DAY)")
    print("  " + "─" * (W - 2))

    gainers = sorted(
        [r for r in darvas_results if r.get("ret_1d") is not None],
        key=lambda x: -x["ret_1d"]
    )[:5]
    losers = sorted(
        [r for r in darvas_results if r.get("ret_1d") is not None],
        key=lambda x: x["ret_1d"]
    )[:5]

    print(f"  Top Gainers: " +
          "  ".join(f"{r['ticker']} ({_arrow(r['ret_1d'])})" for r in gainers))
    print(f"  Top Losers:  " +
          "  ".join(f"{r['ticker']} ({_arrow(r['ret_1d'])})" for r in losers))

    # ── 5. RSI alerts ────────────────────────────────────────────────────────
    print()
    print("  RSI ALERTS")
    print("  " + "─" * (W - 2))

    overbought = [r for r in darvas_results if r.get("rsi") and r["rsi"] >= 70]
    oversold   = [r for r in darvas_results if r.get("rsi") and r["rsi"] <= 30]

    if overbought:
        print("  Overbought (RSI ≥ 70): " +
              ", ".join(f"{r['ticker']} ({r['rsi']:.0f})" for r in overbought))
    if oversold:
        print("  Oversold   (RSI ≤ 30): " +
              ", ".join(f"{r['ticker']} ({r['rsi']:.0f})" for r in oversold))
    if not overbought and not oversold:
        print("  No extreme RSI readings today.")

    print()
    print("═" * W)
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
          f"Source: yfinance  |  herrrickshaw")
    print("═" * W)
    print()


def write_excel_report(
    index_data: dict,
    darvas_results: list[dict],
    piotroski_results: list[dict],
    as_of: datetime,
    xl_path: Path,
) -> None:
    """
    Append today's results as a new date-named sheet in the rolling Excel workbook.
    File: reports/morning_briefing_history.xlsx
    Sheet name: DD-Mon-YY  (e.g. 23-Jun-26)
    If a sheet for today already exists it is replaced.
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed — skipping Excel output.")
        print("         pip install openpyxl")
        return

    sheet_name = as_of.strftime("%d-%b-%y")   # e.g. 23-Jun-26

    # Load existing workbook or create fresh one
    if xl_path.exists():
        wb = load_workbook(xl_path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]             # replace today's sheet if rerun
        ws = wb.create_sheet(sheet_name, 0)  # insert at front
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

    # ── colour palette ────────────────────────────────────────────────────────
    C_HEADER_BG  = "1E3A5F"   # dark navy
    C_HEADER_FG  = "FFFFFF"
    C_TITLE_BG   = "0D2137"
    C_SECTION_BG = "2B4C7E"
    C_BUY_BG     = "C6EFCE"   # light green
    C_BUY_FG     = "276221"
    C_WATCH_BG   = "FFEB9C"   # amber
    C_WATCH_FG   = "9C6500"
    C_AVOID_BG   = "FFC7CE"   # pink-red
    C_AVOID_FG   = "9C0006"
    C_POS_FG     = "276221"
    C_NEG_FG     = "9C0006"
    C_ALT_BG     = "F2F7FC"   # subtle alternating row

    FONT_BODY    = "Arial"
    FONT_MONO    = "Courier New"

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(bold=True, color=C_HEADER_FG, size=10):
        return Font(name=FONT_BODY, bold=bold, color=color, size=size)

    def body_font(bold=False, color="000000", size=10, italic=False):
        return Font(name=FONT_BODY, bold=bold, color=color, size=size, italic=italic)

    def mono_font(size=10, color="000000"):
        return Font(name=FONT_MONO, size=size, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=False)

    def right():
        return Alignment(horizontal="right", vertical="center")

    def left():
        return Alignment(horizontal="left", vertical="center")

    def signal_style(sig: str):
        m = {"BUY": (C_BUY_BG, C_BUY_FG),
             "WATCH": (C_WATCH_BG, C_WATCH_FG),
             "AVOID": (C_AVOID_BG, C_AVOID_FG)}
        bg, fg = m.get(sig, ("FFFFFF", "000000"))
        return fill(bg), Font(name=FONT_BODY, bold=True, color=fg, size=10)

    row = 1

    # ── TITLE ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:M{row}")
    c = ws.cell(row, 1, f"India Market Morning Briefing  —  {as_of.strftime('%A, %d %B %Y')}")
    c.font = Font(name=FONT_BODY, bold=True, size=14, color=C_HEADER_FG)
    c.fill = fill(C_TITLE_BG)
    c.alignment = center()
    ws.row_dimensions[row].height = 26
    row += 1

    ws.merge_cells(f"A{row}:M{row}")
    c = ws.cell(row, 1, f"Generated: {as_of.strftime('%H:%M')}  |  Source: yfinance  |  herrrickshaw")
    c.font = Font(name=FONT_BODY, italic=True, size=9, color="AAAAAA")
    c.fill = fill(C_TITLE_BG)
    c.alignment = center()
    ws.row_dimensions[row].height = 16
    row += 2

    # ── SECTION: INDEX LEVELS ─────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:M{row}")
    c = ws.cell(row, 1, "BENCHMARK LEVELS")
    c.font = Font(name=FONT_BODY, bold=True, size=11, color=C_HEADER_FG)
    c.fill = fill(C_SECTION_BG)
    c.alignment = left()
    ws.row_dimensions[row].height = 20
    row += 1

    idx_headers = ["Index", "Level", "Change", "Change %"]
    for col_i, h in enumerate(idx_headers, 1):
        c = ws.cell(row, col_i, h)
        c.font = hdr_font()
        c.fill = fill(C_HEADER_BG)
        c.alignment = center()
        c.border = bdr
    ws.row_dimensions[row].height = 18
    row += 1

    for name, d in index_data.items():
        chg_fg = C_POS_FG if d["change"] >= 0 else C_NEG_FG
        vals = [name, d["level"], d["change"], d["chg_pct"] / 100]
        fmts = [None, "#,##0.00", "+#,##0.00;-#,##0.00", "+0.00%;-0.00%"]
        for col_i, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row, col_i, v)
            c.font = mono_font(color=(chg_fg if col_i > 1 else "000000"))
            c.alignment = right() if col_i > 1 else left()
            c.border = bdr
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[row].height = 16
        row += 1
    row += 1

    # ── SECTION: DARVAS / BUFFETT ─────────────────────────────────────────────
    ws.merge_cells(f"A{row}:M{row}")
    c = ws.cell(row, 1, "DARVAS / BUFFETT SCAN")
    c.font = Font(name=FONT_BODY, bold=True, size=11, color=C_HEADER_FG)
    c.fill = fill(C_SECTION_BG)
    c.alignment = left()
    ws.row_dimensions[row].height = 20
    row += 1

    darvas_cols = [
        ("Ticker",    12, left()),
        ("Name",      28, left()),
        ("Sector",    18, left()),
        ("CMP (₹)",   10, right()),
        ("52W High",  10, right()),
        ("EMA 50",    10, right()),
        ("RSI",        7, right()),
        ("Ret 1D %",   9, right()),
        ("Ret 1W %",   9, right()),
        ("Ret 1M %",   9, right()),
        ("Score /13",  9, center()),
        ("Signal",    10, center()),
        ("Criteria",  32, left()),
    ]

    for col_i, (h, w, _) in enumerate(darvas_cols, 1):
        c = ws.cell(row, col_i, h)
        c.font = hdr_font()
        c.fill = fill(C_HEADER_BG)
        c.alignment = center()
        c.border = bdr
        ws.column_dimensions[get_column_letter(col_i)].width = w
    ws.row_dimensions[row].height = 18
    row += 1

    for i, r in enumerate(sorted(darvas_results,
                                 key=lambda x: ({"BUY": 0, "WATCH": 1, "AVOID": 2}
                                                .get(x["signal"], 3), -x["score"]))):
        bg_hex = C_ALT_BG if i % 2 == 1 else "FFFFFF"
        crit   = r.get("criteria", {})
        crit_str = "  ".join([
            "✓H"    if crit.get("near_52w_high") else "✗H",
            "✓E50"  if crit.get("above_ema50")   else "✗E50",
            "✓RSI"  if crit.get("rsi_healthy")   else "✗RSI",
            "✓MACD" if crit.get("macd_bull")     else "✗MACD",
            "✓E200" if crit.get("above_ema200")  else "✗E200",
            "✓ROE"  if crit.get("buffett_roe")   else "✗ROE",
            "✓PE"   if crit.get("buffett_pe")    else "✗PE",
            "✓D/E"  if crit.get("buffett_de")    else "✗D/E",
        ])
        row_vals = [
            r["ticker"], r.get("name", ""), r.get("sector", ""),
            r["cmp"], r.get("high_52w"), r.get("ema_50"), r.get("rsi"),
            (r.get("ret_1d") or 0) / 100 if r.get("ret_1d") is not None else None,
            (r.get("ret_1w") or 0) / 100 if r.get("ret_1w") is not None else None,
            (r.get("ret_1m") or 0) / 100 if r.get("ret_1m") is not None else None,
            f"{r['score']}/13",
            r["signal"],
            crit_str,
        ]
        fmts = [None, None, None,
                "#,##0.00", "#,##0.00", "#,##0.00", "0.0",
                "+0.00%;-0.00%", "+0.00%;-0.00%", "+0.00%;-0.00%",
                None, None, None]
        aligns = [col[2] for col in darvas_cols]

        for col_i, (v, fmt, aln) in enumerate(zip(row_vals, fmts, aligns), 1):
            c = ws.cell(row, col_i, v)
            c.border = bdr
            c.alignment = aln
            if fmt:
                c.number_format = fmt

            if col_i == 12:  # Signal column
                sfill, sfont = signal_style(r["signal"])
                c.fill = sfill
                c.font = sfont
            elif col_i in (8, 9, 10) and v is not None:  # return cols
                pct_val = r.get(["ret_1d", "ret_1w", "ret_1m"][col_i - 8])
                fg = C_POS_FG if (pct_val or 0) >= 0 else C_NEG_FG
                c.font = mono_font(color=fg)
                c.fill = fill(bg_hex)
            elif col_i == 1:
                c.font = Font(name=FONT_BODY, bold=True, size=10)
                c.fill = fill(bg_hex)
            else:
                c.font = body_font() if col_i not in (4, 5, 6, 7) else mono_font()
                c.fill = fill(bg_hex)

        ws.row_dimensions[row].height = 16
        row += 1
    row += 1

    # ── SECTION: PIOTROSKI ────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:M{row}")
    c = ws.cell(row, 1, "PIOTROSKI SCAN")
    c.font = Font(name=FONT_BODY, bold=True, size=11, color=C_HEADER_FG)
    c.fill = fill(C_SECTION_BG)
    c.alignment = left()
    ws.row_dimensions[row].height = 20
    row += 1

    pio_cols = [
        ("Ticker",     12), ("Name", 28), ("Sector", 18),
        ("CMP (₹)",    10), ("RSI",   7), ("ROE %",  9),
        ("OPM %",       9), ("P/E",   8), ("D/E",    8),
        ("Score",       9), ("Signal", 10),
    ]

    for col_i, (h, _) in enumerate(pio_cols, 1):
        c = ws.cell(row, col_i, h)
        c.font = hdr_font()
        c.fill = fill(C_HEADER_BG)
        c.alignment = center()
        c.border = bdr
    ws.row_dimensions[row].height = 18
    row += 1

    for i, r in enumerate(sorted(piotroski_results,
                                 key=lambda x: ({"BUY": 0, "WATCH": 1, "AVOID": 2}
                                                .get(x["signal"], 3), -x["score"]))):
        bg_hex = C_ALT_BG if i % 2 == 1 else "FFFFFF"
        row_vals = [
            r["ticker"], r.get("name", ""), r.get("sector", ""),
            r.get("cmp"), r.get("rsi"), r.get("roe"),
            r.get("opm"), r.get("pe"), r.get("debt_to_equity", None),
            f"{r['score']}/{r['max_score']}", r["signal"],
        ]
        fmts = [None, None, None,
                "#,##0.00", "0.0", "0.0%", "0.0%", "0.0x", "0.0x",
                None, None]

        for col_i, (v, fmt) in enumerate(zip(row_vals, fmts), 1):
            # ROE and OPM are stored as plain % values (e.g. 18.5 means 18.5%)
            # convert to decimal for % format
            if col_i in (6, 7) and v is not None:
                v = v / 100
            c = ws.cell(row, col_i, v)
            c.border = bdr
            c.alignment = center() if col_i in (10, 11) else (left() if col_i <= 3 else right())
            if fmt:
                c.number_format = fmt
            if col_i == 11:
                sfill, sfont = signal_style(r["signal"])
                c.fill = sfill
                c.font = sfont
            elif col_i == 1:
                c.font = Font(name=FONT_BODY, bold=True, size=10)
                c.fill = fill(bg_hex)
            else:
                c.font = mono_font() if col_i > 3 else body_font()
                c.fill = fill(bg_hex)

        ws.row_dimensions[row].height = 16
        row += 1

    # ── freeze top rows, auto-filter on darvas table ──────────────────────────
    ws.freeze_panes = "A3"

    wb.save(xl_path)
    print(f"  Excel report saved → {xl_path}  (sheet: {sheet_name})")


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="India market morning briefing")
    parser.add_argument("--symbols", nargs="+", metavar="SYM",
                        help="NSE symbols to scan (default: 20-stock watchlist)")
    parser.add_argument("--nifty50", action="store_true",
                        help="Scan full NIFTY 50 universe (slower)")
    parser.add_argument("--top", type=int, default=0, metavar="N",
                        help="Scan first N stocks from NIFTY 50")
    parser.add_argument("--no-excel", action="store_true",
                        help="Skip writing the Excel report (print-only mode)")
    parser.add_argument("--no-fundamentals", action="store_true",
                        help="Skip yfinance .info call (faster, no PE/ROE/D-E data)")
    args = parser.parse_args()

    if args.symbols:
        universe = [s.upper() for s in args.symbols]
    elif args.nifty50 or args.top:
        universe = NIFTY50_SYMBOLS[:args.top] if args.top else NIFTY50_SYMBOLS
    else:
        universe = DEFAULT_WATCHLIST

    as_of = datetime.now()
    print(f"\n  Scanning {len(universe)} stocks as of {as_of.strftime('%d %b %Y %H:%M')} …")

    # ── fetch index data ──────────────────────────────────────────────────────
    print("  [1/3] Fetching index levels …")
    index_data = fetch_index_data()

    # ── fetch stock data ──────────────────────────────────────────────────────
    print(f"  [2/3] Fetching stock data ({len(universe)} symbols) …")
    records = []
    for i, sym in enumerate(universe, 1):
        print(f"        {i:>3}/{len(universe)}  {sym:<16}", end="\r", flush=True)
        data = fetch_stock_data(sym)
        if data:
            if args.no_fundamentals:
                for key in ("pe", "pb", "roe", "opm", "debt_to_equity",
                            "current_ratio", "revenue_growth", "eps", "beta",
                            "dividend_yield", "market_cap"):
                    data[key] = None
            records.append(data)
    print(f"        Fetched {len(records)}/{len(universe)} stocks successfully.   ")

    if not records:
        print("  No data fetched. Check your internet connection.")
        sys.exit(1)

    df = pd.DataFrame(records)

    # ── run scanners ──────────────────────────────────────────────────────────
    print("  [3/3] Running scanners …")
    darvas_results     = scan_darvas(df)
    piotroski_results  = scan_piotroski(df)

    # ── print report ──────────────────────────────────────────────────────────
    print_briefing(index_data, darvas_results, piotroski_results, as_of)

    # ── Excel report (always, unless --no-excel) ──────────────────────────────
    if not args.no_excel:
        reports_dir = Path(__file__).parent / "reports"
        reports_dir.mkdir(exist_ok=True)
        xl_path = reports_dir / "morning_briefing_history.xlsx"
        write_excel_report(index_data, darvas_results, piotroski_results, as_of, xl_path)


if __name__ == "__main__":
    main()
