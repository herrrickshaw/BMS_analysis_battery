#!/usr/bin/env python3
"""
nifty500_scan.py
================
Full NIFTY 500 morning scan — batched yfinance download, Darvas/Buffett +
Piotroski scoring, date-wise Excel output.

Usage:
    python3 nifty500_scan.py                      # full NIFTY 500 scan
    python3 nifty500_scan.py --fast               # skip fundamentals (~3x faster)
    python3 nifty500_scan.py --batch-size 50      # tickers per yf.download() call
    python3 nifty500_scan.py --output custom.xlsx # override output path
    python3 nifty500_scan.py --top 100            # scan only first N symbols (test run)
    python3 nifty500_scan.py --gems-only          # only run the hidden gems scan
    python3 nifty500_scan.py --no-gems            # skip hidden gems scan

Output:
    reports/nifty500_scan_history.xlsx
      Sheet "Summary"      — index levels + today's scan stats    (replaced each run)
      Sheet "Darvas"       — all 500 Darvas/Buffett results        (replaced each run)
      Sheet "Piotroski"   — all 500 Piotroski results             (replaced each run)
      Sheet "Hidden Gems"  — BUY/WATCH stocks outside NIFTY 500   (replaced each run)
      Sheet "History"      — one summary row appended per run      (cumulative)

Dependencies (standard PyPI — no Claude, no Anthropic SDK required):
    pip3 install yfinance pandas numpy openpyxl requests

Standalone execution (no Claude Code, no app server needed):
    python3 nifty500_scan.py --fast
"""
from __future__ import annotations

import argparse
import math
import sys
import time
import warnings
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

warnings.filterwarnings("ignore")

# Optional Darvas interpreter integration
try:
    from darvas_interpreter import interpret_dataframe as _darvas_interpret
    _HAS_DARVAS = True
except ImportError:
    _HAS_DARVAS = False

try:
    import numpy as np
    import pandas as pd
except ImportError:
    sys.exit("Install dependencies:  pip install yfinance pandas numpy openpyxl requests")

try:
    import yfinance as yf
except ImportError:
    sys.exit("pip install yfinance")

try:
    import requests
except ImportError:
    requests = None  # type: ignore

# ── NIFTY 500 symbol list ─────────────────────────────────────────────────────
# Fetched live from NSE when possible; this hardcoded list is the fallback.

NIFTY500_FALLBACK = [
    # NIFTY 50
    "ADANIENT","ADANIPORTS","APOLLOHOSP","ASIANPAINT","AXISBANK",
    "BAJAJ-AUTO","BAJFINANCE","BAJAJFINSV","BPCL","BHARTIARTL",
    "BRITANNIA","CIPLA","COALINDIA","DIVISLAB","DRREDDY",
    "EICHERMOT","GRASIM","HCLTECH","HDFCBANK","HDFCLIFE",
    "HEROMOTOCO","HINDALCO","HINDUNILVR","ICICIBANK","ITC",
    "INDUSINDBK","INFY","JSWSTEEL","KOTAKBANK","LT",
    "M&M","MARUTI","NTPC","NESTLEIND","ONGC",
    "POWERGRID","RELIANCE","SBILIFE","SHRIRAMFIN","SBIN",
    "SUNPHARMA","TCS","TATACONSUM","TATAMOTORS","TATASTEEL",
    "TECHM","TITAN","ULTRACEMCO","WIPRO","ZOMATO",
    # NIFTY Next 50
    "ABB","ADANIGREEN","ADANIPOWER","AMBUJACEM","AUBANK",
    "AUROPHARMA","BANKBARODA","BEL","BERGEPAINT","CANBK",
    "CHOLAFIN","COLPAL","DLF","GAIL","GODREJCP",
    "GODREJPROP","HAVELLS","ICICIGI","ICICIPRULI","IOC",
    "IRCTC","JINDALSTEL","LICI","LODHA","LTIM",
    "LUPIN","MANKIND","MARICO","MUTHOOTFIN","NHPC",
    "PFC","PIDILITIND","RECLTD","SAILREF","SIEMENS",
    "SRF","TORNTPHARM","TRENT","VBL","VEDL",
    "VOLTAS","YESBANK","ZYDUSLIFE","MCDOWELL-N","ADANITRANS",
    "ATUL","BOSCHLTD","CGPOWER","CUMMINSIND","PGHH",
    # NIFTY Midcap 150
    "AARTIIND","ABCAPITAL","ABFRL","ACC","AIAENG",
    "AJANTPHARM","ALKEM","APOLLOTYRE","ASHOKLEY","ASTRAL",
    "AVANADCOMM","BAJAJHFL","BALKRISIND","BANDHANBNK","BATAINDIA",
    "BHEL","BIOCON","BLUESTAR","BSE","BSOFT",
    "CANFINHOME","CEATLTD","CENTRALBK","CHAMBLFERT","COFORGE",
    "CROMPTON","CSBBANK","CYIENT","DALBHARAT","DEEPAKNTR",
    "DELTACORP","DIXON","DMART","EIDPARRY","ELGIEQUIP",
    "ENGINERSIN","ESCORTS","EXIDEIND","FEDERALBNK","FINPIPE",
    "FLUOROCHEM","FORTIS","GLENMARK","GNFC","GRINDWELL",
    "GSPL","HFCL","HINDPETRO","HONAUT","IBREALEST",
    "IDBI","IDFC","IDFCFIRSTB","IEX","IFCI",
    "IPCALAB","IRCON","JKCEMENT","JKLAKSHMI","JUBLFOOD",
    "KAJARIACER","KALPATPOWR","KEC","KPITTECH","KRBL",
    "KSCL","LALPATHLAB","LAURUSLABS","LICHSGFIN","LINDEINDIA",
    "MANAPPURAM","MFSL","MPHASIS","MRF","MTAR",
    "NATCOPHARM","NAVINFLUOR","NAUKRI","NFL","NMDC",
    "NYKAA","OBEROIRLTY","OFSS","OIL","PAGEIND",
    "PATANJALI","PCJEWELLER","PERSISTENT","PNB","POLYCAB",
    "PRESTIGE","PRINCEPIPES","PTCIL","PVRINOX","RAJESHEXPO",
    "RAMCOCEM","RATNAMANI","REDINGTON","RELAXO","RITES",
    "RPOWER","SAFARI","SANOFI","SCHAEFFLER","SKFINDIA",
    "SONACOMS","STARHEALTH","SUDARSCHEM","SUMICHEM","SUNCLAYLTD",
    "SUNTV","SUPREMEIND","SUVENPHARMA","SYMPHONY","TANLA",
    "TATACHEM","TATACOMM","TATAELXSI","TATAINVEST","TATAPOWER",
    "TEAMLEASE","THYROCARE","TIINDIA","TITAGARH","TORNTPOWER",
    "TTKPRESTIG","TVSHMOTOR","UBL","UCOBANK","UNIONBANK",
    "UNOMINDA","USHAMART","VGUARD","VINATIORGA","WELCORP",
    "WESTLIFE","WHIRLPOOL","ZEEL","ZENTEC","ZODIACLOTH",
    # NIFTY Smallcap 250 (representative)
    "AARTIDRUGS","AAVAS","ACCELYA","ACMESOLAR","ADANIENSOL",
    "ADFFOODS","ADVENZYMES","AGI","AGROPHOS","AHMEDABADBE",
    "AIAENG","AJANTTEX","AKZOINDIA","ALEMBICLTD","ALICON",
    "ALKYLAMINE","ALLCARGO","AMARAJABAT","AMBIKCO","AMNPLST",
    "ANANTRAJ","ANDHRSUGAR","ANGELONE","ANURAS","APCOTEXIND",
    "APEX","APTUS","ARCHIES","ARFIN","AROGRANITE",
    "ARSHIYA","ARVINDSMRT","ASAHIINDIA","ASHIANA","ASHIMASYN",
    "ASPINWALL","ASTEC","ASTRAZEN","ATGL","ATUL",
    "AURIONPRO","AVANTIFEED","AXISCADES","BABAFP","BAJAJCON",
    "BAJAJHIND","BAJEL","BALAJITELE","BALMLAWRIE","BALPHARMA",
    "BALUARTE","BARAMATI","BASF","BAYERCROP","BCCL",
    "BECIL","BEPL","BFINANCE","BFULLEN","BIGBLOC",
    "BIMETAL","BIRLASOFT","BIRIYANIS","BLS","BOROLTD",
    "BPCL","BSLIMITED","CAMLINFINE","CAPACITE","CAPRIHANS",
    "CARYSIL","CASTROLIND","CCL","CESC","CHEMPLASTS",
    "CHENNAISUPR","CHEVIOT","CINELINE","CLEAN","CLNINDIA",
    "CMSINFO","COCHINSHIP","CONTROLPR","COSMOFILMS","CPLILTD",
    "CRAFTSMAN","CREDITACC","CRISIL","DATAPATTNS","DBCORP",
    "DBSTOCKBRO","DBREALTY","DCB","DCCQUEST","DCMSHRIRAM",
    "DEEPIND","DELTACORP","DHANI","DHARMAJ","DHUNINV",
    "DICIND","DLINKINDIA","DNAINDIA","DOLLAR","DPWWORLD",
    "DREDGECORP","DSPAC","ECLERX","EDELWEISS","ELGIRUBCO",
    "EMAMILTD","EMAMIREAL","ENDURANCE","ENIL","EPACK",
    "EPIGRAL","EQUITASBNK","ESABINDIA","ESTER","ETHOSLTD",
    "EVEREADY","EXCEL","EXCELLIND","EXICOM","FINEORG",
    "FINPIPE","FINOFINANCE","FLAIR","FMGOETZE","FOODWORKS",
    "FORCEMOT","GANECOS","GANESHHOUC","GESHIP","GHCL",
    "GICHSGFIN","GILLETTE","GLAND","GLAXO","GLOBUSSPR",
    "GMMPFAUDLR","GODFRYPHLP","GOLDIAM","GPPL","GRANULES",
    "GRAVITA","GREENLAM","GREENPLY","GRSE","GTLINFRA",
    "GUFICBIO","GULFOILLUB","HAPPSTMNDS","HARDWYN","HDFC",
    "HECPROJECT","HEGDE","HEIDELBERG","HESTERBIO","HIMATSEIDE",
    "HINDCOPPER","HINDOILEXP","HINDZINC","HITECH","HLEGLAS",
    "HMT","HOLCIM","HOMEFIRST","HONASA","HUHTAMAKI",
    "IBULHSGFIN","ICICISEC","IDBI","IGARASHI","IGPL",
    "IIFLSEC","IMAGICAWORLD","IMPAL","INDHOTEL","INDIGOPNTS",
    "INDOSTAR","INDSWFTLAB","INFIBEAM","INGERRAND","INOXWIND",
    "INTELLECT","INTENTECH","IOB","IOLCP","ISGEC",
    "ITI","IXIGO","JAIBALAJI","JAMNAAUTO","JAYAGROGN",
    "JBCHEPHARM","JBMA","JCHAC","JHS","JINDALSAW",
    "JKIL","JKPAPER","JLHL","JMFINANCIL","JPASSOCIAT",
    "JSWENERGY","JTEKTINDIA","JUBLINGREA","JUNIPERIND","JUSTDIAL",
    "JYOTICNC","KANSAINER","KAYNES","KDDL","KELLTONTEC",
    "KENNAMET","KESORAMIND","KFINTECH","KHADIM","KILITCH",
    "KIRIINDUS","KIRLOSBROS","KIRLPNU","KITEX","KNR",
    "KOVAI","KPIL","KRSNAA","KSOLVES","KUANTUM",
    "LAXMIMACH","LFLOGISTIC","LGBBROSLTD","LLOYDMETAL","LODHA",
    "LOKESHMA","LSIL","LUMAXIND","LUMAXTECH","LUXIND",
    "MAANALU","MAFANG","MAHAPEXLTD","MAHABANK","MAHINDCIE",
    "MAHLIFE","MAHSCOOTER","MAHSEAMLES","MAITHANALL","MAPMYINDIA",
    "MASTEK","MAYURUNIQ","MBLINFRA","MEDANTA","MEDPLUS",
    "METROPOLIS","MIDHANI","MOLDTKPAC","MOTILALOFS","MSTCLTD",
    "MUKANDLTD","MUNJALSHOW","MURUDCERA","NACLIND","NAMDHARI",
    "NATCOPHARM","NAVA","NAVINFLUOR","NDTV","NEULANDLAB",
    "NEWGEN","NIITLTD","NILKAMAL","NLCINDIA","NOCIL",
    "NRBBEARING","NUVAMA","NUVOCO","OLECTRA","OMAXE",
    "ONEPOINT","OPTIEMUS","ORCHPHARMA","ORIENTELEC","PANAMAPET",
    "PARADEEP","PARAS","PARASHE","PATELENG","PATSPINN",
    "PENIND","PERI","PFIZER","PHOENIXLTD","PIIND",
    "PILANCITA","PILANIINVS","PITTIENG","PNB","PNBGILTS",
    "PNBHOUSING","POLYMED","PRAJ","PRECWIRE","PREMEXPLN",
    "PRICOLLTD","PRINCEPIPES","PRISM","PRIVISCL","PROCTER",
    "PURVA","PVSL","RAJAPALAYAM","RAJRATAN","RAJSREESUG",
    "RAMASTEEL","RANEHOLDIN","RANEENGINE","RANKT","RATEGAIN",
    "RATNAMANI","RAYMOND","RBLBANK","RCAPITEXT","RECLTD",
    "REFEX","REGENCERAM","REMSONSIND","REPCO","RESPONIND",
    "RIIL","RITCO","RKDL","RKFORGE","RMC",
    "ROHITFERRO","ROLEXRINGS","ROSSARI","ROUTE","RPGLIFE",
    "RPPOWER","RSWM","RTNPOWER","RUBYMILLS","RUPA",
    "RUSHIL","SAGCEM","SAHAJAAND","SAKAR","SALSTEEL",
    "SAMHI","SANGAMIND","SANOFI","SARDAEN","SATIN",
    "SBICARD","SBIMF","SEQUENT","SERVOTECH","SFIL",
    "SGIL","SHAKTIPUMP","SHALBY","SHALPAINTS","SHANKARA",
    "SHARDACRP","SHARDAMOTR","SHAREINDIA","SHILPAMED","SHOPERSTOP",
    "SHREDIGCEM","SHREERAMA","SHYAMMETL","SIGACHI","SILVERTUC",
    "SJVN","SKIPPER","SKYGOLD","SMLISUZU","SNOWMAN",
    "SOLARA","SONATSOFTW","SOTL","SOUTHBANK","SPECCHEMLTD",
    "SPENCERS","SRTRANSFIN","STARCEMENT","STCINDIA","STEELCAS",
    "STERLITE","STLTECH","SUBROS","SUDARSCHEM","SUNDRM",
    "SUNFLAG","SUNPHARMA","SUPRAJIT","SURAJEST","SURYAROSNI",
    "SUTLEJTEX","SUVENPHARMA","SWSOLAR","SYNGENE","TAKE",
    "TANLA","TARSONS","TASTYBITE","TATAMETALI","TCIEXP",
    "TCNSCLOTHING","TEJASNET","TEXMOPIPES","THERMAX","THINKINK",
    "THOMASCOOK","TIMETECHNO","TIMKEN","TINPLATE","TIRUMALCHM",
    "TMLIND","TORENT","TORNTPHARM","TPLPLASTEH","TREEHOUSE",
    "TRIDENT","TRIVENI","TTKPRESTIG","TTML","TTNP",
    "TVTODAY","UGARSUGAR","UJJIVANSFB","ULTRACEMCO","UMANG",
    "UNIVASTU","UNIVCABLES","UTIAMC","UTTAMSUGAR","V2RETAIL",
    "VAIBHAVGBL","VALDEL","VARDHMAN","VIPIND","VIPCLOTHNG",
    "VISAKAIND","VISHWARAJ","VSTIND","VSTIL","VMART",
    "WATERBASE","WELSPUNLIV","WENDT","WOCKPHARMA","WONDERLA",
    "XCHANGING","XPRO","YATHARTH","ZENSAR","ZUARI",
]

# deduplicate while preserving order
_seen: set[str] = set()
NIFTY500_FALLBACK = [s for s in NIFTY500_FALLBACK
                     if s not in _seen and not _seen.add(s)]  # type: ignore


# ── fetch live NIFTY 500 list from NSE ───────────────────────────────────────

def fetch_nifty500_from_nse() -> list[str]:
    """Try to download the live NIFTY 500 index CSV from NSE."""
    if requests is None:
        return []
    url = ("https://nsearchives.nseindia.com/content/indices/"
           "ind_nifty500list.csv")
    headers = {
        "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 Chrome/124 Safari/537.36"),
        "Referer": "https://www.nseindia.com/",
    }
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
        df = pd.read_csv(pd.io.common.StringIO(resp.text))
        col = next((c for c in df.columns if "symbol" in c.lower()), None)
        if col:
            syms = df[col].dropna().str.strip().tolist()
            print(f"  Fetched {len(syms)} symbols from NSE live index CSV.")
            return syms
    except Exception as exc:
        print(f"  [WARN] NSE live fetch failed ({exc}), using fallback list.")
    return []


def load_nifty500(data_dir: Path) -> list[str]:
    """
    Priority:
    1. Live NSE index CSV
    2. Local ind_nifty500list.csv (if previously saved)
    3. Hardcoded fallback list
    """
    syms = fetch_nifty500_from_nse()
    if syms:
        # Cache for offline use
        try:
            cache = data_dir / "ind_nifty500list.csv"
            pd.DataFrame({"Symbol": syms}).to_csv(cache, index=False)
        except Exception:
            pass
        return syms

    cache = data_dir / "ind_nifty500list.csv"
    if cache.exists():
        try:
            df = pd.read_csv(cache)
            col = next((c for c in df.columns if "symbol" in c.lower()), None)
            if col:
                syms = df[col].dropna().str.strip().tolist()
                print(f"  Loaded {len(syms)} symbols from cached index file.")
                return syms
        except Exception:
            pass

    print(f"  Using hardcoded NIFTY 500 fallback list ({len(NIFTY500_FALLBACK)} symbols).")
    return NIFTY500_FALLBACK


# ── technical indicators ─────────────────────────────────────────────────────

def _rsi(closes: pd.Series, period: int = 14) -> Optional[float]:
    if len(closes) < period + 1:
        return None
    delta = closes.diff().dropna()
    gain = delta.clip(lower=0).ewm(com=period - 1, min_periods=period).mean()
    loss = (-delta).clip(lower=0).ewm(com=period - 1, min_periods=period).mean()
    rs = gain / loss.replace(0, float("nan"))
    return round(float((100 - 100 / (1 + rs)).iloc[-1]), 2)


def _ema(closes: pd.Series, period: int) -> Optional[float]:
    if len(closes) < period:
        return None
    return round(float(closes.ewm(span=period, adjust=False).mean().iloc[-1]), 2)


def _macd(closes: pd.Series) -> tuple[Optional[float], Optional[float]]:
    if len(closes) < 35:
        return None, None
    macd_line = (closes.ewm(span=12, adjust=False).mean()
                 - closes.ewm(span=26, adjust=False).mean())
    signal = macd_line.ewm(span=9, adjust=False).mean()
    return round(float(macd_line.iloc[-1]), 4), round(float(signal.iloc[-1]), 4)


def _bollinger(closes: pd.Series, period: int = 20
               ) -> tuple[Optional[float], Optional[float], Optional[float]]:
    if len(closes) < period:
        return None, None, None
    sma = closes.rolling(period).mean()
    std = closes.rolling(period).std()
    u, l = float((sma + 2 * std).iloc[-1]), float((sma - 2 * std).iloc[-1])
    c = float(closes.iloc[-1])
    pct = (c - l) / (u - l) if (u - l) != 0 else None
    return round(u, 2), round(l, 2), round(pct, 4) if pct is not None else None


def _stochastic(high: pd.Series, low: pd.Series, close: pd.Series,
                period: int = 14, smooth: int = 3
                ) -> tuple[Optional[float], Optional[float]]:
    if len(close) < period + smooth:
        return None, None
    lo_min = low.rolling(period).min()
    hi_max = high.rolling(period).max()
    denom = (hi_max - lo_min).replace(0, float("nan"))
    k = 100 * (close - lo_min) / denom
    ks = k.rolling(smooth).mean()
    ds = ks.rolling(smooth).mean()
    return round(float(ks.iloc[-1]), 2), round(float(ds.iloc[-1]), 2)


def _atr(high: pd.Series, low: pd.Series, close: pd.Series,
         period: int = 14) -> Optional[float]:
    if len(close) < period + 1:
        return None
    prev = close.shift(1)
    tr = pd.concat([high - low, (high - prev).abs(), (low - prev).abs()],
                   axis=1).max(axis=1)
    return round(float(tr.ewm(com=period - 1, min_periods=period).mean().iloc[-1]), 4)


def _vol_ratio(volume: pd.Series, period: int = 20) -> Optional[float]:
    if len(volume) < period + 1:
        return None
    avg = float(volume.rolling(period).mean().iloc[-1])
    return round(float(volume.iloc[-1]) / avg, 2) if avg else None


def _ret(closes: pd.Series, n: int) -> Optional[float]:
    if len(closes) < n + 1:
        return None
    s, e = float(closes.iloc[-(n + 1)]), float(closes.iloc[-1])
    return round((e - s) / s * 100, 2) if s else None


def _pct(v) -> Optional[float]:
    try:
        return round(float(v) * 100, 2) if v is not None else None
    except Exception:
        return None


def _f(v) -> Optional[float]:
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
    except Exception:
        return None


# ── batch OHLCV download ─────────────────────────────────────────────────────

def batch_download(symbols: list[str], period: str = "1y",
                   batch_size: int = 100) -> dict[str, pd.DataFrame]:
    """
    Download 1-year daily OHLCV for all symbols using yf.download() in batches.
    Returns dict: bare_symbol → DataFrame(Open,High,Low,Close,Volume).
    """
    tickers = [f"{s}.NS" for s in symbols]
    result: dict[str, pd.DataFrame] = {}

    for i in range(0, len(tickers), batch_size):
        batch = tickers[i: i + batch_size]
        syms  = symbols[i: i + batch_size]
        pct   = (i + len(batch)) / len(tickers) * 100
        print(f"  Downloading batch {i // batch_size + 1} "
              f"({i + 1}–{i + len(batch)}/{len(tickers)})  [{pct:.0f}%]", flush=True)

        try:
            raw = yf.download(
                batch, period=period, auto_adjust=True,
                progress=False, threads=True,
            )
            if raw.empty:
                continue

            # Multi-ticker download nests columns as (metric, ticker)
            if isinstance(raw.columns, pd.MultiIndex):
                for sym, ticker in zip(syms, batch):
                    try:
                        df = raw.xs(ticker, axis=1, level=1)[
                            ["Open", "High", "Low", "Close", "Volume"]
                        ].dropna(subset=["Close"])
                        if len(df) >= 30:
                            result[sym] = df
                    except Exception:
                        pass
            else:
                # Single ticker in batch
                df = raw[["Open", "High", "Low", "Close", "Volume"]].dropna(
                    subset=["Close"])
                if len(df) >= 30 and syms:
                    result[syms[0]] = df

        except Exception as exc:
            print(f"  [WARN] Batch {i // batch_size + 1} failed: {exc}")

        time.sleep(0.5)  # polite rate limiting

    return result


# ── build indicator record for one symbol ────────────────────────────────────

def build_record(sym: str, hist: pd.DataFrame, info: dict) -> dict:
    closes = hist["Close"]
    high   = hist["High"]
    low    = hist["Low"]
    volume = hist["Volume"]

    cmp      = round(float(closes.iloc[-1]), 2)
    high_52w = round(float(high.tail(252).max()), 2)
    low_52w  = round(float(low.tail(252).min()), 2)

    macd_v, macd_s = _macd(closes)
    bb_u, bb_l, bb_pct = _bollinger(closes)
    stk, std_v = _stochastic(high, low, closes)

    return {
        "ticker":         sym,
        "name":           info.get("longName") or info.get("shortName") or sym,
        "sector":         info.get("sector", ""),
        "industry":       info.get("industry", ""),
        "cmp":            cmp,
        "high_52w":       high_52w,
        "low_52w":        low_52w,
        "ema_20":         _ema(closes, 20),
        "ema_50":         _ema(closes, 50),
        "ema_200":        _ema(closes, 200),
        "rsi":            _rsi(closes),
        "macd":           macd_v,
        "macd_signal":    macd_s,
        "bb_upper":       bb_u,
        "bb_lower":       bb_l,
        "bb_pct":         bb_pct,
        "stoch_k":        stk,
        "stoch_d":        std_v,
        "atr_14":         _atr(high, low, closes),
        "volume_ratio":   _vol_ratio(volume),
        "volume":         int(volume.iloc[-1]),
        # fundamentals (None if --fast)
        "pe":             _f(info.get("trailingPE")),
        "pb":             _f(info.get("priceToBook")),
        "roe":            _pct(info.get("returnOnEquity")),
        "opm":            _pct(info.get("operatingMargins")),
        "debt_to_equity": _f(info.get("debtToEquity")),
        "market_cap":     info.get("marketCap"),
        "beta":           _f(info.get("beta")),
        "current_ratio":  _f(info.get("currentRatio")),
        "revenue_growth": _pct(info.get("revenueGrowth")),
        "eps":            _f(info.get("trailingEps")),
        "dividend_yield": _f(info.get("dividendYield")),
        # returns
        "ret_1d":  _ret(closes, 1),
        "ret_1w":  _ret(closes, 5),
        "ret_1m":  _ret(closes, 21),
        "ret_3m":  _ret(closes, 63),
        "ret_6m":  _ret(closes, 126),
        "ret_1y":  _ret(closes, 252),
        "_exchange": "NSE",
    }


# ── fetch fundamentals (optional) ────────────────────────────────────────────

def fetch_fundamentals_bulk(symbols: list[str],
                            batch_size: int = 10) -> dict[str, dict]:
    """
    Fetch .info for each symbol in small serial batches.
    Slow but gives PE/ROE/OPM/D-E data.
    """
    infos: dict[str, dict] = {}
    total = len(symbols)
    for i, sym in enumerate(symbols, 1):
        print(f"  Fundamentals {i}/{total}  {sym:<16}", end="\r", flush=True)
        try:
            info = yf.Ticker(f"{sym}.NS").info
            infos[sym] = info
        except Exception:
            infos[sym] = {}
        if i % batch_size == 0:
            time.sleep(0.3)
    print()
    return infos


# ── Darvas / Buffett scanner ──────────────────────────────────────────────────

def _col(df: pd.DataFrame, name: str) -> pd.Series:
    return pd.to_numeric(df[name], errors="coerce") if name in df.columns \
        else pd.Series(np.nan, index=df.index)


def _vcrit(passes: pd.Series, valid: pd.Series) -> pd.Series:
    out = pd.Series(None, index=passes.index, dtype=object)
    out[valid] = passes[valid]
    return out


def _pts(crit: pd.Series, weight: int = 1) -> pd.Series:
    return (crit == True).astype(int) * weight  # noqa: E712


def scan_darvas(df: pd.DataFrame) -> pd.DataFrame:
    d = df.reset_index(drop=True)
    cmp  = _col(d, "cmp");   h52 = _col(d, "high_52w"); l52 = _col(d, "low_52w")
    e20  = _col(d, "ema_20"); e50 = _col(d, "ema_50");  e200 = _col(d, "ema_200")
    rsi  = _col(d, "rsi");   roe = _col(d, "roe");      pe  = _col(d, "pe")
    de   = _col(d, "debt_to_equity")
    macd = _col(d, "macd");  msig = _col(d, "macd_signal")
    vr   = _col(d, "volume_ratio"); bb = _col(d, "bb_pct")
    stk  = _col(d, "stoch_k"); std = _col(d, "stoch_d")

    h52s = h52.where(h52 > 0)
    v1  = cmp.notna() & h52.gt(0); v2  = cmp.notna() & e50.notna()
    v3  = cmp.notna() & h52.notna() & l52.notna() & (h52 - l52).gt(0)
    v4  = rsi.notna(); v5 = roe.notna(); v6 = pe.notna(); v7 = de.notna()
    v8  = cmp.notna() & e200.notna(); v9  = macd.notna() & msig.notna()
    v10 = vr.notna();  v11 = e20.notna() & e50.notna()
    v12 = bb.notna();  v13 = stk.notna() & std.notna()

    c1  = _vcrit((h52 - cmp) / h52s <= 0.15, v1)
    c2  = _vcrit(cmp > e50, v2)
    c3  = _vcrit((cmp - l52) / (h52 - l52) >= 0.60, v3)
    c4  = _vcrit(rsi.between(40, 68), v4)
    c5  = _vcrit(roe > 12, v5);  c6 = _vcrit(pe.between(5, 30), v6)
    c7  = _vcrit(de < 1.0, v7);  c8 = _vcrit(cmp > e200, v8)
    c9  = _vcrit(macd > msig, v9); c10 = _vcrit(vr > 1.5, v10)
    c11 = _vcrit(e20 > e50, v11); c12 = _vcrit(bb > 0.5, v12)
    c13 = _vcrit(stk.between(20, 80) & (stk > std), v13)

    score = (_pts(c1, 2) + _pts(c2, 2) + _pts(c3) + _pts(c4) + _pts(c5) +
             _pts(c6) + _pts(c7) + _pts(c8) + _pts(c9) + _pts(c10) +
             _pts(c11) + _pts(c12) + _pts(c13))

    signal = np.where(score >= 5, "BUY", np.where(score >= 3, "WATCH", "AVOID"))

    out = d.copy()
    out["darvas_score"]  = score
    out["darvas_signal"] = signal
    out["near_52w_high"] = c1.map({True: "Y", False: "N"}).fillna("—")
    out["above_ema50"]   = c2.map({True: "Y", False: "N"}).fillna("—")
    out["range_str"]     = c3.map({True: "Y", False: "N"}).fillna("—")
    out["rsi_ok"]        = c4.map({True: "Y", False: "N"}).fillna("—")
    out["macd_bull"]     = c9.map({True: "Y", False: "N"}).fillna("—")
    out["above_ema200"]  = c8.map({True: "Y", False: "N"}).fillna("—")
    out["roe_ok"]        = c5.map({True: "Y", False: "N"}).fillna("—")
    out["pe_ok"]         = c6.map({True: "Y", False: "N"}).fillna("—")
    out["de_ok"]         = c7.map({True: "Y", False: "N"}).fillna("—")
    return out.sort_values(
        ["darvas_signal", "darvas_score"],
        key=lambda s: s.map({"BUY": 0, "WATCH": 1, "AVOID": 2}) if s.name == "darvas_signal" else -s,
    ).reset_index(drop=True)


def scan_piotroski(df: pd.DataFrame) -> pd.DataFrame:
    d = df.reset_index(drop=True)
    cmp  = _col(d, "cmp");   h52 = _col(d, "high_52w")
    e20  = _col(d, "ema_20"); e50 = _col(d, "ema_50")
    rsi  = _col(d, "rsi");   roe = _col(d, "roe");    opm = _col(d, "opm")
    pb   = _col(d, "pb");    de  = _col(d, "debt_to_equity")
    cr   = _col(d, "current_ratio"); rg = _col(d, "revenue_growth")
    atr  = _col(d, "atr_14"); stk = _col(d, "stoch_k"); std = _col(d, "stoch_d")

    v1 = roe.notna(); v2 = de.notna(); v3 = opm.notna(); v4 = rg.notna()
    v5 = cmp.notna() & e50.notna(); v6 = rsi.notna()
    v7 = cmp.notna() & h52.gt(0) & h52.notna()
    v8 = cr.notna(); v9 = pb.notna()
    v10 = e20.notna() & e50.notna(); v11 = atr.notna() & cmp.gt(0)

    c1  = _vcrit(roe > 0, v1);   c2 = _vcrit(de < 0.5, v2)
    c3  = _vcrit(opm > 15, v3);  c4 = _vcrit(rg > 0, v4)
    c5  = _vcrit(cmp > e50, v5); c6 = _vcrit(rsi.between(40, 70), v6)
    c7  = _vcrit((h52 - cmp) / h52.where(h52 > 0) <= 0.20, v7)
    c8  = _vcrit(cr > 1.5, v8);  c9 = _vcrit(pb < 3, v9)
    c10 = _vcrit(e20 > e50, v10); c11 = _vcrit((atr / cmp) * 100 < 3, v11)

    score = (_pts(c1) + _pts(c2) + _pts(c3) + _pts(c4) + _pts(c5) +
             _pts(c6) + _pts(c7) + _pts(c8) + _pts(c9) + _pts(c10) + _pts(c11))

    max_s = sum(v.astype(int) for v in [v1, v2, v3, v4, v5, v6, v7, v8, v9, v10, v11])
    buy_thr   = np.ceil(max_s * 0.65).astype(int)
    watch_thr = np.ceil(max_s * 0.40).astype(int)
    signal = np.where(score >= buy_thr, "BUY",
             np.where(score >= watch_thr, "WATCH", "AVOID"))

    out = d.copy()
    out["pio_score"]  = score
    out["pio_max"]    = max_s
    out["pio_signal"] = signal
    return out.sort_values(
        ["pio_signal", "pio_score"],
        key=lambda s: s.map({"BUY": 0, "WATCH": 1, "AVOID": 2}) if s.name == "pio_signal" else -s,
    ).reset_index(drop=True)


# ── index levels ─────────────────────────────────────────────────────────────

def fetch_index_data() -> dict:
    indices = {
        "NIFTY 50":   "^NSEI",
        "SENSEX":     "^BSESN",
        "NIFTY BANK": "^NSEBANK",
        "NIFTY IT":   "^CNXIT",
        "NIFTY MID":  "NIFTY_MID_SELECT.NS",
        "NIFTY SMALL":"^CNXSC",
    }
    result = {}
    try:
        raw = yf.download(" ".join(indices.values()), period="5d",
                          auto_adjust=True, progress=False)
        closes = (raw["Close"] if "Close" in raw.columns
                  else raw.xs("Close", axis=1, level=0))
        for name, sym in indices.items():
            try:
                col = closes[sym].dropna()
                if len(col) >= 2:
                    lv = float(col.iloc[-1]); pv = float(col.iloc[-2])
                    chg = lv - pv
                    result[name] = {
                        "level":   round(lv, 2),
                        "change":  round(chg, 2),
                        "chg_pct": round(chg / pv * 100, 2) if pv else 0,
                    }
            except Exception:
                pass
    except Exception as exc:
        print(f"  [WARN] Index fetch failed: {exc}")
    return result


# ── Excel writer ─────────────────────────────────────────────────────────────

def write_excel(
    darvas_df: pd.DataFrame,
    piotroski_df: pd.DataFrame,
    index_data: dict,
    as_of: datetime,
    xl_path: Path,
) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("pip install openpyxl")

    date_label = as_of.strftime("%d-%b-%y")   # e.g. 23-Jun-26

    # Load or create workbook
    if xl_path.exists():
        wb = load_workbook(xl_path)
    else:
        wb = Workbook()
        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # Helper styles
    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color="999999")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_t = Border(left=thin, right=thin, top=thick, bottom=thick)

    C = {
        "navy":      "1E3A5F",
        "dark_navy": "0D2137",
        "mid_navy":  "2B4C7E",
        "buy_bg":    "C6EFCE", "buy_fg":   "276221",
        "watch_bg":  "FFEB9C", "watch_fg": "9C6500",
        "avoid_bg":  "FFC7CE", "avoid_fg": "9C0006",
        "pos_fg":    "276221", "neg_fg":   "9C0006",
        "alt":       "F2F7FC", "white":    "FFFFFF",
        "hdr_fg":    "FFFFFF",
    }

    def fill(h): return PatternFill("solid", fgColor=h)
    def fnt(bold=False, color="000000", size=10, italic=False, name="Arial"):
        return Font(name=name, bold=bold, color=color, size=size, italic=italic)
    def mono(color="000000", size=10): return Font(name="Courier New", size=size, color=color)
    def aln(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def sig_style(sig):
        m = {"BUY":   (C["buy_bg"],   C["buy_fg"]),
             "WATCH": (C["watch_bg"], C["watch_fg"]),
             "AVOID": (C["avoid_bg"], C["avoid_fg"])}
        bg, fg = m.get(sig, (C["white"], "000000"))
        return fill(bg), fnt(bold=True, color=fg)

    def write_title(ws, row, text, n_cols=15):
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        c = ws.cell(row, 1, text)
        c.font = fnt(bold=True, size=13, color=C["hdr_fg"])
        c.fill = fill(C["dark_navy"])
        c.alignment = aln("center")
        ws.row_dimensions[row].height = 24
        return row + 1

    def write_section_header(ws, row, text, n_cols=15):
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        c = ws.cell(row, 1, text)
        c.font = fnt(bold=True, size=11, color=C["hdr_fg"])
        c.fill = fill(C["mid_navy"])
        c.alignment = aln("left")
        ws.row_dimensions[row].height = 20
        return row + 1

    def write_col_headers(ws, row, headers, widths):
        for col_i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row, col_i, h)
            c.font = fnt(bold=True, color=C["hdr_fg"], size=9)
            c.fill = fill(C["navy"])
            c.alignment = aln("center")
            c.border = bdr_t
            ws.column_dimensions[get_column_letter(col_i)].width = w
        ws.row_dimensions[row].height = 18
        return row + 1

    # ── SHEET 1: Summary ─────────────────────────────────────────────────────
    sname = "Summary"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname, 0)
    ws.sheet_properties.tabColor = "1E3A5F"

    row = write_title(ws, 1, f"NIFTY 500 Scan  —  {as_of.strftime('%A, %d %B %Y')}  |  {as_of.strftime('%H:%M IST')}")
    row += 1

    # Index levels
    row = write_section_header(ws, row, "BENCHMARK LEVELS")
    row = write_col_headers(ws, row, ["Index", "Level", "Change", "Change %"], [20, 14, 12, 12])
    for i, (name, d) in enumerate(index_data.items()):
        fg = C["pos_fg"] if d["change"] >= 0 else C["neg_fg"]
        bg = C["alt"] if i % 2 else C["white"]
        vals = [name, d["level"], d["change"], d["chg_pct"] / 100]
        fmts = [None, "#,##0.00", "+#,##0.00;-#,##0.00", "+0.00%;-0.00%"]
        for col_i, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row, col_i, v)
            c.font = mono(color=(fg if col_i > 1 else "000000"))
            c.fill = fill(bg)
            c.alignment = aln("right" if col_i > 1 else "left")
            c.border = bdr
            if fmt: c.number_format = fmt
        ws.row_dimensions[row].height = 16
        row += 1
    row += 1

    # Scan stats
    d_buys  = (darvas_df["darvas_signal"] == "BUY").sum()
    d_watch = (darvas_df["darvas_signal"] == "WATCH").sum()
    d_avoid = (darvas_df["darvas_signal"] == "AVOID").sum()
    p_buys  = (piotroski_df["pio_signal"] == "BUY").sum()
    p_watch = (piotroski_df["pio_signal"] == "WATCH").sum()

    row = write_section_header(ws, row, "SCAN SUMMARY")
    stats = [
        ("Total stocks scanned",  len(darvas_df)),
        ("Darvas/Buffett  BUY",   d_buys),
        ("Darvas/Buffett  WATCH", d_watch),
        ("Darvas/Buffett  AVOID", d_avoid),
        ("Piotroski  BUY",        p_buys),
        ("Piotroski  WATCH",      p_watch),
        ("Run date/time",         as_of.strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (label, val) in enumerate(stats):
        bg = C["alt"] if i % 2 else C["white"]
        for col_i, v in enumerate([label, str(val)], 1):
            c = ws.cell(row, col_i, v)
            c.font = fnt(bold=(col_i == 1))
            c.fill = fill(bg)
            c.alignment = aln("left")
            c.border = bdr
        ws.row_dimensions[row].height = 16
        row += 1

    ws.freeze_panes = "A3"

    # ── SHEET 2: Darvas ───────────────────────────────────────────────────────
    sname = "Darvas"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname, 1)
    ws.sheet_properties.tabColor = "276221"

    # Interpreter columns appended after criteria flags
    D_HDRS = ["Ticker","Name","Sector","CMP (₹)","52W Hi","52W Lo",
              "EMA 20","EMA 50","EMA 200","RSI","Vol Ratio",
              "Ret 1D%","Ret 1W%","Ret 1M%","Ret 3M%","Ret 6M%","Ret 1Y%",
              "PE","ROE%","D/E","Score","Signal",
              "H","E50","RSI","MACD","E200","ROE","PE","DE",
              "DV State","DV Action","Entry (₹)","Stop (₹)","Risk %",
              "Target 1 (₹)","Rwd 1 %","Target 2 (₹)","Rwd 2 %","Target 3 (₹)","Rwd 3 %"]
    D_WIDS = [12,28,18,10,10,10,9,9,9,7,9,
              8,8,8,8,8,8,
              7,7,7,7,10,
              4,4,4,4,4,4,4,4,
              16,36,11,11,8,
              12,8,12,8,12,8]

    row = write_title(ws, 1, f"Darvas / Buffett Scan  —  {date_label}  |  BUY:{d_buys}  WATCH:{d_watch}  AVOID:{d_avoid}", len(D_HDRS))
    row = write_col_headers(ws, row, D_HDRS, D_WIDS)

    # Map state → fill/font colours for interpreter column
    DV_STATE_STYLE = {
        "BREAKOUT_FRESH": ("C6EFCE","276221"),
        "BREAKOUT_OLD":   ("DDFFDD","276221"),
        "NEAR_TOP":       ("FFEB9C","9C6500"),
        "IN_BOX":         ("DDEEFF","003366"),
        "FORMING":        ("EEE0FF","4B0082"),
        "BREAKDOWN":      ("FFC7CE","9C0006"),
        "NO_BOX":         ("F0F0F0","888888"),
    }

    ret_key_map = {12:"ret_1d",13:"ret_1w",14:"ret_1m",15:"ret_3m",16:"ret_6m",17:"ret_1y"}

    for i, (_, r) in enumerate(darvas_df.iterrows()):
        bg  = C["alt"] if i % 2 else C["white"]
        sig = str(r.get("darvas_signal", ""))
        is_actionable = sig in ("BUY", "WATCH")

        def _v(k):
            val = r.get(k)
            return val if val is not None and pd.notna(val) else None

        dv_state  = str(r.get("dv_state", "")) if is_actionable else ""
        dv_action = str(r.get("dv_action","")) if is_actionable else ""

        row_vals = [
            r["ticker"], r.get("name",""), r.get("sector",""),
            _v("cmp"), _v("high_52w"), _v("low_52w"),
            _v("ema_20"), _v("ema_50"), _v("ema_200"), _v("rsi"), _v("volume_ratio"),
            (_v("ret_1d") or 0)/100 if _v("ret_1d") is not None else None,
            (_v("ret_1w") or 0)/100 if _v("ret_1w") is not None else None,
            (_v("ret_1m") or 0)/100 if _v("ret_1m") is not None else None,
            (_v("ret_3m") or 0)/100 if _v("ret_3m") is not None else None,
            (_v("ret_6m") or 0)/100 if _v("ret_6m") is not None else None,
            (_v("ret_1y") or 0)/100 if _v("ret_1y") is not None else None,
            _v("pe"), (_v("roe") or 0)/100 if _v("roe") is not None else None,
            _v("debt_to_equity"),
            r.get("darvas_score"), sig,
            r.get("near_52w_high","—"), r.get("above_ema50","—"),
            r.get("rsi_ok","—"),        r.get("macd_bull","—"),
            r.get("above_ema200","—"),  r.get("roe_ok","—"),
            r.get("pe_ok","—"),         r.get("de_ok","—"),
            # ── Interpreter columns ──────────────────────────────────────────
            dv_state,
            dv_action,
            _v("dv_entry")   if is_actionable else None,
            _v("dv_stop")    if is_actionable else None,
            _v("dv_risk_pct")if is_actionable else None,
            _v("dv_target1") if is_actionable else None,
            _v("dv_rwd1_pct")if is_actionable else None,
            _v("dv_target2") if is_actionable else None,
            _v("dv_rwd2_pct")if is_actionable else None,
            _v("dv_target3") if is_actionable else None,
            _v("dv_rwd3_pct")if is_actionable else None,
        ]

        price_cols  = {4,5,6,7,8,9,10}
        ret_cols    = {12,13,14,15,16,17,18}
        ratio_cols  = {19,20,21}
        crit_cols   = set(range(23, 31))
        sig_col     = 22
        dv_state_col = 31
        dv_action_col= 32
        dv_price_cols= {33,34,36,38,40}   # entry, stop, T1, T2, T3
        dv_pct_cols  = {35,37,39,41}       # risk%, rwd1%, rwd2%, rwd3%

        fmts = {
            **{c: "#,##0.00"       for c in price_cols},
            **{c: "+0.00%;-0.00%"  for c in ret_cols},
            19: "0.0x", 20: "0.0%", 21: "0.0x",
            **{c: "#,##0.00"       for c in dv_price_cols},
            **{c: "0.00%"          for c in dv_pct_cols},
        }

        for col_i, v in enumerate(row_vals, 1):
            c = ws.cell(row, col_i, v)
            c.border = bdr

            if col_i in ret_cols and v is not None:
                orig = _v(ret_key_map.get(col_i, ""))
                fg = C["pos_fg"] if (orig or 0) >= 0 else C["neg_fg"]
                c.font = mono(color=fg); c.fill = fill(bg); c.alignment = aln("right")

            elif col_i == sig_col:
                sf, sft = sig_style(sig)
                c.fill = sf; c.font = sft; c.alignment = aln("center")

            elif col_i in crit_cols:
                is_y = str(v) == "Y"
                c.font = fnt(bold=True, color=(C["pos_fg"] if is_y else C["neg_fg"]))
                c.fill = fill(bg); c.alignment = aln("center")

            elif col_i == dv_state_col and dv_state:
                bg_s, fg_s = DV_STATE_STYLE.get(dv_state, ("FFFFFF","000000"))
                c.font = fnt(bold=True, color=fg_s)
                c.fill = fill(bg_s); c.alignment = aln("center")

            elif col_i == dv_action_col:
                c.font = fnt(italic=True, size=9); c.fill = fill(bg); c.alignment = aln("left")

            elif col_i in dv_price_cols and is_actionable:
                colours = {33: "0070C0", 34: "9C0006",
                           36: "276221", 38: "276221", 40: "276221"}
                c.font = mono(color=colours.get(col_i,"000000"), size=10)
                c.fill = fill(bg); c.alignment = aln("right")

            elif col_i in dv_pct_cols and is_actionable:
                colours = {35: "9C0006", 37: "276221", 39: "276221", 41: "276221"}
                # convert plain % float to decimal for % number format
                if v is not None:
                    c.value = v / 100
                c.font = mono(color=colours.get(col_i,"000000"))
                c.fill = fill(bg); c.alignment = aln("right")

            elif col_i == 1:
                c.font = fnt(bold=True); c.fill = fill(bg); c.alignment = aln("left")

            elif col_i in price_cols | ratio_cols:
                c.font = mono(); c.fill = fill(bg); c.alignment = aln("right")

            else:
                c.font = fnt(); c.fill = fill(bg)
                c.alignment = aln("right" if col_i > 3 else "left")

            if col_i in fmts and col_i not in dv_pct_cols:
                c.number_format = fmts[col_i]
            elif col_i in dv_pct_cols:
                c.number_format = "+0.00%;-0.00%"

        ws.row_dimensions[row].height = 15
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(D_HDRS))}{row - 1}"

    # ── SHEET 3: Piotroski ────────────────────────────────────────────────────
    sname = "Piotroski"
    if sname in wb.sheetnames:
        del wb[sname]
    ws = wb.create_sheet(sname, 2)
    ws.sheet_properties.tabColor = "9C6500"

    P_HDRS = ["Ticker","Name","Sector","CMP (₹)","RSI",
              "ROE%","OPM%","D/E","P/B","Current R","Rev Grw%",
              "EMA 50","Score","Max","Signal"]
    P_WIDS = [12,28,18,10,7,8,8,8,7,10,10,10,7,5,10]

    row = write_title(ws, 1, f"Piotroski Scan  —  {date_label}  |  BUY:{p_buys}  WATCH:{p_watch}", len(P_HDRS))
    row = write_col_headers(ws, row, P_HDRS, P_WIDS)

    for i, (_, r) in enumerate(piotroski_df.iterrows()):
        bg  = C["alt"] if i % 2 else C["white"]
        sig = str(r.get("pio_signal", ""))

        def _v(k): return r.get(k) if pd.notna(r.get(k, None)) else None

        row_vals = [
            r["ticker"], r.get("name",""), r.get("sector",""),
            _v("cmp"), _v("rsi"),
            (_v("roe") or 0)/100  if _v("roe")  is not None else None,
            (_v("opm") or 0)/100  if _v("opm")  is not None else None,
            _v("debt_to_equity"), _v("pb"), _v("current_ratio"),
            (_v("revenue_growth") or 0)/100 if _v("revenue_growth") is not None else None,
            _v("ema_50"),
            r.get("pio_score"), r.get("pio_max"), sig,
        ]
        num_cols = {4,5,6,7,8,9,10,11,12}
        sig_col  = 15
        fmts = {4:"#,##0.00", 5:"0.0", 6:"0.0%", 7:"0.0%",
                8:"0.0x", 9:"0.0x", 10:"0.0", 11:"0.0%", 12:"#,##0.00"}

        for col_i, v in enumerate(row_vals, 1):
            c = ws.cell(row, col_i, v)
            c.border = bdr
            if col_i == sig_col:
                sf, sft = sig_style(sig)
                c.fill = sf; c.font = sft; c.alignment = aln("center")
            elif col_i == 1:
                c.font = fnt(bold=True); c.fill = fill(bg); c.alignment = aln("left")
            elif col_i in num_cols:
                c.font = mono(); c.fill = fill(bg); c.alignment = aln("right")
                if col_i in fmts: c.number_format = fmts[col_i]
            else:
                c.font = fnt(); c.fill = fill(bg)
                c.alignment = aln("right" if col_i > 3 else "left")
        ws.row_dimensions[row].height = 15
        row += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(P_HDRS))}{row - 1}"

    # ── SHEET 4: History (append one row per run) ─────────────────────────────
    sname = "History"
    if sname not in wb.sheetnames:
        ws_h = wb.create_sheet(sname, 3)
        ws_h.sheet_properties.tabColor = "555555"
        hist_hdrs = ["Date","Time","Stocks Scanned",
                     "Darvas BUY","Darvas WATCH","Darvas AVOID",
                     "Piotroski BUY","Piotroski WATCH",
                     "NIFTY 50","NIFTY Chg%","SENSEX","SENSEX Chg%"]
        for col_i, h in enumerate(hist_hdrs, 1):
            c = ws_h.cell(1, col_i, h)
            c.font = fnt(bold=True, color=C["hdr_fg"])
            c.fill = fill(C["navy"])
            c.alignment = aln("center")
            c.border = bdr
            ws_h.column_dimensions[get_column_letter(col_i)].width = 14
        ws_h.row_dimensions[1].height = 18
    else:
        ws_h = wb[sname]

    next_row = ws_h.max_row + 1
    hist_vals = [
        as_of.strftime("%Y-%m-%d"), as_of.strftime("%H:%M"),
        len(darvas_df), d_buys, d_watch, d_avoid, p_buys, p_watch,
        index_data.get("NIFTY 50",  {}).get("level",  ""),
        index_data.get("NIFTY 50",  {}).get("chg_pct",""),
        index_data.get("SENSEX",    {}).get("level",  ""),
        index_data.get("SENSEX",    {}).get("chg_pct",""),
    ]
    bg = C["alt"] if next_row % 2 else C["white"]
    for col_i, v in enumerate(hist_vals, 1):
        c = ws_h.cell(next_row, col_i, v)
        c.font = fnt()
        c.fill = fill(bg)
        c.alignment = aln("right" if col_i > 2 else "center")
        c.border = bdr
    ws_h.row_dimensions[next_row].height = 15

    wb.save(xl_path)
    print(f"\n  ✓  Saved → {xl_path}")
    print(f"     Sheets: Summary | Darvas | Piotroski | History")
    print(f"     Darvas  — BUY: {d_buys}  WATCH: {d_watch}  AVOID: {d_avoid}")
    print(f"     Piotroski — BUY: {p_buys}  WATCH: {p_watch}")
    return wb, xl_path


def write_gems_sheet(
    gems_darvas: pd.DataFrame,
    gems_piotroski: pd.DataFrame,
    as_of: datetime,
    xl_path: Path,
) -> None:
    """
    Append / replace a 'Hidden Gems' sheet in the existing workbook.
    Shows stocks from the full NSE universe (outside NIFTY 500) that
    pass Darvas BUY/WATCH or Piotroski BUY/WATCH criteria.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    date_label = as_of.strftime("%d-%b-%y")
    wb = load_workbook(xl_path)

    thin  = Side(style="thin",   color="CCCCCC")
    thick = Side(style="medium", color="999999")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    bdr_t = Border(left=thin, right=thin, top=thick, bottom=thick)

    C = {
        "navy":      "1E3A5F", "dark_navy": "0D2137", "mid_navy": "2B4C7E",
        "gem_bg":    "E8D5FF", "gem_fg":   "4B0082",   # purple for gems
        "buy_bg":    "C6EFCE", "buy_fg":   "276221",
        "watch_bg":  "FFEB9C", "watch_fg": "9C6500",
        "avoid_bg":  "FFC7CE", "avoid_fg": "9C0006",
        "pos_fg":    "276221", "neg_fg":   "9C0006",
        "alt":       "F2F7FC", "white":    "FFFFFF",
        "hdr_fg":    "FFFFFF",
    }

    def fill(h): return PatternFill("solid", fgColor=h)
    def fnt(bold=False, color="000000", size=10, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    def mono(color="000000", size=10):
        return Font(name="Courier New", size=size, color=color)
    def aln(h="left", v="center"):
        return Alignment(horizontal=h, vertical=v)

    def sig_style(sig):
        m = {"BUY":   (C["buy_bg"],   C["buy_fg"]),
             "WATCH": (C["watch_bg"], C["watch_fg"])}
        bg, fg = m.get(sig, (C["avoid_bg"], C["avoid_fg"]))
        return fill(bg), fnt(bold=True, color=fg)

    def write_title(ws, row, text, n_cols=16):
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        c = ws.cell(row, 1, text)
        c.font = fnt(bold=True, size=13, color=C["hdr_fg"])
        c.fill = fill("4B0082")  # purple title for gems
        c.alignment = aln("center")
        ws.row_dimensions[row].height = 24
        return row + 1

    def write_section(ws, row, text, n_cols=16):
        ws.merge_cells(f"A{row}:{get_column_letter(n_cols)}{row}")
        c = ws.cell(row, 1, text)
        c.font = fnt(bold=True, size=11, color=C["hdr_fg"])
        c.fill = fill("6A0DAD")
        c.alignment = aln("left")
        ws.row_dimensions[row].height = 20
        return row + 1

    def write_col_headers(ws, row, headers, widths):
        for col_i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row, col_i, h)
            c.font = fnt(bold=True, color=C["hdr_fg"], size=9)
            c.fill = fill(C["navy"])
            c.alignment = aln("center")
            c.border = bdr_t
            ws.column_dimensions[get_column_letter(col_i)].width = w
        ws.row_dimensions[row].height = 18
        return row + 1

    # Create / replace sheet
    sname = "Hidden Gems"
    if sname in wb.sheetnames:
        del wb[sname]
    # Insert after Piotroski (index 3)
    sheet_idx = min(3, len(wb.sheetnames))
    ws = wb.create_sheet(sname, sheet_idx)
    ws.sheet_properties.tabColor = "7B2FBE"

    d_signals = gems_darvas["darvas_signal"] if not gems_darvas.empty else pd.Series(dtype=str)
    p_signals = gems_piotroski["pio_signal"] if not gems_piotroski.empty else pd.Series(dtype=str)

    d_buy   = (d_signals == "BUY").sum()
    d_watch = (d_signals == "WATCH").sum()
    p_buy   = (p_signals == "BUY").sum()
    p_watch = (p_signals == "WATCH").sum()

    row = write_title(
        ws, 1,
        f"Hidden Gems (outside NIFTY 500)  —  {date_label}  |  "
        f"Darvas BUY:{d_buy} WATCH:{d_watch}  |  Piotroski BUY:{p_buy} WATCH:{p_watch}",
    )

    # ── explanation row ───────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:P{row}")
    c = ws.cell(row, 1,
        "Stocks from the full NSE universe (~2,368) that are NOT in NIFTY 500 "
        "but pass Darvas/Buffett or Piotroski criteria. Sorted: BUY → WATCH, then by score.")
    c.font = fnt(italic=True, size=9, color="555555")
    c.alignment = aln("left")
    ws.row_dimensions[row].height = 14
    row += 2

    # ── Darvas gems ───────────────────────────────────────────────────────────
    G_HDRS = ["Ticker","Name","Sector","CMP (₹)","52W Hi","EMA 50","EMA 200",
              "RSI","Vol Ratio","Ret 1D%","Ret 1W%","Ret 1M%","Ret 1Y%",
              "PE","ROE%","Score","Signal",
              "DV State","DV Action","Entry (₹)","Stop (₹)","Risk %",
              "Target 1 (₹)","Rwd 1 %","Target 2 (₹)","Rwd 2 %","Target 3 (₹)","Rwd 3 %"]
    G_WIDS = [12,28,18,10,10,10,10,7,9,8,8,8,8,7,7,7,10,
              16,36,11,11,8,12,8,12,8,12,8]

    GDV_STATE_STYLE = {
        "BREAKOUT_FRESH": ("C6EFCE","276221"),
        "BREAKOUT_OLD":   ("DDFFDD","276221"),
        "NEAR_TOP":       ("FFEB9C","9C6500"),
        "IN_BOX":         ("DDEEFF","003366"),
        "FORMING":        ("EEE0FF","4B0082"),
        "BREAKDOWN":      ("FFC7CE","9C0006"),
        "NO_BOX":         ("F0F0F0","888888"),
    }

    row = write_section(ws, row, f"DARVAS / BUFFETT HIDDEN GEMS  (BUY:{d_buy}  WATCH:{d_watch})", len(G_HDRS))
    row = write_col_headers(ws, row, G_HDRS, G_WIDS)

    gems_d_filtered = gems_darvas[gems_darvas["darvas_signal"].isin(["BUY", "WATCH"])] \
        if not gems_darvas.empty else gems_darvas

    for i, (_, r) in enumerate(gems_d_filtered.iterrows()):
        bg  = C["alt"] if i % 2 else C["white"]
        sig = str(r.get("darvas_signal", ""))
        is_actionable = sig in ("BUY", "WATCH")

        def _v(k):
            val = r.get(k)
            return val if val is not None and pd.notna(val) else None

        ret_keys  = ["ret_1d", "ret_1w", "ret_1m", "ret_1y"]
        dv_state  = str(r.get("dv_state", ""))  if is_actionable else ""
        dv_action = str(r.get("dv_action",""))  if is_actionable else ""

        row_vals = [
            r["ticker"], r.get("name",""), r.get("sector",""),
            _v("cmp"), _v("high_52w"), _v("ema_50"), _v("ema_200"),
            _v("rsi"), _v("volume_ratio"),
            (_v("ret_1d") or 0)/100 if _v("ret_1d") is not None else None,
            (_v("ret_1w") or 0)/100 if _v("ret_1w") is not None else None,
            (_v("ret_1m") or 0)/100 if _v("ret_1m") is not None else None,
            (_v("ret_1y") or 0)/100 if _v("ret_1y") is not None else None,
            _v("pe"),
            (_v("roe") or 0)/100 if _v("roe") is not None else None,
            r.get("darvas_score"), sig,
            # ── interpreter columns ──────────────────────────────────────
            dv_state,
            dv_action,
            _v("dv_entry")    if is_actionable else None,
            _v("dv_stop")     if is_actionable else None,
            _v("dv_risk_pct") if is_actionable else None,
            _v("dv_target1")  if is_actionable else None,
            _v("dv_rwd1_pct") if is_actionable else None,
            _v("dv_target2")  if is_actionable else None,
            _v("dv_rwd2_pct") if is_actionable else None,
            _v("dv_target3")  if is_actionable else None,
            _v("dv_rwd3_pct") if is_actionable else None,
        ]

        ret_cols     = {10, 11, 12, 13}
        num_cols     = {4, 5, 6, 7, 8, 9}
        sig_col      = 17
        dv_state_col = 18
        dv_action_col= 19
        dv_price_cols= {20, 21, 23, 25, 27}
        dv_pct_cols  = {22, 24, 26, 28}

        fmts = {4:"#,##0.00",5:"#,##0.00",6:"#,##0.00",7:"#,##0.00",
                8:"0.0",9:"0.0",
                14:"0.0x",15:"0.0%",
                **{c:"+0.00%;-0.00%" for c in ret_cols},
                **{c:"#,##0.00"      for c in dv_price_cols},
                **{c:"+0.00%;-0.00%" for c in dv_pct_cols}}

        for col_i, v in enumerate(row_vals, 1):
            c = ws.cell(row, col_i, v)
            c.border = bdr
            if col_i == sig_col:
                sf, sft = sig_style(sig)
                c.fill = sf; c.font = sft; c.alignment = aln("center")
            elif col_i == 1:
                c.font = fnt(bold=True, color="4B0082")
                c.fill = fill(bg); c.alignment = aln("left")
            elif col_i in ret_cols and v is not None:
                orig_val = _v(ret_keys[col_i - 10]) or 0
                fg = C["pos_fg"] if orig_val >= 0 else C["neg_fg"]
                c.font = mono(color=fg); c.fill = fill(bg); c.alignment = aln("right")
                c.number_format = "+0.00%;-0.00%"
            elif col_i == dv_state_col and dv_state:
                bg_s, fg_s = GDV_STATE_STYLE.get(dv_state, ("FFFFFF","000000"))
                c.font = fnt(bold=True, color=fg_s)
                c.fill = fill(bg_s); c.alignment = aln("center")
            elif col_i == dv_action_col:
                c.font = fnt(italic=True, size=9); c.fill = fill(bg); c.alignment = aln("left")
            elif col_i in dv_price_cols and is_actionable:
                clr = {20:"0070C0",21:"9C0006",23:"276221",25:"276221",27:"276221"}
                c.font = mono(color=clr.get(col_i,"000000"), size=10)
                c.fill = fill(bg); c.alignment = aln("right")
                c.number_format = "#,##0.00"
            elif col_i in dv_pct_cols and is_actionable:
                clr = {22:"9C0006",24:"276221",26:"276221",28:"276221"}
                if v is not None:
                    c.value = v / 100
                c.font = mono(color=clr.get(col_i,"000000"))
                c.fill = fill(bg); c.alignment = aln("right")
                c.number_format = "+0.00%;-0.00%"
            elif col_i in num_cols:
                c.font = mono(); c.fill = fill(bg); c.alignment = aln("right")
                if col_i in fmts: c.number_format = fmts[col_i]
            else:
                c.font = fnt(); c.fill = fill(bg)
                c.alignment = aln("right" if col_i > 3 else "left")
                if col_i in fmts: c.number_format = fmts[col_i]
        ws.row_dimensions[row].height = 15
        row += 1

    row += 1

    # ── Piotroski gems ────────────────────────────────────────────────────────
    P_HDRS = ["Ticker","Name","Sector","CMP (₹)","RSI",
              "ROE%","OPM%","D/E","P/B","Current R","Rev Grw%",
              "EMA 50","Score","Max","Signal"]
    P_WIDS = [12,28,18,10,7,8,8,8,7,10,10,10,7,5,10]

    row = write_section(ws, row, f"PIOTROSKI HIDDEN GEMS  (BUY:{p_buy}  WATCH:{p_watch})", len(P_HDRS))
    row = write_col_headers(ws, row, P_HDRS, P_WIDS)

    gems_p_filtered = gems_piotroski[gems_piotroski["pio_signal"].isin(["BUY", "WATCH"])] \
        if not gems_piotroski.empty else gems_piotroski

    for i, (_, r) in enumerate(gems_p_filtered.iterrows()):
        bg  = C["alt"] if i % 2 else C["white"]
        sig = str(r.get("pio_signal", ""))

        def _v(k):
            val = r.get(k)
            return val if val is not None and pd.notna(val) else None

        row_vals = [
            r["ticker"], r.get("name",""), r.get("sector",""),
            _v("cmp"), _v("rsi"),
            (_v("roe") or 0)/100  if _v("roe")  is not None else None,
            (_v("opm") or 0)/100  if _v("opm")  is not None else None,
            _v("debt_to_equity"), _v("pb"), _v("current_ratio"),
            (_v("revenue_growth") or 0)/100 if _v("revenue_growth") is not None else None,
            _v("ema_50"), r.get("pio_score"), r.get("pio_max"), sig,
        ]
        num_cols = {4,5,6,7,8,9,10,11,12}
        sig_col  = 15
        fmts = {4:"#,##0.00",5:"0.0",6:"0.0%",7:"0.0%",
                8:"0.0x",9:"0.0x",10:"0.0",11:"0.0%",12:"#,##0.00"}

        for col_i, v in enumerate(row_vals, 1):
            c = ws.cell(row, col_i, v)
            c.border = bdr
            if col_i == sig_col:
                sf, sft = sig_style(sig)
                c.fill = sf; c.font = sft; c.alignment = aln("center")
            elif col_i == 1:
                c.font = fnt(bold=True, color="4B0082")
                c.fill = fill(bg); c.alignment = aln("left")
            elif col_i in num_cols:
                c.font = mono(); c.fill = fill(bg); c.alignment = aln("right")
                if col_i in fmts: c.number_format = fmts[col_i]
            else:
                c.font = fnt(); c.fill = fill(bg)
                c.alignment = aln("right" if col_i > 3 else "left")
        ws.row_dimensions[row].height = 15
        row += 1

    ws.freeze_panes = "A3"

    wb.save(xl_path)
    print(f"  ✓  Hidden Gems sheet written  "
          f"(Darvas: {len(gems_d_filtered)}  Piotroski: {len(gems_p_filtered)})")


# ── hidden gems subroutine ────────────────────────────────────────────────────

def load_full_nse_universe(data_dir: Path) -> list[str]:
    """
    Load all NSE equities from the local nse_equity_list.csv (~2,368 stocks).
    Falls back to an empty list if the file is not present.
    """
    csv_path = data_dir / "nse_equity_list.csv"
    if not csv_path.exists():
        print("  [WARN] data/nse_equity_list.csv not found — run the app once to download it.")
        return []
    try:
        df = pd.read_csv(csv_path)
        col = next((c for c in df.columns if "symbol" in c.upper()), df.columns[0])
        syms = df[col].dropna().str.strip().str.upper().tolist()
        # Keep only plain equity symbols (no spaces, no slashes)
        syms = [s for s in syms if s.isalnum() or "-" in s or "&" in s]
        print(f"  Loaded {len(syms)} symbols from nse_equity_list.csv")
        return syms
    except Exception as exc:
        print(f"  [WARN] Could not read nse_equity_list.csv: {exc}")
        return []


def run_hidden_gems(
    nifty500_set: set[str],
    data_dir: Path,
    fast: bool,
    batch_size: int,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Scan the full NSE universe minus NIFTY 500 for hidden BUY/WATCH opportunities.

    Steps:
      1. Load all ~2,368 NSE symbols from nse_equity_list.csv
      2. Exclude symbols already in NIFTY 500
      3. Batch-download OHLCV via yf.download()
      4. Optionally fetch fundamentals
      5. Run both scanners
      6. Return (darvas_df, piotroski_df, hist_map) — hist_map re-used by interpreter
    """
    print("\n" + "─" * 60)
    print("  HIDDEN GEMS SCAN  (full NSE universe minus NIFTY 500)")
    print("─" * 60)

    all_nse = load_full_nse_universe(data_dir)
    if not all_nse:
        empty = pd.DataFrame()
        return empty, empty, {}

    gems_universe = [s for s in all_nse if s not in nifty500_set]
    print(f"  Universe: {len(all_nse)} total  —  {len(nifty500_set)} NIFTY 500"
          f"  =  {len(gems_universe)} stocks to scan")

    print(f"\n  Downloading OHLCV for {len(gems_universe)} stocks …")
    t0 = time.time()
    hist_map = batch_download(gems_universe, batch_size=batch_size)
    print(f"  Downloaded {len(hist_map)}/{len(gems_universe)} in {time.time()-t0:.0f}s")

    infos: dict[str, dict] = {s: {} for s in gems_universe}
    if not fast and hist_map:
        print(f"\n  Fetching fundamentals for {len(hist_map)} gems …")
        print("  (Use --fast to skip  —  this adds ~10–30 min)")
        infos = fetch_fundamentals_bulk(list(hist_map.keys()))

    records = []
    for sym, hist in hist_map.items():
        try:
            records.append(build_record(sym, hist, infos.get(sym, {})))
        except Exception as exc:
            print(f"  [WARN] {sym}: {exc}")

    if not records:
        empty = pd.DataFrame()
        return empty, empty, {}

    df = pd.DataFrame(records)
    print(f"\n  Running scanners on {len(df)} gem candidates …")
    darvas_df    = scan_darvas(df)
    piotroski_df = scan_piotroski(df)

    d_buy   = (darvas_df["darvas_signal"] == "BUY").sum()
    d_watch = (darvas_df["darvas_signal"] == "WATCH").sum()
    p_buy   = (piotroski_df["pio_signal"] == "BUY").sum()
    p_watch = (piotroski_df["pio_signal"] == "WATCH").sum()

    print(f"\n  Darvas  — BUY: {d_buy}  WATCH: {d_watch}")
    print(f"  Piotroski — BUY: {p_buy}  WATCH: {p_watch}")

    # Quick preview of top gems
    top = darvas_df[darvas_df["darvas_signal"] == "BUY"].head(10)
    if not top.empty:
        print(f"\n  Top Darvas BUYs outside NIFTY 500:")
        for _, r in top.iterrows():
            print(f"    {r['ticker']:<14} score {r['darvas_score']:>2}/13  "
                  f"CMP {r['cmp']:>8.2f}  RSI {r.get('rsi',0):>5.1f}  "
                  f"{r.get('sector','')[:20]}")

    return darvas_df, piotroski_df, hist_map


# ── terminal summary ──────────────────────────────────────────────────────────

def print_summary(darvas_df: pd.DataFrame, piotroski_df: pd.DataFrame,
                  index_data: dict, as_of: datetime) -> None:
    W = 70
    print("\n" + "═" * W)
    print(f"  NIFTY 500 SCAN  —  {as_of.strftime('%d %b %Y  %H:%M')}")
    print("═" * W)

    if index_data:
        print()
        for name, d in index_data.items():
            arrow = "▲" if d["change"] >= 0 else "▼"
            print(f"  {name:<14}  {d['level']:>10,.2f}  "
                  f"{arrow} {d['change']:+,.2f}  ({d['chg_pct']:+.2f}%)")

    buys = darvas_df[darvas_df["darvas_signal"] == "BUY"]
    print(f"\n  Darvas/Buffett  BUY: {len(buys)}  "
          f"WATCH: {(darvas_df['darvas_signal']=='WATCH').sum()}  "
          f"AVOID: {(darvas_df['darvas_signal']=='AVOID').sum()}")

    if not buys.empty:
        print(f"\n  {'TICKER':<14}{'CMP':>9}{'RSI':>6}  "
              f"{'1D%':>7}{'1M%':>7}  SCORE  SECTOR")
        print("  " + "─" * (W - 2))
        for _, r in buys.head(20).iterrows():
            ret1d = f"{r['ret_1d']:+.1f}%" if pd.notna(r.get("ret_1d")) else "  —  "
            ret1m = f"{r['ret_1m']:+.1f}%" if pd.notna(r.get("ret_1m")) else "  —  "
            print(f"  {r['ticker']:<14}{r['cmp']:>9.2f}"
                  f"{r['rsi'] or 0:>6.1f}  "
                  f"{ret1d:>7}{ret1m:>7}  "
                  f"{r['darvas_score']:>3}/13  {r.get('sector','')[:18]}")

    p_buys = piotroski_df[piotroski_df["pio_signal"] == "BUY"]
    print(f"\n  Piotroski  BUY: {len(p_buys)}  "
          f"WATCH: {(piotroski_df['pio_signal']=='WATCH').sum()}")
    if not p_buys.empty:
        print(f"  Top: " + ", ".join(p_buys["ticker"].head(15).tolist()))

    print("\n" + "═" * W + "\n")


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="NIFTY 500 scan + Hidden Gems — batched yfinance + Excel output",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--fast", action="store_true",
                        help="Skip yfinance .info call (no PE/ROE/D-E, ~3x faster)")
    parser.add_argument("--batch-size", type=int, default=100, metavar="N",
                        help="Tickers per yf.download() call (default: 100)")
    parser.add_argument("--top", type=int, default=0, metavar="N",
                        help="Scan only first N NIFTY 500 symbols (for testing)")
    parser.add_argument("--output", type=str, default="", metavar="PATH",
                        help="Override Excel output path")
    parser.add_argument("--no-excel", action="store_true",
                        help="Terminal output only, no Excel file")
    parser.add_argument("--no-gems", action="store_true",
                        help="Skip the Hidden Gems scan (NIFTY 500 only)")
    parser.add_argument("--gems-only", action="store_true",
                        help="Run only the Hidden Gems scan, skip NIFTY 500")
    parser.add_argument("--darvas", action="store_true",
                        help="Run Darvas interpreter on BUY/WATCH stocks (entry/stop/target)")
    args = parser.parse_args()

    root_dir    = Path(__file__).parent
    data_dir    = root_dir / "data"
    reports_dir = root_dir / "reports"
    reports_dir.mkdir(exist_ok=True)

    xl_path = Path(args.output) if args.output \
        else reports_dir / "nifty500_scan_history.xlsx"

    as_of = datetime.now()
    print(f"\n{'='*60}")
    print(f"  NSE FULL SCAN  —  {as_of.strftime('%d %b %Y  %H:%M')}")
    print(f"{'='*60}\n")

    # ── Load NIFTY 500 symbol list ────────────────────────────────────────────
    print("  [1/4] Loading NIFTY 500 symbol list …")
    nifty500_symbols = load_nifty500(data_dir)
    nifty500_set = set(nifty500_symbols)

    darvas_df    = pd.DataFrame()
    piotroski_df = pd.DataFrame()
    index_data   = {}

    if not args.gems_only:
        symbols = nifty500_symbols[: args.top] if args.top else nifty500_symbols
        if args.top:
            print(f"  Limited to first {len(symbols)} symbols (--top {args.top}).")
        else:
            print(f"  Symbols loaded: {len(symbols)}")

        # ── Fetch index levels ────────────────────────────────────────────────
        print("\n  [2/4] Fetching index levels …")
        index_data = fetch_index_data()

        # ── Batch OHLCV download ──────────────────────────────────────────────
        print(f"\n  [3/4] Downloading OHLCV (batch_size={args.batch_size}) …")
        t0 = time.time()
        hist_map = batch_download(symbols, batch_size=args.batch_size)
        print(f"  Downloaded {len(hist_map)}/{len(symbols)} stocks in {time.time()-t0:.0f}s")

        # ── Fetch fundamentals ────────────────────────────────────────────────
        infos: dict[str, dict] = {s: {} for s in symbols}
        if not args.fast and hist_map:
            print(f"\n  [4/4] Fetching fundamentals ({len(hist_map)} stocks) …")
            print("         Use --fast to skip (~5–15 min extra).\n")
            infos = fetch_fundamentals_bulk(list(hist_map.keys()))
        else:
            print("\n  [4/4] Fundamentals skipped (--fast).")

        # ── Build records + scan ──────────────────────────────────────────────
        records = []
        for sym, hist in hist_map.items():
            try:
                records.append(build_record(sym, hist, infos.get(sym, {})))
            except Exception as exc:
                print(f"  [WARN] {sym}: {exc}")

        if not records:
            print("  No records built. Exiting.")
            sys.exit(1)

        df = pd.DataFrame(records)
        print(f"\n  Running scanners on {len(df)} NIFTY 500 stocks …")
        darvas_df    = scan_darvas(df)
        piotroski_df = scan_piotroski(df)

        print_summary(darvas_df, piotroski_df, index_data, as_of)

        # ── Darvas interpretation (always runs when module is available) ─────────
        if _HAS_DARVAS:
            actionable = darvas_df[darvas_df["darvas_signal"].isin(["BUY", "WATCH"])]
            print(f"\n  Running Darvas interpreter on {len(actionable)} BUY/WATCH stocks …")
            darvas_df = _darvas_interpret(darvas_df, hist_map=hist_map)
            print()

        if not args.no_excel:
            print(f"  Writing Excel → {xl_path} …")
            write_excel(darvas_df, piotroski_df, index_data, as_of, xl_path)

        # ── Write standalone Darvas signals Excel ─────────────────────────────
        if args.darvas and _HAS_DARVAS and not args.no_excel:
            from darvas_interpreter import DarvasSignal, write_signals_excel, _build_signal
            dv_cols = [c for c in darvas_df.columns if c.startswith("dv_")]
            if dv_cols:
                # Reconstruct DarvasSignal objects for the Excel writer
                dv_signals = []
                for _, row in darvas_df[
                    darvas_df["darvas_signal"].isin(["BUY","WATCH"])
                ].iterrows():
                    ds = DarvasSignal(
                        ticker        = str(row["ticker"]),
                        name          = str(row.get("name","")),
                        current_price = row.get("cmp"),
                        state         = str(row.get("dv_state","NO_BOX")),
                        action        = str(row.get("dv_action","")),
                        entry_price   = row.get("dv_entry"),
                        stop_loss     = row.get("dv_stop"),
                        target_1      = row.get("dv_target1"),
                        target_2      = row.get("dv_target2"),
                        target_3      = row.get("dv_target3"),
                        risk_pct      = row.get("dv_risk_pct"),
                        reward_1_pct  = row.get("dv_rwd1_pct"),
                        reward_2_pct  = row.get("dv_rwd2_pct"),
                        reward_3_pct  = row.get("dv_rwd3_pct"),
                        box_top       = row.get("dv_box_top"),
                        box_bottom    = row.get("dv_box_bottom"),
                        box_width_pct = row.get("dv_box_w_pct"),
                        box_age_days  = row.get("dv_box_age"),
                        notes         = str(row.get("dv_notes","")),
                    )
                    dv_signals.append(ds)
                dv_xl = reports_dir / "darvas_signals_history.xlsx"
                print(f"  Writing Darvas signals → {dv_xl} …")
                write_signals_excel(dv_signals, dv_xl)
    else:
        print("  --gems-only: skipping NIFTY 500 scan.")
        # Still need a workbook to attach the gems sheet to
        if not args.no_excel:
            from openpyxl import Workbook
            if not xl_path.exists():
                wb = Workbook()
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]
                wb.save(xl_path)

    # ── Hidden Gems scan ──────────────────────────────────────────────────────
    if not args.no_gems:
        gems_darvas, gems_piotroski, gems_hist_map = run_hidden_gems(
            nifty500_set, data_dir, fast=args.fast, batch_size=args.batch_size
        )
        # Run interpreter on gems (zero extra network calls — uses already-downloaded OHLCV)
        if _HAS_DARVAS and not gems_darvas.empty:
            gems_actionable = gems_darvas[gems_darvas["darvas_signal"].isin(["BUY", "WATCH"])]
            print(f"\n  Running Darvas interpreter on {len(gems_actionable)} gem BUY/WATCH stocks …")
            gems_darvas = _darvas_interpret(gems_darvas, hist_map=gems_hist_map)
            print()
        if not args.no_excel and xl_path.exists():
            print(f"\n  Writing Hidden Gems sheet → {xl_path} …")
            write_gems_sheet(gems_darvas, gems_piotroski, as_of, xl_path)
    else:
        print("\n  Hidden Gems scan skipped (--no-gems).")


if __name__ == "__main__":
    main()
