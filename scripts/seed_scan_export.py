#!/usr/bin/env python3
"""
Full pipeline:
  1. Build instrument CSVs for Hong Kong, Canada, China (if missing)
  2. Seed all markets into Cassandra
  3. Fetch quotes (yfinance) for markets that have 0 quotes
  4. Run Darvas/Buffett + Piotroski daily scan across all markets
  5. Export results to Excel with per-market sheets + a summary sheet

Usage:
    python3 scripts/seed_scan_export.py [--skip-fetch] [--out reports/daily_scan.xlsx]
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
import time
import warnings
from datetime import datetime, timezone
from pathlib import Path

warnings.filterwarnings('ignore')

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / 'data'
sys.path.insert(0, str(ROOT / 'backend'))

# ── step 0: connect Cassandra ─────────────────────────────────────────────────

from db import cassandra_client as cass

print("Connecting to Cassandra...")
if not cass.connect():
    sys.exit("Cassandra unavailable — is it running on 127.0.0.1:9042?")
print("  Connected.\n")


# ── step 1: build missing instrument CSVs ────────────────────────────────────

def _write_csv(path: Path, fieldnames: list, rows: list) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    return len(rows)


def build_hk_list():
    path = DATA / 'hk_list.csv'
    rows = []
    try:
        import FinanceDataReader as fdr
        df = fdr.StockListing('HKEX')
        for _, r in df.iterrows():
            code_raw = str(r.get('Code', r.get('Symbol', ''))).strip()
            name = str(r.get('Name', r.get('CompanyName', ''))).strip()
            if not code_raw or not name:
                continue
            try:
                code = str(int(float(code_raw))).zfill(4)
            except Exception:
                code = code_raw
            if code.isdigit():
                rows.append({'symbol': code, 'name': name, 'exchange': 'HKEX',
                             'yf_ticker': code + '.HK'})
    except Exception as e:
        print(f"    FDR failed ({e}), using Hang Seng 50 fallback")

    if not rows:
        hs50 = [
            ('0700', 'Tencent Holdings'), ('0005', 'HSBC Holdings'),
            ('0939', 'China Construction Bank'), ('1398', 'ICBC'),
            ('3988', 'Bank of China'), ('2318', 'Ping An Insurance'),
            ('0941', 'China Mobile'), ('0883', 'CNOOC'),
            ('0388', 'Hong Kong Exchanges'), ('1299', 'AIA Group'),
            ('0016', 'Sun Hung Kai Properties'), ('0001', 'CK Hutchison'),
            ('2628', 'China Life Insurance'), ('0823', 'Link REIT'),
            ('0002', 'CLP Holdings'), ('0003', 'HK & China Gas'),
            ('0006', 'Power Assets'), ('0011', 'Hang Seng Bank'),
            ('0012', 'Henderson Land'), ('0017', 'New World Development'),
            ('0019', 'Swire Pacific'), ('0027', 'Galaxy Entertainment'),
            ('0066', 'MTR Corporation'), ('0101', 'Hang Lung Properties'),
            ('0175', 'Geely Automobile'), ('0267', 'CITIC'),
            ('0291', 'China Resources Beer'), ('0316', 'Orient Overseas'),
            ('0386', 'Sinopec'), ('0669', 'Techtronic Industries'),
            ('0688', 'China Overseas Land'), ('0762', 'China Unicom'),
            ('0857', 'PetroChina'), ('0868', 'Xinyi Glass'),
            ('0960', 'Longfor Group'), ('1038', 'CK Infrastructure'),
            ('1044', 'Hengan International'), ('1093', 'CSPC Pharmaceutical'),
            ('1109', 'China Resources Land'), ('1177', 'Sino Biopharmaceutical'),
            ('1211', 'BYD Company'), ('1997', 'Wharf Real Estate'),
            ('2020', 'ANTA Sports'), ('2269', 'WuXi Biologics'),
            ('2313', 'Shenzhou International'), ('2382', 'Sunny Optical'),
            ('3690', 'Meituan'), ('6098', 'Country Garden Services'),
            ('9988', 'Alibaba Group'), ('9618', 'JD.com'),
        ]
        rows = [{'symbol': c, 'name': n, 'exchange': 'HKEX', 'yf_ticker': c + '.HK'}
                for c, n in hs50]

    n = _write_csv(path, ['symbol', 'name', 'exchange', 'yf_ticker'], rows)
    print(f"  Hong Kong list: {n} stocks → {path.name}")


def build_canada_list():
    path = DATA / 'canada_list.csv'
    rows = []
    try:
        import FinanceDataReader as fdr
        df = fdr.StockListing('TSX')
        for _, r in df.iterrows():
            sym = str(r.get('Symbol', r.get('Code', ''))).strip()
            name = str(r.get('Name', r.get('CompanyName', ''))).strip()
            if sym and name and not sym.endswith('.TO'):
                rows.append({'symbol': sym, 'name': name, 'exchange': 'TSX',
                             'yf_ticker': sym + '.TO'})
    except Exception as e:
        print(f"    FDR failed ({e}), using TSX 60 fallback")

    if not rows:
        tsx60 = [
            ('SHOP', 'Shopify'), ('CNR', 'Canadian National Railway'),
            ('TD', 'Toronto-Dominion Bank'), ('ENB', 'Enbridge'),
            ('RY', 'Royal Bank of Canada'), ('CP', 'Canadian Pacific Kansas City'),
            ('BN', 'Brookfield Corp'), ('BAM', 'Brookfield Asset Management'),
            ('BCE', 'BCE Inc'), ('SU', 'Suncor Energy'),
            ('TRI', 'Thomson Reuters'), ('ATD', 'Alimentation Couche-Tard'),
            ('ABX', 'Barrick Gold'), ('AEM', 'Agnico Eagle Mines'),
            ('BNS', 'Bank of Nova Scotia'), ('BMO', 'Bank of Montreal'),
            ('MFC', 'Manulife Financial'), ('SLF', 'Sun Life Financial'),
            ('TRP', 'TC Energy'), ('CNQ', 'Canadian Natural Resources'),
            ('CCO', 'Cameco'), ('CSU', 'Constellation Software'),
            ('DOL', 'Dollarama'), ('EMA', 'Emera'), ('FTS', 'Fortis'),
            ('GWO', 'Great-West Lifeco'), ('H', 'Hydro One'),
            ('IFC', 'Intact Financial'), ('IMO', 'Imperial Oil'),
            ('K', 'Kinross Gold'), ('L', 'Loblaw Companies'),
            ('MRU', 'Metro Inc'), ('NA', 'National Bank of Canada'),
            ('NTR', 'Nutrien'), ('OVV', 'Ovintiv'),
            ('POW', 'Power Corporation of Canada'),
            ('PPL', 'Pembina Pipeline'),
            ('QSR', 'Restaurant Brands International'),
            ('RCI.B', 'Rogers Communications'), ('SAP', 'Saputo'),
            ('T', 'TELUS'), ('TOU', 'Tourmaline Oil'),
            ('WCN', 'Waste Connections'), ('WFG', 'West Fraser Timber'),
            ('WN', 'George Weston'), ('WPM', 'Wheaton Precious Metals'),
            ('X', 'TMX Group'), ('AQN', 'Algonquin Power'),
            ('CAR.UN', 'Canadian Apartment REIT'), ('CTC.A', 'Canadian Tire'),
            ('EQB', 'EQB Inc'), ('FM', 'First Quantum Minerals'),
            ('GFL', 'GFL Environmental'), ('PKI', 'Parkland Corp'),
            ('GIB.A', 'CGI Inc'), ('ALA', 'AltaGas'),
            ('ACO.X', 'ATCO'), ('CCA', 'Cogeco Communications'),
        ]
        rows = [{'symbol': s, 'name': n, 'exchange': 'TSX', 'yf_ticker': s + '.TO'}
                for s, n in tsx60]

    n = _write_csv(path, ['symbol', 'name', 'exchange', 'yf_ticker'], rows)
    print(f"  Canada list: {n} stocks → {path.name}")


def build_china_list():
    path = DATA / 'china_list.csv'
    rows = []
    try:
        import akshare as ak
        df = ak.stock_info_a_code_name()
        for _, r in df.iterrows():
            code = str(r['code']).strip().zfill(6)
            name = str(r['name']).strip()
            suffix = '.SS' if code.startswith('6') else '.SZ'
            rows.append({'symbol': code, 'name': name,
                         'exchange': 'SSE' if suffix == '.SS' else 'SZSE',
                         'yf_ticker': code + suffix})
    except Exception as e:
        print(f"    akshare failed ({e}); China list empty — skipping China")
        return

    n = _write_csv(path, ['symbol', 'name', 'exchange', 'yf_ticker'], rows)
    print(f"  China list: {n} stocks → {path.name}")


print("=== Step 1: Building instrument CSVs ===")
if not (DATA / 'hk_list.csv').exists():
    build_hk_list()
else:
    print(f"  Hong Kong list: already exists")

if not (DATA / 'canada_list.csv').exists():
    build_canada_list()
else:
    print(f"  Canada list: already exists")

if not (DATA / 'china_list.csv').exists():
    build_china_list()
else:
    print(f"  China list: already exists")
print()


# ── step 2: seed all markets ──────────────────────────────────────────────────

print("=== Step 2: Seeding markets ===")
# Clear market_db cache so new CSVs are picked up
import parsers.market_db as mdb
mdb._CACHE.clear()

from db.seeder import seed_market, MARKETS

for market in MARKETS:
    result = seed_market(market, force=False)
    if result.get('skipped'):
        s = cass.session()
        r = s.execute(f'SELECT COUNT(*) FROM {cass.KEYSPACE}.instruments WHERE market = %s',
                      (market,)).one()
        print(f"  {market:12s}  already seeded ({int(r[0]):,} instruments)")
    elif result.get('error'):
        print(f"  {market:12s}  ERROR: {result['error']}")
    else:
        print(f"  {market:12s}  seeded {result['inserted']:,} instruments")
print()


# ── step 3: fetch quotes for markets with 0 quotes ───────────────────────────

parser = argparse.ArgumentParser(add_help=False)
parser.add_argument('--skip-fetch', action='store_true')
parser.add_argument('--out', default=str(ROOT / 'reports' / 'daily_scan.xlsx'))
args, _ = parser.parse_known_args()

print("=== Step 3: Quote status ===")
s_cass = cass.session()
needs_fetch = []
has_quotes = []
for market in MARKETS:
    q = s_cass.execute(
        f'SELECT COUNT(*) FROM {cass.KEYSPACE}.stock_quotes WHERE market = %s', (market,)
    ).one()
    count = int(q[0])
    instr = s_cass.execute(
        f'SELECT COUNT(*) FROM {cass.KEYSPACE}.instruments WHERE market = %s', (market,)
    ).one()
    instr_count = int(instr[0])
    if count == 0 and instr_count > 0:
        needs_fetch.append(market)
        print(f"  {market:12s}  quotes=0 / {instr_count:,} instruments  → will fetch")
    elif instr_count == 0:
        print(f"  {market:12s}  no instruments — skipping")
    else:
        has_quotes.append(market)
        print(f"  {market:12s}  quotes={count:,} ✓")
print()

if needs_fetch and not args.skip_fetch:
    print(f"=== Fetching quotes for: {', '.join(needs_fetch)} ===")
    from db.bulk_fetcher import fetch_market_quotes
    for market in needs_fetch:
        print(f"  Fetching {market}…", flush=True)
        t0 = time.time()
        fetch_market_quotes(market, batch_size=50, max_workers=4, with_fundamentals=False)
        q = s_cass.execute(
            f'SELECT COUNT(*) FROM {cass.KEYSPACE}.stock_quotes WHERE market = %s', (market,)
        ).one()
        elapsed = time.time() - t0
        print(f"  {market}: {int(q[0]):,} quotes written in {elapsed:.0f}s")
        has_quotes.append(market)
    print()
elif needs_fetch:
    print(f"  --skip-fetch set; skipping fetch for: {', '.join(needs_fetch)}\n")


# ── step 4: run daily scan ────────────────────────────────────────────────────

print("=== Step 4: Running daily scan ===")
from db.quote_updater import get_market_quotes_df
from scanners.daily_scanner import scan_darvas, scan_piotroski

CURRENCY = {
    'india': '₹', 'us': '$', 'europe': '€', 'japan': '¥',
    'korea': '₩', 'china': '¥', 'hong_kong': 'HK$', 'canada': 'C$',
}
LABEL = {
    'india': 'India', 'us': 'US', 'europe': 'Europe', 'japan': 'Japan',
    'korea': 'Korea', 'china': 'China', 'hong_kong': 'Hong Kong', 'canada': 'Canada',
}

all_darvas: list[dict] = []
all_piotroski: list[dict] = []

scan_markets = [m for m in MARKETS if m in has_quotes or m in needs_fetch]
for market in scan_markets:
    df = get_market_quotes_df(market)
    if df.empty:
        print(f"  {market:12s}  no quote data — skipped")
        continue
    cur = CURRENCY.get(market, '')
    lbl = LABEL.get(market, market)
    d_rows = scan_darvas(df)
    p_rows = scan_piotroski(df)
    for row in d_rows:
        if row.get('signal') in ('BUY', 'WATCH'):
            row.update({'market': market, 'market_label': lbl, 'currency': cur})
            all_darvas.append(row)
    for row in p_rows:
        if row.get('signal') in ('BUY', 'WATCH'):
            row.update({'market': market, 'market_label': lbl, 'currency': cur})
            all_piotroski.append(row)
    print(f"  {market:12s}  {len(df):,} stocks  →  "
          f"Darvas BUY={sum(1 for r in d_rows if r.get('signal')=='BUY')} "
          f"WATCH={sum(1 for r in d_rows if r.get('signal')=='WATCH')}  |  "
          f"Piotroski BUY={sum(1 for r in p_rows if r.get('signal')=='BUY')} "
          f"WATCH={sum(1 for r in p_rows if r.get('signal')=='WATCH')}")

all_darvas.sort(key=lambda r: r.get('score', 0) or 0, reverse=True)
all_piotroski.sort(key=lambda r: r.get('score', 0) or 0, reverse=True)
print(f"\n  Total Darvas signals: {len(all_darvas)}")
print(f"  Total Piotroski signals: {len(all_piotroski)}\n")


# ── step 5: export to Excel ───────────────────────────────────────────────────

print(f"=== Step 5: Exporting to Excel ===")
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUT = Path(args.out)
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# colour palette
C_HEADER_BG = '1e293b'
C_HEADER_FG = 'e2e8f0'
C_BUY_BG    = '052e16'
C_BUY_FG    = '4ade80'
C_WATCH_BG  = '431407'
C_WATCH_FG  = 'fb923c'
C_ALT_BG    = '0f172a'
C_BASE_BG   = '020617'
C_BORDER    = '334155'

def _fill(hex_: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_)

def _font(hex_: str, bold=False, sz=10) -> Font:
    return Font(color=hex_, bold=bold, size=sz, name='Calibri')

def _border() -> Border:
    s = Side(style='thin', color=C_BORDER)
    return Border(bottom=s)

def _fmt_number(v, decimals=2):
    if v is None:
        return '—'
    try:
        return round(float(v), decimals)
    except Exception:
        return '—'

def _pct(v):
    if v is None:
        return '—'
    try:
        return round(float(v), 2)
    except Exception:
        return '—'


DARVAS_COLS = [
    ('Market',      'market_label',   14),
    ('Ticker',      'ticker',         12),
    ('Company',     'name',           30),
    ('Signal',      'signal',         8),
    ('Score',       'score',          7),
    ('Max Score',   'max_score',      9),
    ('Price',       'cmp',            10),
    ('RSI',         'rsi',            7),
    ('RSI Signal',  'rsi_signal',     10),
    ('EMA-50',      'ema_50',         10),
    ('EMA-200',     'ema_200',        10),
    ('52W High',    'high_52w',       10),
    ('52W Low',     'low_52w',        10),
    ('1D %',        'ret_1d',         8),
    ('1W %',        'ret_1w',         8),
    ('1M %',        'ret_1m',         8),
    ('3M %',        'ret_3m',         8),
    ('6M %',        'ret_6m',         8),
    ('1Y %',        'ret_1y',         8),
    ('P/E',         'pe',             8),
    ('P/B',         'pb',             8),
    ('ROE %',       'roe',            8),
    ('D/E',         'debt_to_equity', 8),
    ('OPM %',       'opm',            8),
    ('Mkt Cap',     'market_cap',     12),
    ('Volume',      'volume',         12),
    ('Vol Ratio',   'volume_ratio',   10),
    ('Beta',        'beta',           7),
    ('Curr Ratio',  'current_ratio',  10),
    ('MACD',        'macd',           10),
    ('Sector',      'sector',         20),
    ('Industry',    'industry',       25),
    ('Exchange',    'exchange',       14),
    ('Near 52W Hi', 'near_52w_high',  12),
    ('Above EMA50', 'above_ema50',    12),
    ('Currency',    'currency',       8),
]

PIOTROSKI_COLS = DARVAS_COLS[:]  # same columns, different data


def write_sheet(wb, title: str, rows: list[dict], cols: list[tuple]) -> None:
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '4f46e5'

    # freeze header
    ws.freeze_panes = 'A2'

    # header row
    for ci, (label, _, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font      = _font(C_HEADER_FG, bold=True, sz=10)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 20

    # data rows
    for ri, row in enumerate(rows, 2):
        sig = row.get('signal', '')
        if sig == 'BUY':
            row_bg, name_fg = C_BUY_BG, C_BUY_FG
        else:
            row_bg, name_fg = C_WATCH_BG, C_WATCH_FG
        alt_bg = C_ALT_BG if ri % 2 == 0 else C_BASE_BG

        for ci, (_, key, _) in enumerate(cols, 1):
            raw = row.get(key)

            # format value
            if key in ('ret_1d', 'ret_1w', 'ret_1m', 'ret_3m', 'ret_6m', 'ret_1y',
                        'roe', 'opm'):
                val = _pct(raw)
            elif key in ('cmp', 'ema_50', 'ema_200', 'high_52w', 'low_52w',
                          'pe', 'pb', 'debt_to_equity', 'volume_ratio',
                          'beta', 'current_ratio', 'macd', 'rsi'):
                val = _fmt_number(raw)
            elif key in ('market_cap', 'volume'):
                val = _fmt_number(raw, 0) if raw is not None and raw != '—' else '—'
            elif key in ('near_52w_high', 'above_ema50'):
                val = 'Y' if raw is True else ('N' if raw is False else '—')
            elif key == 'score':
                try:
                    val = int(raw) if raw is not None else '—'
                except Exception:
                    val = '—'
            else:
                val = raw if raw is not None else '—'

            cell = ws.cell(row=ri, column=ci, value=val)

            # colour
            if key == 'name':
                cell.font = _font(name_fg, bold=True)
                cell.fill = _fill(row_bg)
            elif key == 'signal':
                cell.font = _font(name_fg, bold=True)
                cell.fill = _fill(row_bg)
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.font = _font('94a3b8')
                cell.fill = _fill(alt_bg)

            # right-align numbers
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal='right')

            cell.border = _border()

    # auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def write_summary(wb, darvas: list, piotroski: list) -> None:
    ws = wb.create_sheet(title='Summary', index=0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '059669'

    scanned_at = datetime.now().strftime('%Y-%m-%d %H:%M')

    # Title block
    ws.merge_cells('A1:H1')
    t = ws['A1']
    t.value = f'Daily Scan Report — {scanned_at}'
    t.font  = Font(color='e2e8f0', bold=True, size=14, name='Calibri')
    t.fill  = _fill('0f172a')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:H2')
    sub = ws['A2']
    sub.value = 'Darvas/Buffett & Piotroski · All Markets · BUY & WATCH signals'
    sub.font  = Font(color='64748b', size=9, name='Calibri')
    sub.fill  = _fill('0f172a')
    sub.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    # Per-scan summary headers
    headers = ['Scan', 'Market', 'BUY', 'WATCH', 'Total']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font  = _font(C_HEADER_FG, bold=True)
        c.fill  = _fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal='center')
        c.border = _border()

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8

    ri = 5
    for scan_name, rows in [('Darvas/Buffett', darvas), ('Piotroski', piotroski)]:
        markets_seen = sorted({r['market'] for r in rows})
        for mi, market in enumerate(markets_seen):
            mrows = [r for r in rows if r['market'] == market]
            buy   = sum(1 for r in mrows if r.get('signal') == 'BUY')
            watch = sum(1 for r in mrows if r.get('signal') == 'WATCH')
            bg = C_ALT_BG if ri % 2 == 0 else C_BASE_BG
            vals = [scan_name if mi == 0 else '', LABEL.get(market, market),
                    buy, watch, len(mrows)]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font   = _font('94a3b8')
                c.fill   = _fill(bg)
                c.border = _border()
                if isinstance(v, int):
                    c.alignment = Alignment(horizontal='right')
            ri += 1

        # totals row
        total_buy   = sum(1 for r in rows if r.get('signal') == 'BUY')
        total_watch = sum(1 for r in rows if r.get('signal') == 'WATCH')
        for ci, v in enumerate([f'{scan_name} TOTAL', '', total_buy,
                                  total_watch, len(rows)], 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font   = _font('e2e8f0', bold=True)
            c.fill   = _fill('1e3a5f')
            c.border = _border()
            if isinstance(v, int):
                c.alignment = Alignment(horizontal='right')
        ri += 2


# Build workbook
write_summary(wb, all_darvas, all_piotroski)

write_sheet(wb, 'Darvas-Buffett (All)', all_darvas, DARVAS_COLS)
write_sheet(wb, 'Piotroski (All)',      all_piotroski, DARVAS_COLS)

# Per-market sheets for Darvas
for market in scan_markets:
    mrows = [r for r in all_darvas if r.get('market') == market]
    if mrows:
        lbl = LABEL.get(market, market)
        write_sheet(wb, f'D-{lbl}', mrows, DARVAS_COLS)

# Per-market sheets for Piotroski
for market in scan_markets:
    mrows = [r for r in all_piotroski if r.get('market') == market]
    if mrows:
        lbl = LABEL.get(market, market)
        write_sheet(wb, f'P-{lbl}', mrows, DARVAS_COLS)

wb.save(str(OUT))
print(f"  Saved → {OUT}")
print(f"\nDone. {len(all_darvas)} Darvas + {len(all_piotroski)} Piotroski signals exported.")
